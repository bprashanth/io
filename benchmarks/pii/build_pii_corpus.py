#!/usr/bin/env python3
"""Deterministic synthetic NGO-style PII corpus generator.

Everything produced here is fabricated (Verhoeff-valid but synthetic Aadhaar
numbers, made-up names/villages/identifiers). Nothing is a real person's
data. Ground truth (column classes and character-offset spans) is generated
by construction alongside the content, never recovered by regex afterwards,
and every offset is re-verified against the written files before this script
exits successfully.
"""

import json
import os
import random
import string
import uuid as uuidlib
from datetime import date, timedelta

import pandas as pd
import openpyxl

SEED = 20260821
random.seed(SEED)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(SCRIPT_DIR, "corpus")
os.makedirs(OUT_DIR, exist_ok=True)

REF_DATE = date(2026, 8, 21)
SUMMARY = []

# --------------------------------------------------------------------------
# Verhoeff checksum, used to make synthetic Aadhaar numbers structurally valid
# --------------------------------------------------------------------------
_D = (
    (0, 1, 2, 3, 4, 5, 6, 7, 8, 9), (1, 2, 3, 4, 0, 6, 7, 8, 9, 5),
    (2, 3, 4, 0, 1, 7, 8, 9, 5, 6), (3, 4, 0, 1, 2, 8, 9, 5, 6, 7),
    (4, 0, 1, 2, 3, 9, 5, 6, 7, 8), (5, 9, 8, 7, 6, 0, 4, 3, 2, 1),
    (6, 5, 9, 8, 7, 1, 0, 4, 3, 2), (7, 6, 5, 9, 8, 2, 1, 0, 4, 3),
    (8, 7, 6, 5, 9, 3, 2, 1, 0, 4), (9, 8, 7, 6, 5, 4, 3, 2, 1, 0),
)
_P = (
    (0, 1, 2, 3, 4, 5, 6, 7, 8, 9), (1, 5, 7, 6, 2, 8, 3, 0, 9, 4),
    (5, 8, 0, 3, 7, 9, 6, 1, 4, 2), (8, 9, 1, 6, 0, 4, 3, 5, 2, 7),
    (9, 4, 5, 3, 1, 2, 6, 8, 7, 0), (4, 2, 8, 6, 5, 7, 3, 9, 0, 1),
    (2, 7, 9, 3, 8, 0, 6, 4, 1, 5), (7, 0, 4, 6, 9, 1, 3, 2, 5, 8),
)
_INV = (0, 4, 3, 2, 1, 5, 6, 7, 8, 9)


def verhoeff_check_digit(num_str):
    c = 0
    for i, ch in enumerate(reversed(num_str)):
        c = _D[c][_P[(i + 1) % 8][int(ch)]]
    return _INV[c]


def verhoeff_validate(num_str):
    c = 0
    for i, ch in enumerate(reversed(num_str)):
        c = _D[c][_P[i % 8][int(ch)]]
    return c == 0


def gen_aadhaar():
    base = str(random.randint(2, 9)) + "".join(str(random.randint(0, 9)) for _ in range(10))
    digits = base + str(verhoeff_check_digit(base))
    assert verhoeff_validate(digits)
    if random.random() < 0.5:
        return digits
    return f"{digits[0:4]} {digits[4:8]} {digits[8:12]}"


# --------------------------------------------------------------------------
# Name / place / identifier pools
# --------------------------------------------------------------------------
FIRST_MALE = [
    "Ramesh", "Suresh", "Mahesh", "Dinesh", "Ganesh", "Prakash", "Vinod", "Anil", "Sunil", "Ravi",
    "Ashok", "Vijay", "Sanjay", "Rajesh", "Mukesh", "Deepak", "Amit", "Rohit", "Vikram", "Arjun",
    "Rahul", "Rakesh", "Manoj", "Santosh", "Sachin", "Nitin", "Milind", "Shrikant", "Balasaheb", "Vasant",
    "Baban", "Tukaram", "Namdev", "Dattatray", "Popat", "Eknath", "Bhaskar", "Shivaji", "Yashwant", "Bapu",
    "Mohan", "Gopal", "Krishna", "Hari", "Raju", "Ramu", "Ismail", "Ayub", "Salman", "Irfan",
    "Rizwan", "Farooq", "Aslam", "Iqbal", "Zakir", "Nasir", "Gurmeet", "Harpreet", "Jaspreet", "Amarjeet",
    "Baljeet", "Manpreet", "Joseph", "Thomas", "Stephen", "Xavier", "Anthony", "Francis", "Pradeep", "Umesh",
    "Ramdas", "Baliram", "Shankar", "Ganpat", "Dagdu", "Vishal", "Pankaj", "Aakash", "Kiran", "Naveen",
]
FIRST_FEMALE = [
    "Sunita", "Anita", "Kavita", "Savita", "Rekha", "Meena", "Geeta", "Sita", "Radha", "Lata",
    "Usha", "Nirmala", "Shobha", "Vandana", "Pratibha", "Sushila", "Kamala", "Vimla", "Shakuntala", "Yashoda",
    "Mangala", "Sarita", "Manisha", "Rupali", "Snehal", "Priyanka", "Pooja", "Neha", "Kavya", "Divya",
    "Swati", "Bhagyashree", "Archana", "Suvarna", "Jyoti", "Kalpana", "Vaishali", "Ashwini", "Sushma", "Sarla",
    "Parvati", "Lakshmi", "Saraswati", "Durga", "Fatima", "Ayesha", "Zainab", "Sakina", "Rukhsana", "Shabana",
    "Nasreen", "Rehana", "Amina", "Harjeet", "Simran", "Navjot", "Rajwinder", "Gurleen", "Mary", "Sheela",
    "Rosy", "Grace", "Lily", "Sudha", "Padma", "Lalita", "Kumud", "Indira", "Champa", "Sarojini", "Tara", "Deepa",
]
SURNAMES = [
    "Kulkarni", "Deshmukh", "Patil", "Joshi", "Pawar", "More", "Shinde", "Jadhav", "Gaikwad", "Bhosale",
    "Kale", "Chavan", "Sawant", "Deshpande", "Kadam", "Sharma", "Verma", "Yadav", "Singh", "Kumar",
    "Mishra", "Tiwari", "Gupta", "Chaudhary", "Rai", "Prasad", "Thakur", "Pandey", "Dubey", "Reddy",
    "Naidu", "Iyer", "Iyengar", "Nair", "Menon", "Pillai", "Rao", "Chettiar", "Krishnan", "Khan",
    "Sheikh", "Ansari", "Qureshi", "Sayyed", "Pathan", "Mirza", "Siddiqui", "Gill", "Sidhu", "Bajwa",
    "Dhillon", "Kaur", "DSouza", "Fernandes", "Pereira", "Lobo", "Rodrigues", "Fernandez", "Thomas", "George",
    "Mathew", "Waghmare", "Salunkhe", "Zende", "Bhagat", "Chougule", "Nikam", "Mane", "Bhoir", "Gadhave",
    "Bansal", "Agarwal", "Chandra", "Mahato", "Kumari", "Ranjan", "Kushwaha", "Paswan", "Manjhi", "Yadav",
]

AREAS = [
    {"state": "Maharashtra", "district": "Pune", "talukas": ["Haveli", "Mulshi", "Junnar", "Khed", "Baramati"],
     "villages": ["Wadgaon", "Kharadi", "Loni Kalbhor", "Chikhali", "Manjari", "Uruli Kanchan", "Shirur",
                  "Talegaon Dhamdhere", "Ambegaon Budruk", "Rajgurunagar", "Saswad", "Nira", "Supa", "Bhor",
                  "Alandi", "Dehu", "Nanded Phata", "Kadus", "Chakan", "Kesnand"],
     "bbox": (18.30, 18.90, 73.30, 74.30)},
    {"state": "Bihar", "district": "Gaya", "talukas": ["Belaganj", "Sherghati", "Wazirganj", "Tekari", "Konch"],
     "villages": ["Rampur", "Bishunpur", "Chandpur", "Devipur", "Bakhtiyarpur Tola", "Mahuar", "Amawan",
                  "Karpi", "Barachatti", "Imamganj", "Dobhi", "Fatehpur", "Atri", "Guraru", "Paraiya"],
     "bbox": (24.30, 25.20, 84.50, 85.60)},
    {"state": "Bihar", "district": "Muzaffarpur", "talukas": ["Kanti", "Kurhani", "Sakra", "Bochaha", "Musahri"],
     "villages": ["Harnaut Tola", "Chandwara", "Rampur Sokho", "Deopur", "Bariarpur", "Basaith",
                  "Kalyanpur", "Motipur", "Paroo", "Sahebganj"],
     "bbox": (25.90, 26.40, 85.00, 85.70)},
]

SCHOOLS = ["Zilla Parishad Primary School", "New English School", "Vidya Mandir High School",
           "Saraswati Vidyalaya", "Government Middle School", "Adarsh Vidyalaya", "Municipal School",
           "Rashtriya Balika Vidyalaya", "Lokmanya Tilak Vidyalaya", "Sant Gadge Baba School"]
SCHEMES = ["Pre-Matric Scholarship Scheme", "Post-Matric Scholarship Scheme", "Savitribai Phule Scholarship",
           "Rajarshi Shahu Maharaj Scholarship", "Minority Welfare Scholarship", "Merit-cum-Means Scholarship",
           "EBC Scholarship Scheme"]
CITIES = ["Pune", "Mumbai", "Nagpur", "Nashik", "Aurangabad", "Kolhapur", "Patna", "Gaya", "Muzaffarpur",
          "Delhi", "Bengaluru", "Thane", "Solapur"]
CASTES = ["SC", "ST", "OBC", "General", "NT", "VJNT", "OBC-NT"]
BANK_CODES = ["SBIN", "HDFC", "ICIC", "UTIB", "PUNB", "BARB", "CNRB", "UBIN", "IOBA", "MAHB"]
PSPS = ["okaxis", "okhdfcbank", "oksbi", "ybl", "paytm", "ibl", "axl"]
STATUSES = ["Approved", "Pending", "Rejected", "Under Review", "Disbursed"]
MODES = ["Cash", "Cheque", "UPI", "NEFT", "Card", "DD"]


def rand_name(gender=None):
    if gender is None:
        gender = random.choice(["M", "F"])
    first = random.choice(FIRST_MALE if gender == "M" else FIRST_FEMALE)
    surname = random.choice(SURNAMES)
    style = random.choices(["fs", "sf", "init", "lower", "upper", "first"],
                            weights=[52, 15, 10, 8, 8, 7])[0]
    if style == "fs":
        name = f"{first} {surname}"
    elif style == "sf":
        name = f"{surname} {first}"
    elif style == "init":
        name = f"{random.choice(string.ascii_uppercase)}. {random.choice(string.ascii_uppercase)}. {surname}"
    elif style == "lower":
        name = f"{first} {surname}".lower()
    elif style == "upper":
        name = f"{first} {surname}".upper()
    else:
        name = first
    return {"name": name, "gender": gender, "first": first, "surname": surname}


def gen_phone():
    d = str(random.randint(6, 9)) + "".join(str(random.randint(0, 9)) for _ in range(9))
    fmt = random.choice(["plain", "intl", "dashed"])
    if fmt == "plain":
        return d
    if fmt == "intl":
        return f"+91 {d[:5]} {d[5:]}"
    return f"0{d[:5]}-{d[5:]}"


def gen_email(first, surname):
    f = first.lower()
    s = surname.lower().replace("'", "")
    style = random.choice([0, 1, 2])
    if style == 0:
        local = f"{f}.{s}{random.randint(1, 99)}"
    elif style == 1:
        local = f"{f}{s[:3]}{random.randint(10, 999)}"
    else:
        local = f"{f}_{s}"
    domain = random.choice(["gmail.com", "yahoo.in", "rediffmail.com", "outlook.com", "gmail.com"])
    return f"{local}@{domain}"


def gen_pan():
    letters = [random.choice(string.ascii_uppercase) for _ in range(5)]
    letters[3] = "P"
    digits = "".join(str(random.randint(0, 9)) for _ in range(4))
    last = random.choice(string.ascii_uppercase)
    return "".join(letters) + digits + last


def gen_bank_account():
    length = random.randint(11, 16)
    return str(random.randint(1, 9)) + "".join(str(random.randint(0, 9)) for _ in range(length - 1))


def gen_ifsc():
    code = random.choice(BANK_CODES)
    rest = "".join(random.choice(string.ascii_uppercase + string.digits) for _ in range(6))
    return f"{code}0{rest}"


def gen_upi(first, surname):
    local = f"{first.lower()}{surname.lower()[:3]}{random.randint(10, 999)}"
    return f"{local}@{random.choice(PSPS)}"


def gen_dob_age(min_age, max_age):
    age = random.randint(min_age, max_age)
    start = date(REF_DATE.year - age - 1, REF_DATE.month, min(REF_DATE.day, 28))
    end = date(REF_DATE.year - age, REF_DATE.month, min(REF_DATE.day, 28))
    span = max((end - start).days, 1)
    return start + timedelta(days=random.randint(0, span)), age


def gen_ration_card(state):
    code = "MH" if state == "Maharashtra" else "BR"
    return f"{code}{random.randint(10, 33):02d}{random.randint(1000000, 9999999)}"


def gen_voter_id():
    letters = "".join(random.choice(string.ascii_uppercase) for _ in range(3))
    digits = "".join(str(random.randint(0, 9)) for _ in range(7))
    return letters + digits


def gen_vehicle():
    state = random.choice(["MH", "BR"])
    rto = random.randint(1, 50)
    letters = "".join(random.choice(string.ascii_uppercase) for _ in range(2))
    digits = random.randint(1000, 9999)
    return f"{state}{rto:02d}{letters}{digits}"


def gen_address(village, taluka, district, state):
    house = random.randint(1, 450)
    return f"H.No. {house}, {village}, {taluka} Taluka, {district}, {state}"


def gen_gps(bbox):
    lat = round(random.uniform(bbox[0], bbox[1]), 4)
    lon = round(random.uniform(bbox[2], bbox[3]), 4)
    return lat, lon


def compose(parts):
    """Concatenate strings and (text, class) tuples; return (full_text, spans)."""
    s = ""
    spans = []
    for p in parts:
        if isinstance(p, tuple):
            text, cls = p
            start = len(s)
            s += text
            spans.append({"start": start, "end": len(s), "class": cls, "text": text})
        else:
            s += p
    return s, spans


def write_json(path, obj):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, ensure_ascii=False)


def record_summary(name, rows, col_classes, spans):
    pii_cols = 0
    for v in col_classes.values():
        classes = v if isinstance(v, list) else [v]
        if any(c not in ("none", "record_id_non_pii") for c in classes):
            pii_cols += 1
    SUMMARY.append({"file": name, "rows": rows, "pii_columns": pii_cols, "spans": len(spans)})


def verify_tabular_csv(path, cells):
    df = pd.read_csv(path, dtype=str, keep_default_na=False, na_filter=False)
    for c in cells:
        cell_text = df.iloc[c["row"]][c["column"]]
        assert cell_text[c["start"]:c["end"]] == c["text"], f"span mismatch in {path} row {c['row']} col {c['column']}: {c}"
    print(f"verified {len(cells)} spans in {os.path.basename(path)}")


def verify_tabular_xlsx(path, cells):
    wb = openpyxl.load_workbook(path, data_only=True)
    header_cache = {}
    for c in cells:
        ws = wb[c["sheet"]]
        if c["sheet"] not in header_cache:
            headers = [cell.value for cell in ws[1]]
            header_cache[c["sheet"]] = {h: i + 1 for i, h in enumerate(headers)}
        col_idx = header_cache[c["sheet"]][c["column"]]
        val = ws.cell(row=c["row"] + 2, column=col_idx).value
        val = "" if val is None else str(val)
        assert val[c["start"]:c["end"]] == c["text"], f"span mismatch in {path} sheet {c['sheet']} row {c['row']} col {c['column']}: {c}"
    print(f"verified {len(cells)} spans in {os.path.basename(path)}")


def verify_txt(path, spans, expected_content):
    with open(path, encoding="utf-8") as f:
        content = f.read()
    assert content == expected_content, f"content mismatch on reread of {path}"
    for s in spans:
        assert content[s["start"]:s["end"]] == s["text"], f"span mismatch in {path}: {s}"
    print(f"verified {len(spans)} spans in {os.path.basename(path)}")


# --------------------------------------------------------------------------
# File 1: scholarship_applicants.csv
# --------------------------------------------------------------------------
REMARKS_PLAIN = [
    "Documents pending verification.", "Marksheet not attested by school.",
    "Waiting for bank passbook copy.", "Case reviewed by committee, approved.",
    "Photo missing from file, requested again.", "Income certificate expired, renewal asked.",
    "Application complete, forwarded to district office.", "Caste certificate under verification.",
    "Bank account details mismatch, resubmission needed.", "School has confirmed attendance record.",
    "Awaiting scholarship disbursal from state.", "Duplicate application, merged with earlier record.",
]


def remarks_with_pii():
    relative = rand_name()
    phone = gen_phone()
    kind = random.choice(["mother_call", "follow_up", "submit", "shifted", "guardian"])
    if kind == "mother_call":
        parts = ["Spoke to ", (relative["name"], "person_name"), "'s mother, call ", (phone, "phone"), " for update."]
    elif kind == "follow_up":
        parts = ["Follow up with ", (relative["name"], "person_name"), " at ", (phone, "phone"), "."]
    elif kind == "submit":
        parts = [(relative["name"], "person_name"), " to submit updated income certificate, contact ", (phone, "phone"), "."]
    elif kind == "shifted":
        parts = ["Family shifted house, new contact person ", (relative["name"], "person_name"), " (", (phone, "phone"), ")."]
    else:
        parts = ["Guardian ", (relative["name"], "person_name"), " confirmed enrollment."]
    return compose(parts)


def build_scholarship():
    n = 300
    columns = ["Sr No", "Naam", "Pita ka naam", "DOB", "Umar", "Mob", "Alt Contact", "Aadhar No", "School",
               "Gaon", "Taluka", "District", "Category", "Family Income", "Marks %", "Status", "Remarks",
               "col_17", "IFSC", "Name of scheme"]
    rows, cells = [], []
    for i in range(n):
        area = random.choice(AREAS)
        village = random.choice(area["villages"])
        child = rand_name()
        father = rand_name(gender="M")
        if random.random() < 0.8:
            father["surname"] = child["surname"]
            father["name"] = f"{father['first']} {father['surname']}"
        dob, age = gen_dob_age(10, 22)
        mob = gen_phone()
        if random.random() < 0.5:
            alt, alt_cls = gen_email(child["first"], child["surname"]), "email"
        else:
            alt, alt_cls = gen_phone(), "phone"
        if random.random() < 0.30:
            remark_text, remark_spans = remarks_with_pii()
        else:
            remark_text, remark_spans = random.choice(REMARKS_PLAIN), []
        for sp in remark_spans:
            cells.append({"sheet": None, "row": i, "column": "Remarks", **sp})
        rows.append({
            "Sr No": i + 1, "Naam": child["name"], "Pita ka naam": father["name"],
            "DOB": dob.strftime("%d-%m-%Y"), "Umar": age, "Mob": mob, "Alt Contact": alt,
            "Aadhar No": gen_aadhaar(), "School": f"{random.choice(SCHOOLS)}, {village}",
            "Gaon": village, "Taluka": random.choice(area["talukas"]), "District": area["district"],
            "Category": random.choice(CASTES), "Family Income": random.randint(15000, 250000),
            "Marks %": round(random.uniform(35, 98), 1), "Status": random.choice(STATUSES),
            "Remarks": remark_text, "col_17": gen_bank_account(), "IFSC": gen_ifsc(),
            "Name of scheme": random.choice(SCHEMES),
        })
    df = pd.DataFrame(rows, columns=columns)
    path = os.path.join(OUT_DIR, "scholarship_applicants.csv")
    df.to_csv(path, index=False)
    col_classes = {
        "Sr No": "record_id_non_pii", "Naam": "person_name", "Pita ka naam": "person_name", "DOB": "dob",
        "Umar": "age", "Mob": "phone", "Alt Contact": ["email", "phone"], "Aadhar No": "aadhaar", "School": "none",
        "Gaon": "village", "Taluka": "none", "District": "none", "Category": "caste_category",
        "Family Income": "none", "Marks %": "none", "Status": "none", "Remarks": "free_text_with_pii",
        "col_17": "bank_account", "IFSC": "ifsc", "Name of scheme": "none",
    }
    write_json(os.path.join(OUT_DIR, "scholarship_applicants.columns.json"), col_classes)
    write_json(os.path.join(OUT_DIR, "scholarship_applicants.cells.json"), cells)
    verify_tabular_csv(path, cells)
    record_summary("scholarship_applicants.csv", n, col_classes, cells)


# --------------------------------------------------------------------------
# File 2: donor_transactions.csv
# --------------------------------------------------------------------------
DONOR_NOTES_PLAIN = [
    "Long time supporter, met at annual event.", "Corporate CSR partner, quarterly pledge.",
    "Wants anonymity in annual report.", "First time donor via website campaign.",
    "Prefers email communication only.", "Recurring monthly donor since last year.",
    "Donation made in memory of a family member.", "Requested 80G receipt by post.",
]


def donor_note_with_pii():
    person = rand_name()
    phone = gen_phone()
    kind = random.choice(["called", "contact", "address"])
    if kind == "called":
        parts = ["Called ", (person["name"], "person_name"), " to confirm pledge, ", (phone, "phone"), "."]
    elif kind == "contact":
        parts = ["Alternate contact person: ", (person["name"], "person_name"), " (", (phone, "phone"), ")."]
    else:
        area = random.choice(AREAS)
        village = random.choice(area["villages"])
        addr = gen_address(village, random.choice(area["talukas"]), area["district"], area["state"])
        parts = ["Donor communicated new address for receipt: ", (addr, "address"), "."]
    return compose(parts)


def build_donors():
    n = 250
    columns = ["Receipt No", "Date", "Donor", "Email", "PAN", "Amount", "Mode", "UPI ID", "City", "Note"]
    rows, cells = [], []
    start_date = date(2025, 4, 1)
    for i in range(n):
        donor = rand_name()
        d = start_date + timedelta(days=random.randint(0, 500))
        upi = "" if random.random() < 0.40 else gen_upi(donor["first"], donor["surname"])
        if random.random() < 0.25:
            note_text, note_spans = donor_note_with_pii()
        else:
            note_text, note_spans = random.choice(DONOR_NOTES_PLAIN), []
        for sp in note_spans:
            cells.append({"sheet": None, "row": i, "column": "Note", **sp})
        rows.append({
            "Receipt No": f"RCPT{2025000 + i}", "Date": d.strftime("%d-%m-%Y"), "Donor": donor["name"],
            "Email": gen_email(donor["first"], donor["surname"]), "PAN": gen_pan(),
            "Amount": random.choice([500, 1000, 2500, 5000, 10000, 25000, 50000, 100000]) + random.randint(0, 99),
            "Mode": random.choice(MODES), "UPI ID": upi, "City": random.choice(CITIES), "Note": note_text,
        })
    df = pd.DataFrame(rows, columns=columns)
    path = os.path.join(OUT_DIR, "donor_transactions.csv")
    df.to_csv(path, index=False)
    col_classes = {
        "Receipt No": "record_id_non_pii", "Date": "none", "Donor": "person_name", "Email": "email",
        "PAN": "pan", "Amount": "none", "Mode": "none", "UPI ID": "upi_id", "City": "none",
        "Note": "free_text_with_pii",
    }
    write_json(os.path.join(OUT_DIR, "donor_transactions.columns.json"), col_classes)
    write_json(os.path.join(OUT_DIR, "donor_transactions.cells.json"), cells)
    verify_tabular_csv(path, cells)
    record_summary("donor_transactions.csv", n, col_classes, cells)


# --------------------------------------------------------------------------
# File 3: child_fitness_scores.xlsx
# --------------------------------------------------------------------------
FITNESS_NOTES_PLAIN = [
    "Regular attendance, good improvement.", "Needs encouragement during warm-up.",
    "Missed two sessions due to rain.", "Shows strong stamina for age group.",
    "Recovering from minor ankle sprain.", "Enthusiastic participant, peer motivator.",
]


def fitness_note_with_pii():
    person = rand_name()
    phone = gen_phone()
    if random.random() < 0.5:
        parts = ["Parent ", (person["name"], "person_name"), " requested schedule change, ", (phone, "phone"), "."]
    else:
        parts = ["Flagged to coach by ", (person["name"], "person_name"), ", reachable at ", (phone, "phone"), "."]
    return compose(parts)


def build_fitness_xlsx():
    n = 200
    coaches = [rand_name()["name"] for _ in range(15)]
    children = []
    for i in range(n):
        area = random.choice(AREAS)
        village = random.choice(area["villages"])
        child = rand_name()
        dob, age = gen_dob_age(6, 16)
        children.append({
            "id": f"CFS{i + 1:04d}", "name": child["name"], "gender": "Male" if child["gender"] == "M" else "Female",
            "age": age, "dob": dob, "school": f"{random.choice(SCHOOLS)}, {village}",
            "coach": random.choice(coaches), "village": village, "parent_phone": gen_phone(),
        })
    columns = ["Child ID", "Child name", "Gender", "Age", "DOB", "School", "Coach", "Height cm", "Weight kg",
               "Shuttle run sec", "Sit ups", "Location", "Parent contact", "Notes"]
    cells = []
    sheets = {}
    for sheet_name, delta_months in (("Baseline", 0), ("Endline", 5)):
        rows = []
        for i, c in enumerate(children):
            if random.random() < 0.15:
                note_text, note_spans = fitness_note_with_pii()
            else:
                note_text, note_spans = random.choice(FITNESS_NOTES_PLAIN), []
            for sp in note_spans:
                cells.append({"sheet": sheet_name, "row": i, "column": "Notes", **sp})
            rows.append({
                "Child ID": c["id"], "Child name": c["name"], "Gender": c["gender"], "Age": c["age"],
                "DOB": c["dob"].strftime("%d-%m-%Y"), "School": c["school"], "Coach": c["coach"],
                "Height cm": round(random.uniform(105, 155) + delta_months * 0.3, 1),
                "Weight kg": round(random.uniform(16, 45) + delta_months * 0.2, 1),
                "Shuttle run sec": round(random.uniform(9, 18) - delta_months * 0.05, 1),
                "Sit ups": random.randint(5, 30) + (2 if delta_months else 0),
                "Location": c["village"], "Parent contact": c["parent_phone"], "Notes": note_text,
            })
        sheets[sheet_name] = pd.DataFrame(rows, columns=columns)
    path = os.path.join(OUT_DIR, "child_fitness_scores.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    col_classes = {
        "Child ID": "record_id_non_pii", "Child name": "person_name", "Gender": "none", "Age": "age", "DOB": "dob",
        "School": "none", "Coach": "person_name", "Height cm": "none", "Weight kg": "none",
        "Shuttle run sec": "none", "Sit ups": "none", "Location": "village", "Parent contact": "phone",
        "Notes": "free_text_with_pii",
    }
    write_json(os.path.join(OUT_DIR, "child_fitness_scores.columns.json"), col_classes)
    write_json(os.path.join(OUT_DIR, "child_fitness_scores.cells.json"), cells)
    verify_tabular_xlsx(path, cells)
    record_summary("child_fitness_scores.xlsx", n * 2, col_classes, cells)


# --------------------------------------------------------------------------
# File 4: household_survey.csv
# --------------------------------------------------------------------------
Q26_PLAIN = [
    "Water source is a shared borewell.", "Household reported no major issues this quarter.",
    "House has pucca roof, kutcha walls.", "Family relies on MNREGA work for income.",
    "No school-age children currently out of school.", "Toilet constructed under Swachh Bharat scheme.",
]


def q26_with_pii():
    person = rand_name()
    if random.random() < 0.5:
        parts = ["Visited again, spoke with ", (person["name"], "person_name"), " who confirmed details."]
    else:
        parts = ["Neighbor ", (person["name"], "person_name"), " helped locate the household."]
    return compose(parts)


def build_household_survey():
    n = 150
    enumerators = [rand_name()["name"] for _ in range(10)]
    columns = (["_id", "_uuid", "start", "end", "enumerator_name", "hh_head_name", "respondent_name",
                "respondent_phone", "village", "gps_lat", "gps_lon", "ration_card_no"]
               + [f"q{str(k).zfill(2)}" for k in range(1, 26)] + ["q26_remarks", "_submission_time"])
    rows, cells = [], []
    survey_start = date(2026, 2, 1)
    for i in range(n):
        area = random.choice(AREAS)
        village = random.choice(area["villages"])
        lat, lon = gen_gps(area["bbox"])
        head = rand_name(gender="M")
        if random.random() < 0.6:
            respondent = head
        else:
            respondent = rand_name()
            respondent["surname"] = head["surname"]
            respondent["name"] = f"{respondent['first']} {respondent['surname']}"
        d = survey_start + timedelta(days=random.randint(0, 150))
        start_dt = f"{d.isoformat()}T{random.randint(8, 17):02d}:{random.randint(0, 59):02d}:00"
        end_dt = f"{d.isoformat()}T{random.randint(8, 17):02d}:{random.randint(0, 59):02d}:00"
        sub_dt = f"{d.isoformat()}T{random.randint(18, 20):02d}:{random.randint(0, 59):02d}:00"
        if random.random() < 0.20:
            remark_text, remark_spans = q26_with_pii()
        else:
            remark_text, remark_spans = random.choice(Q26_PLAIN), []
        for sp in remark_spans:
            cells.append({"sheet": None, "row": i, "column": "q26_remarks", **sp})
        row = {
            "_id": 90000 + i, "_uuid": str(uuidlib.UUID(int=random.getrandbits(128))),
            "start": start_dt, "end": end_dt, "enumerator_name": random.choice(enumerators),
            "hh_head_name": head["name"], "respondent_name": respondent["name"],
            "respondent_phone": gen_phone(), "village": village, "gps_lat": lat, "gps_lon": lon,
            "ration_card_no": gen_ration_card(area["state"]),
        }
        for k in range(1, 26):
            key = f"q{str(k).zfill(2)}"
            if k % 3 == 0:
                row[key] = random.choice(["Yes", "No", "Sometimes"])
            elif k % 3 == 1:
                row[key] = random.randint(0, 10)
            else:
                row[key] = random.choice(["Poor", "Average", "Good"])
        row["q26_remarks"] = remark_text
        row["_submission_time"] = sub_dt
        rows.append(row)
    df = pd.DataFrame(rows, columns=columns)
    path = os.path.join(OUT_DIR, "household_survey.csv")
    df.to_csv(path, index=False)
    col_classes = {
        "_id": "record_id_non_pii", "_uuid": "record_id_non_pii", "start": "none", "end": "none",
        "enumerator_name": "person_name", "hh_head_name": "person_name", "respondent_name": "person_name",
        "respondent_phone": "phone", "village": "village", "gps_lat": "gps", "gps_lon": "gps",
        "ration_card_no": "ration_card",
    }
    for k in range(1, 26):
        col_classes[f"q{str(k).zfill(2)}"] = "none"
    col_classes["q26_remarks"] = "free_text_with_pii"
    col_classes["_submission_time"] = "none"
    write_json(os.path.join(OUT_DIR, "household_survey.columns.json"), col_classes)
    write_json(os.path.join(OUT_DIR, "household_survey.cells.json"), cells)
    verify_tabular_csv(path, cells)
    record_summary("household_survey.csv", n, col_classes, cells)


# --------------------------------------------------------------------------
# File 5: field_whatsapp_chat.txt
# --------------------------------------------------------------------------
NONPII_CHATTER = [
    "ok", "theek hai", "haan bhej diya", "kal milte hai", "done", "noted",
    "Aaj 45 bacchon ka checkup hua.", "Kul kharcha 1200 rupaye hua.",
    "Meeting 3 baje hai kal.", "15 forms complete kiye aaj.",
    "Session 2 ghante chala.", "Total 8 gaon cover ho gaye is hafte.",
    "Budget 5000 rupaye approve hua.", "60% attendance thi aaj.",
]


def build_whatsapp():
    senders = [rand_name()["name"] for _ in range(12)]
    lines_target = 180
    aadhaar_idx = random.randint(int(lines_target * 0.3), int(lines_target * 0.6))
    email_idx = random.randint(int(lines_target * 0.15), int(lines_target * 0.5))
    while email_idx == aadhaar_idx:
        email_idx = random.randint(int(lines_target * 0.15), int(lines_target * 0.5))
    content = ""
    spans = []
    cur_date = date(2026, 3, 1)
    for i in range(lines_target):
        cur_date += timedelta(days=random.choice([0, 0, 0, 1]))
        hh, mm = random.randint(7, 21), random.randint(0, 59)
        sender = random.choice(senders)
        area = random.choice(AREAS)
        village = random.choice(area["villages"])
        if i == aadhaar_idx:
            person = rand_name()
            parts = [(person["name"], "person_name"), " ka Aadhaar number ", (gen_aadhaar(), "aadhaar"),
                     " list mein verify karna hai."]
        elif i == email_idx:
            person = rand_name()
            parts = ["Report ", (gen_email(person["first"], person["surname"]), "email"), " pe bhejo please, aaj hi."]
        else:
            kind = random.random()
            if kind < 0.18:
                person = rand_name()
                parts = [(person["name"], "person_name"), " ka number hai ", (gen_phone(), "phone"), ", call kar lena."]
            elif kind < 0.34:
                person = rand_name()
                _, age = gen_dob_age(4, 70)
                parts = [(person["name"], "person_name"), " ", (village, "village"),
                          " mein rehte hai, unki beti ", (str(age), "age"), " saal ki hai."]
            elif kind < 0.48:
                parts = ["Kal ", (village, "village"), " jana hai session ke liye."]
            elif kind < 0.62:
                person = rand_name()
                parts = [(person["name"], "person_name"), " aaj camp me aayi thi, follow up karna hai."]
            elif kind < 0.75:
                person = rand_name()
                parts = ["Beneficiary ", (person["name"], "person_name"), " se baat hui, number ",
                          (gen_phone(), "phone"), " save kar lo."]
            else:
                parts = [random.choice(NONPII_CHATTER)]
        msg_text, msg_spans = compose(parts)
        line_prefix = f"{cur_date.strftime('%d/%m/%y')}, {hh:02d}:{mm:02d} - "
        sender_part = f"{sender}: "
        sender_offset = len(content) + len(line_prefix)
        spans.append({"start": sender_offset, "end": sender_offset + len(sender), "class": "person_name", "text": sender})
        base_offset = sender_offset + len(sender_part)
        for sp in msg_spans:
            spans.append({"start": base_offset + sp["start"], "end": base_offset + sp["end"],
                          "class": sp["class"], "text": sp["text"]})
        content += line_prefix + sender_part + msg_text + "\n"
    path = os.path.join(OUT_DIR, "field_whatsapp_chat.txt")
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    write_json(os.path.join(OUT_DIR, "field_whatsapp_chat.spans.json"), spans)
    verify_txt(path, spans, content)
    SUMMARY.append({"file": "field_whatsapp_chat.txt", "rows": lines_target, "pii_columns": "n/a", "spans": len(spans)})


# --------------------------------------------------------------------------
# File 6: field_observation_report.txt
# --------------------------------------------------------------------------
def build_report():
    n_paragraphs = 25
    special = random.sample(range(n_paragraphs), 7)
    phone_idxs = special[0:2]
    aadhaar_idx, email_idx, vehicle_idx, address_idx, voter_idx = special[2:7]
    content = ""
    spans = []
    officer = rand_name()
    for i in range(n_paragraphs):
        area = random.choice(AREAS)
        village = random.choice(area["villages"])
        taluka = random.choice(area["talukas"])
        beneficiary = rand_name()
        _, age = gen_dob_age(3, 75)
        parts = [f"Visit {i + 1}: The team, led by field officer ", (officer["name"], "person_name"),
                 ", reached ", (village, "village"), f", {taluka} taluka, {area['district']} district, ",
                 f"around {random.randint(9, 17)}:{random.randint(0, 59):02d} hrs. ",
                 "The household of ", (beneficiary["name"], "person_name"), ", aged ",
                 (str(age), "age"), " years, was surveyed. "]
        pct, cnt, rup = random.randint(10, 95), random.randint(2, 60), random.choice([200, 500, 750, 1200, 2500, 4000])
        parts.append(f"Attendance in the last cycle was {pct} percent, with {cnt} children present and an outlay of Rs {rup}. ")
        if i == aadhaar_idx:
            parts += [(beneficiary["name"], "person_name"), "'s Aadhaar number ", (gen_aadhaar(), "aadhaar"),
                      " was recorded for the ration card update. "]
        if i in phone_idxs:
            parts += ["Contact was noted as ", (gen_phone(), "phone"), " for follow-up. "]
        if i == email_idx:
            parts += ["Photographs and consent forms were emailed to ",
                      (gen_email(beneficiary["first"], beneficiary["surname"]), "email"), " the same evening. "]
        if i == vehicle_idx:
            parts += ["The team travelled in vehicle ", (gen_vehicle(), "vehicle_number"),
                      " provided by the district office. "]
        if i == address_idx:
            addr = gen_address(village, taluka, area["district"], area["state"])
            parts += ["The household address was recorded as ", (addr, "address"), " for the case file. "]
        if i == voter_idx:
            parts += ["Voter ID card was checked against EPIC number ", (gen_voter_id(), "voter_id"),
                      " to confirm residence. "]
        if random.random() < 0.4:
            lat, lon = gen_gps(area["bbox"])
            parts += ["GPS location logged at ", (f"{lat}, {lon}", "gps"), ". "]
        parts.append(f"{random.randint(1, 10)} households were covered in this locality this week.")
        para_text, para_spans = compose(parts)
        offset = len(content)
        for sp in para_spans:
            spans.append({"start": offset + sp["start"], "end": offset + sp["end"], "class": sp["class"], "text": sp["text"]})
        content += para_text + "\n\n"
    path = os.path.join(OUT_DIR, "field_observation_report.txt")
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    write_json(os.path.join(OUT_DIR, "field_observation_report.spans.json"), spans)
    verify_txt(path, spans, content)
    SUMMARY.append({"file": "field_observation_report.txt", "rows": n_paragraphs, "pii_columns": "n/a", "spans": len(spans)})


# --------------------------------------------------------------------------
README_TEXT = """# PII corpus ground truth

Synthetic NGO-style fixtures for benchmarking local PII redaction engines.
Everything in this directory is fabricated: names, phone numbers, Aadhaar/PAN
numbers, bank details, villages, GPS points. Aadhaar numbers are structurally
valid (pass the Verhoeff checksum) but are not real allotted numbers. Built by
`../build_pii_corpus.py` with `random.seed(20260821)`; regenerate with the
same interpreter to reproduce byte-for-byte.

## Taxonomy

    person_name, phone, email, aadhaar, pan, bank_account, ifsc, upi_id, dob,
    age, address, village, gps, caste_category, ration_card, voter_id,
    vehicle_number, record_id_non_pii, free_text_with_pii, none

## Ground truth files

For each tabular file `<basename>.{csv,xlsx}`:

- `<basename>.columns.json`: `{column_name: class}`. A column that mixes two
  entity types cell-by-cell (e.g. "Alt Contact" holds either an email or a
  phone number) is given a **list** of classes, e.g. `["email", "phone"]`,
  instead of a single string.
- `<basename>.cells.json`: a flat list of PII spans found inside free-text
  cells (Remarks / Note / Notes / q26_remarks columns), one entry per span:
  `{"sheet": "Baseline"|null, "row": 0-based data row (header excluded),
  "column": name, "start": char offset in the cell string, "end": ...,
  "class": ..., "text": exact substring}`. `sheet` is `null` for CSV files.
  Cells with no embedded PII contribute no entries.

For each narrative `.txt` file:

- `<basename>.spans.json`: a flat list of `{"start", "end", "class", "text"}`
  character offsets into the whole file's content. In the WhatsApp export,
  the sender name in `dd/mm/yy, hh:mm - Sender Name: message` is itself a
  `person_name` span.

All offsets were produced by construction (the generator inserts each PII
value and records its own start/end at insertion time) and were re-verified
after writing: every file was reopened from disk and
`content[start:end] == text` was asserted for every recorded span before the
generator was allowed to report success.

## Files

1. `scholarship_applicants.csv` (300 rows) - Hinglish scholarship register.
2. `donor_transactions.csv` (250 rows) - donation ledger.
3. `child_fitness_scores.xlsx` (200 children x 2 sheets: Baseline, Endline).
4. `household_survey.csv` (150 rows, KoBo-style export with q01..q25
   indicator columns).
5. `field_whatsapp_chat.txt` (~180 lines) - WhatsApp export format.
6. `field_observation_report.txt` (25 paragraphs) - narrative field visit
   report.

## Judgement calls

- `Taluka`, `District`, `City` are marked `none`: the taxonomy only has a
  `village` class for place-level PII, and these broader administrative
  units are shared by thousands of people, so they are not treated as
  identifying on their own.
- `School` (institution name) and `Name of scheme` are `none`: they are not
  personal identifiers, and "Name of scheme" is deliberately a header that
  *looks* like it might be sensitive but is not.
- `Coach` and `enumerator_name` are `person_name` even though they name
  staff rather than beneficiaries - the taxonomy has no separate class for
  staff vs. beneficiary names.
- `gps_lat` and `gps_lon` are each independently labelled `gps`, even though
  a single coordinate alone is less identifying than the pair.
- `start`, `end`, `_submission_time`, and `Date` (donor ledger) are `none`:
  timestamps have no dedicated class in the taxonomy and are not treated as
  PII by themselves here.
- `col_17` (scholarship file) is a deliberately generic header that actually
  holds bank account numbers; `Name of scheme` is a header that looks
  sensitive but is not; `q01..q25` are generic KoBo-style indicator headers
  that hold no PII.
"""


def main():
    build_scholarship()
    build_donors()
    build_fitness_xlsx()
    build_household_survey()
    build_whatsapp()
    build_report()
    with open(os.path.join(OUT_DIR, "README.md"), "w", encoding="utf-8") as f:
        f.write(README_TEXT)
    print()
    header = f"{'file':32} {'rows':>6} {'pii_cols':>9} {'spans':>7}"
    print(header)
    print("-" * len(header))
    for s in SUMMARY:
        print(f"{s['file']:32} {str(s['rows']):>6} {str(s['pii_columns']):>9} {s['spans']:>7}")


if __name__ == "__main__":
    main()
