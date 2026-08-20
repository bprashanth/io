#!/usr/bin/env python3
"""Build deterministic synthetic Excel, PDF, and CSV screening inputs."""

from __future__ import annotations

import csv
import datetime
import io
import re
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


ROOT = Path(__file__).resolve().parents[1] / "cases"


def normalize_zip(path: Path) -> None:
    """Remove wall-clock timestamps from an OOXML zip container."""
    source = path.read_bytes()
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(source), "r") as reader, zipfile.ZipFile(
        output, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as writer:
        for name in sorted(reader.namelist()):
            old = reader.getinfo(name)
            info = zipfile.ZipInfo(name, date_time=(2026, 8, 20, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = old.external_attr
            info.create_system = old.create_system
            contents = reader.read(name)
            # openpyxl overwrites the modified property with wall-clock UTC in
            # Workbook.save(), even when Workbook.properties.modified was set.
            if name == "docProps/core.xml":
                contents = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]+(</dcterms:modified>)",
                    rb"\g<1>2026-08-20T00:00:00Z\g<2>",
                    contents,
                )
            writer.writestr(info, contents)
    path.write_bytes(output.getvalue())


def build_xlsx() -> None:
    output = ROOT / "dev-xlsx-health-001" / "inputs" / "maternal_health.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    rows = [
        ("Gaya", 2021, 1000, 680, 600),
        ("Gaya", 2022, 1100, 792, 693),
        ("Gaya", 2023, 1200, 900, 816),
        ("Nalanda", 2021, 900, 675, 630),
        ("Nalanda", 2022, 1000, 780, 720),
        ("Nalanda", 2023, 1100, 880, 825),
        ("Purnia", 2021, 1200, 720, 660),
        ("Purnia", 2022, 1300, 819, 728),
        ("Purnia", 2023, 1400, 924, 812),
    ]
    workbook = Workbook()
    fixed_time = datetime.datetime(2026, 8, 20, 0, 0, 0)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    data = workbook.active
    data.title = "District Data"
    data.append(
        [
            "district",
            "year",
            "pregnancies_registered",
            "institutional_deliveries",
            "postnatal_check_48h",
        ]
    )
    for row in rows:
        data.append(row)
    for cell in data[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    data.freeze_panes = "A2"
    data.auto_filter.ref = data.dimensions
    for width, letter in zip((16, 10, 25, 25, 24), "ABCDE", strict=True):
        data.column_dimensions[letter].width = width

    notes = workbook.create_sheet("Indicator Notes")
    notes.append(["indicator", "definition", "formula", "unit", "source"])
    notes.append(
        [
            "Institutional delivery coverage",
            "Share of registered pregnancies with an institutional delivery",
            "institutional_deliveries / pregnancies_registered * 100",
            "percent",
            "synthetic maternal-health workbook fixture",
        ]
    )
    notes.append(
        [
            "Postnatal check within 48 hours coverage",
            "Share of registered pregnancies with a recorded postnatal check within 48 hours",
            "postnatal_check_48h / pregnancies_registered * 100",
            "percent",
            "synthetic maternal-health workbook fixture",
        ]
    )
    notes.append(
        [
            "Important",
            "Illustrative benchmark data; not official statistics and not person-level records",
            "",
            "",
            "synthetic maternal-health workbook fixture",
        ]
    )
    for cell in notes[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="548235")
    for width, letter in zip((38, 74, 62, 14, 46), "ABCDE", strict=True):
        notes.column_dimensions[letter].width = width
    workbook.save(output)
    normalize_zip(output)


def build_irregular_xlsx() -> None:
    output = ROOT / "dev-xlsx-headers-001" / "inputs" / "school_attendance_nested.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    fixed_time = datetime.datetime(2026, 8, 20, 0, 0, 0)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    report = workbook.active
    report.title = "Attendance Report"

    def add_table(title_row: int, number: int, level: str, rows: list[tuple[object, ...]]) -> None:
        report.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=5)
        title = report.cell(title_row, 1, f"Table {number}. {level} school attendance by block, sex and year")
        title.font = Font(bold=True, color="FFFFFF", size=12)
        title.fill = PatternFill("solid", fgColor="1F4E78" if level == "Primary" else "548235")
        header = title_row + 2
        report.merge_cells(start_row=header, start_column=1, end_row=header + 1, end_column=1)
        report.cell(header, 1, "Block")
        report.merge_cells(start_row=header, start_column=2, end_row=header, end_column=3)
        report.cell(header, 2, "Boys attendance (%)")
        report.merge_cells(start_row=header, start_column=4, end_row=header, end_column=5)
        report.cell(header, 4, "Girls attendance (%)")
        for column, value in enumerate(("2022", "2023", "2022", "2023"), start=2):
            report.cell(header + 1, column, value)
        for row in range(header, header + 2):
            for column in range(1, 6):
                cell = report.cell(row, column)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="5B9BD5")
        for values in rows:
            report.append(values)

    add_table(1, 1, "Primary", [
        ("Tekari", 78, 82, 75, 80),
        ("Wazirganj", 72, 76, 70, 74),
        ("Atri", 68, 73, 66, 71),
    ])
    add_table(11, 2, "Secondary", [
        ("Tekari", 70, 74, 68, 73),
        ("Wazirganj", 64, 68, 62, 66),
        ("Atri", 60, 65, 58, 63),
    ])
    report.freeze_panes = "B5"
    for width, letter in zip((20, 18, 18, 18, 18), "ABCDE", strict=True):
        report.column_dimensions[letter].width = width

    notes = workbook.create_sheet("Read Me")
    notes.merge_cells("A1:D1")
    notes["A1"] = "About this workbook"
    notes["A1"].font = Font(bold=True, color="FFFFFF")
    notes["A1"].fill = PatternFill("solid", fgColor="7F6000")
    notes.append(["source", "synthetic school-attendance nested workbook fixture", None, None])
    notes.append(["unit", "percent", None, None])
    notes.append(["warning", "Illustrative aggregate benchmark data; not official statistics", None, None])

    decoy = workbook.create_sheet("Enrolment Raw")
    decoy.append(["school_code", "block", "year", "enrolled"])
    for index in range(1, 31):
        decoy.append([f"S{index:03d}", ("Tekari", "Wazirganj", "Atri")[index % 3], 2023, 100 + index])
    for cell in decoy[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="A5A5A5")

    workbook.save(output)
    normalize_zip(output)


def build_pdf() -> None:
    output = ROOT / "dev-pdf-health-001" / "inputs" / "facility_delivery_report.pdf"
    output.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        str(output), pagesize=A4, leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
        title="Illustrative District Facility Delivery Report",
        author="NGO dashboard benchmark",
    )
    story = [
        Paragraph("Illustrative District Facility Delivery Report", styles["Title"]),
        Spacer(1, 8 * mm),
        Paragraph("Reporting years: 2021 to 2023", styles["Heading2"]),
        Spacer(1, 3 * mm),
        Paragraph(
            "This short digital PDF is a synthetic benchmark fixture. It contains no "
            "personal records and must not be described as official statistics.",
            styles["BodyText"],
        ),
        Spacer(1, 5 * mm),
        Paragraph("Indicator definition", styles["Heading2"]),
        Paragraph(
            "Facility delivery coverage is the percentage of recorded deliveries that "
            "occurred in a health facility. The report supplies percentages only; it "
            "does not supply counts or causal explanations.",
            styles["BodyText"],
        ),
        Spacer(1, 5 * mm),
        Paragraph(
            "Source label: synthetic facility-delivery PDF fixture. The data table is "
            "Table 1 on page 2.",
            styles["BodyText"],
        ),
        PageBreak(),
        Paragraph("Table 1. Facility delivery coverage by district and year", styles["Heading1"]),
        Spacer(1, 4 * mm),
    ]
    table_data = [
        ["District", "2021 (%)", "2022 (%)", "2023 (%)"],
        ["Gaya", "62.0", "67.0", "72.0"],
        ["Nalanda", "70.0", "74.0", "79.0"],
        ["Purnia", "55.0", "61.0", "66.0"],
        ["Kishanganj", "50.0", "56.0", "60.0"],
    ]
    table = Table(table_data, colWidths=[60 * mm, 32 * mm, 32 * mm, 32 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#7F8C8D")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF2F8")]),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.extend(
        [
            table,
            Spacer(1, 5 * mm),
            Paragraph(
                "Source: synthetic facility-delivery PDF fixture. Unit: percent. "
                "Values are illustrative and are not official statistics.",
                styles["BodyText"],
            ),
        ]
    )
    def invariant_canvas(*args, **kwargs):
        kwargs["invariant"] = 1
        return canvas.Canvas(*args, **kwargs)

    document.build(story, canvasmaker=invariant_canvas)


def build_safe_csv() -> None:
    output = ROOT / "dev-safe-programme-001" / "inputs" / "women_livelihood_outcomes.csv"
    output.parent.mkdir(parents=True, exist_ok=True)
    rows = [
        ("Gaya", 2022, 1000, 800, 520),
        ("Gaya", 2023, 1100, 902, 616),
        ("Nalanda", 2022, 900, 765, 540),
        ("Nalanda", 2023, 1000, 870, 650),
        ("Purnia", 2022, 1200, 840, 504),
        ("Purnia", 2023, 1300, 949, 598),
        ("Kishanganj", 2022, 800, 520, 280),
        ("Kishanganj", 2023, 900, 612, 342),
    ]
    with output.open("w", newline="") as stream:
        writer = csv.writer(stream, lineterminator="\n")
        writer.writerow(
            [
                "district",
                "year",
                "women_enrolled",
                "completed_training",
                "employed_at_6_months",
                "source",
            ]
        )
        for row in rows:
            writer.writerow((*row, "synthetic women livelihood outcomes fixture"))


def main() -> None:
    build_xlsx()
    build_irregular_xlsx()
    build_pdf()
    build_safe_csv()


if __name__ == "__main__":
    main()
