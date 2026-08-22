#!/usr/bin/env python3
"""Run the frozen v2 plain-language-to-SQL gate against an OpenAI endpoint."""

from __future__ import annotations

import argparse
import hashlib
import itertools
import json
import math
import re
import signal
import time
import urllib.request
from pathlib import Path
from typing import Any

import duckdb
import pandas as pd
import sqlglot
from openpyxl import load_workbook
from sqlglot import expressions as exp


ROOT = Path(__file__).resolve().parents[2]


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, ensure_ascii=False) + "\n")


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load_table(spec: dict[str, Any]) -> pd.DataFrame:
    path = ROOT / spec["path"]
    if spec["kind"] == "csv":
        return pd.read_csv(path)
    if spec["kind"] == "xlsx":
        book = load_workbook(path, data_only=True, read_only=True)
        sheet = book[spec["sheet"]]
        rows = list(sheet.iter_rows(values_only=True))
        return pd.DataFrame(rows[1:], columns=rows[0])
    raise ValueError(f"unsupported table kind: {spec['kind']}")


def create_database(manifest: dict[str, Any]) -> tuple[duckdb.DuckDBPyConnection, dict[str, Any]]:
    db = duckdb.connect(":memory:", config={"enable_external_access": "false"})
    evidence: dict[str, Any] = {"tables": []}
    for index, spec in enumerate(manifest["tables"]):
        frame = load_table(spec)
        temporary = f"input_frame_{index}"
        db.register(temporary, frame)
        db.execute(f'CREATE TABLE "{spec["name"]}" AS SELECT * FROM "{temporary}"')
        db.unregister(temporary)
        evidence["tables"].append({
            "name": spec["name"],
            "path": spec["path"],
            "sha256": sha256(ROOT / spec["path"]),
            "rows": len(frame),
            "columns": [{"name": str(name), "type": str(frame[name].dtype)} for name in frame.columns],
        })
    return db, evidence


def ddl(db: duckdb.DuckDBPyConnection, manifest: dict[str, Any]) -> str:
    statements = []
    for table in manifest["tables"]:
        name = table["name"]
        columns = db.execute(
            "SELECT column_name, data_type FROM information_schema.columns "
            "WHERE table_name = ? ORDER BY ordinal_position",
            [name],
        ).fetchall()
        body = ",\n  ".join(f'"{column}" {kind}' for column, kind in columns)
        statements.append(f'CREATE TABLE "{name}" (\n  {body}\n);')
    return "\n\n".join(statements)


def make_prompt(question: str, schema: str, reference: str, repair: str | None = None) -> str:
    parts = [
        "You are a SQLite expert. Read the database schema and return exactly one read-only SQL query that answers the user question.",
        "Do not explain the query. Do not invent tables, columns or values. Use NULL handling and floating-point arithmetic where needed.",
        "[User Question]",
        question,
        "[Database Schema]",
        schema,
        "[Reference Information]",
        reference,
    ]
    if repair:
        parts.extend(["[Previous Query Error]", repair, "Return one corrected read-only query."])
    parts.extend(["[User Question]", question, "```sql"])
    return "\n".join(parts)


def known_categories(db: duckdb.DuckDBPyConnection, manifest: dict[str, Any], limit: int = 50) -> dict[str, dict[str, list[str]]]:
    """Distinct values of low-cardinality text columns, as the live shell shows them."""
    result: dict[str, dict[str, list[str]]] = {}
    for table in manifest["tables"]:
        name = table["name"]
        columns = db.execute(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_name = ? AND data_type IN ('VARCHAR', 'TEXT') ORDER BY ordinal_position",
            [name],
        ).fetchall()
        per_table: dict[str, list[str]] = {}
        for (column,) in columns:
            quoted = column.replace('"', '""')
            values = [str(row[0]) for row in db.execute(
                f'SELECT DISTINCT "{quoted}" FROM "{name}" WHERE "{quoted}" IS NOT NULL ORDER BY 1 LIMIT {limit + 1}'
            ).fetchall()]
            if len(values) <= limit:
                per_table[column] = values
        result[name] = per_table
    return result


def make_shell_prompt(question: str, schema: str, categories: dict[str, Any], repair: str | None = None,
                      plus: bool = False, io: bool | str = False) -> str:
    """The prompt family used by run_local_first_insight.py (single turn, no prior turns).
    ``plus`` adds one generic aggregate-row rule that the live shell does not have yet.
    ``io`` adds the two rules the io desktop app ships with (quoting/UNION alignment, fuzzy matching hint)."""
    sections = [
        "You translate ordinary questions into exactly one read-only DuckDB SELECT or WITH query.",
        "Return SQL only. Never write files, install software, calculate values yourself, or invent columns.",
        "Use only the supplied schema. Preserve relevant scope from earlier turns unless the newest question changes it.",
        "A requested rate must use its stated numerator and denominator. Percentage-point change means later percentage minus earlier percentage, not relative growth.",
        "When the user explicitly asks for percent or percentage, multiply numerator divided by denominator by 100.",
        "When the user asks to compare years or groups without asking for a change, keep those year/group columns and requested measures in the output.",
        "When several known category values are named, filter them separately with IN rather than concatenating them.",
        *(["Aggregate or total rows that sit inside a detail table (for example a 'Total' label in a district column) are not detail rows; exclude them when the user asks about the detail level."] if plus else []),
        "Use NULLIF for denominators. Missing values remain missing and are not zero.",
        *(["Quote column names with double quotes when they contain spaces or odd characters. When two tables describe the same thing with different column names, UNION ALL them with aligned aliases."] if io else []),
        *(["For approximate name matching across tables use jaro_winkler_similarity(lower(trim(a)), lower(trim(b))) >= 0.9, keep only the single best match per row (QUALIFY ROW_NUMBER() OVER (PARTITION BY row ORDER BY similarity DESC) = 1), then LEFT JOIN from the reference table so rows with zero matches are kept."] if io == "full" else []),
        f"SCHEMA:\n{schema}",
        f"KNOWN CATEGORICAL VALUES (local or trusted 27B tier only):\n{json.dumps(categories, ensure_ascii=False)}",
        f"CURRENT QUESTION:\n{question}",
    ]
    if repair:
        sections.append(f"REJECTED QUERY ERROR:\n{repair}\nReturn a complete corrected query.")
    return "\n\n".join(sections)


def endpoint_json(url: str, body: dict[str, Any], timeout: int, api_key: str | None = None) -> dict[str, Any]:
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    request = urllib.request.Request(
        url,
        data=json.dumps(body).encode(),
        headers=headers,
        method="POST",
    )
    def timed_out(_signum: int, _frame: Any) -> None:
        raise TimeoutError(f"model response exceeded {timeout} seconds")

    previous_handler = signal.signal(signal.SIGALRM, timed_out)
    signal.setitimer(signal.ITIMER_REAL, timeout)
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return json.load(response)
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, previous_handler)


def call_model(
    endpoint: str,
    model: str,
    prompt: str,
    timeout: int,
    api_key: str | None,
    max_tokens: int,
    reasoning_effort: str | None,
    temperature: float,
    no_think_template: bool = False,
) -> tuple[str, dict[str, Any]]:
    started = time.monotonic()
    body: dict[str, Any] = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    if reasoning_effort == "none":
        # Thinking models otherwise spend the whole budget in the hidden
        # reasoning channel and return empty content, which the gate counts
        # as a failure. Ask the provider to disable thinking.
        body["reasoning"] = {"enabled": False}
    elif reasoning_effort:
        body["reasoning"] = {"effort": reasoning_effort}
    if no_think_template:
        # llama.cpp / vLLM honour the chat-template switch, not the OpenRouter field.
        body["chat_template_kwargs"] = {"enable_thinking": False}
    raw = endpoint_json(
        endpoint.rstrip("/") + "/chat/completions",
        body,
        timeout,
        api_key,
    )
    duration = round(time.monotonic() - started, 3)
    text = raw["choices"][0]["message"].get("content")
    if not text:
        raise ValueError("model returned no answer content")
    return text, {
        "id": raw.get("id"),
        "model": raw.get("model"),
        "finish_reason": raw["choices"][0].get("finish_reason"),
        "usage": raw.get("usage"),
        "duration_seconds": duration,
    }


def extract_sql(text: str) -> str:
    fenced = re.search(r"```(?:sql)?\s*(.*?)```", text, flags=re.I | re.S)
    candidate = fenced.group(1) if fenced else text
    start = re.search(r"\b(?:WITH|SELECT)\b", candidate, flags=re.I)
    if not start:
        raise ValueError("response contains no SELECT or WITH query")
    candidate = candidate[start.start():].strip().removesuffix("```").strip()
    parsed = sqlglot.parse(candidate, read="sqlite")
    if len(parsed) != 1 or not isinstance(parsed[0], exp.Query):
        raise ValueError("response is not exactly one read-only query")
    forbidden = (exp.Insert, exp.Update, exp.Delete, exp.Create, exp.Drop, exp.Command, exp.Copy)
    if any(isinstance(node, forbidden) for node in parsed[0].walk()):
        raise ValueError("query contains a forbidden operation")
    return parsed[0].sql(dialect="duckdb")


def clean(value: Any) -> Any:
    if value is None or pd.isna(value):
        return None
    if hasattr(value, "item"):
        value = value.item()
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


def rows(db: duckdb.DuckDBPyConnection, query: str) -> list[list[Any]]:
    db.execute("EXPLAIN " + query)
    return [[clean(value) for value in row] for row in db.execute(query).fetchall()]


def equal_value(left: Any, right: Any, tolerance: float | tuple[float, float]) -> bool:
    """Compare two cells. ``tolerance`` is either one value used for both the
    relative and absolute tolerance (the v2 gate contract) or a
    ``(rel_tol, abs_tol)`` pair (holdout-v2 and later, where gold values are
    unrounded and a model's ROUND(x, 1) must still count as equal)."""
    if isinstance(left, bool) or isinstance(right, bool):
        return left == right
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        rel, absolute = tolerance if isinstance(tolerance, tuple) else (tolerance, tolerance)
        return math.isclose(float(left), float(right), rel_tol=rel, abs_tol=absolute)
    return left == right


def manifest_tolerance(manifest: dict[str, Any]) -> float | tuple[float, float]:
    comparison = manifest["comparison"]
    if "numeric_abs_tolerance" in comparison or "numeric_rel_tolerance" in comparison:
        return (
            float(comparison.get("numeric_rel_tolerance", 1e-06)),
            float(comparison.get("numeric_abs_tolerance", 1e-06)),
        )
    return float(comparison["numeric_tolerance"])


def equal_row(left: list[Any], right: list[Any], tolerance: float) -> bool:
    return len(left) == len(right) and all(equal_value(a, b, tolerance) for a, b in zip(left, right))


def equal_rows(
    actual: list[list[Any]],
    expected: list[list[Any]],
    tolerance: float,
    order_matters: bool,
) -> bool:
    if not expected:
        return not actual
    actual_widths = {len(row) for row in actual}
    expected_widths = {len(row) for row in expected}
    if len(actual_widths) != 1 or len(expected_widths) != 1:
        return False
    actual_width = next(iter(actual_widths))
    expected_width = next(iter(expected_widths))
    if len(actual) == len(expected) and actual_width >= expected_width:
        for indices in itertools.combinations(range(actual_width), expected_width):
            projected = [[row[index] for index in indices] for row in actual]
            if order_matters:
                if all(equal_row(left, right, tolerance) for left, right in zip(projected, expected)):
                    return True
                continue
            unmatched = list(expected)
            for row in projected:
                match = next(
                    (index for index, candidate in enumerate(unmatched) if equal_row(row, candidate, tolerance)),
                    None,
                )
                if match is None:
                    break
                unmatched.pop(match)
            else:
                if not unmatched:
                    return True
    if order_matters:
        return False

    # A dashboard query may return the same obligations in a wide comparison:
    # one row per leading group with repeated period/measure blocks. Accept that
    # only when each expected long-form suffix occurs contiguously in the
    # corresponding wide row and the group sets are identical.
    groups: list[tuple[Any, list[list[Any]]]] = []
    for expected_row in expected:
        group = next((entry for entry in groups if equal_value(entry[0], expected_row[0], tolerance)), None)
        if group is None:
            group = (expected_row[0], [])
            groups.append(group)
        group[1].append(expected_row[1:])
    if len(actual) != len(groups) or any(not row for row in actual):
        return False
    for key, suffixes in groups:
        candidate = next((row for row in actual if equal_value(row[0], key, tolerance)), None)
        if candidate is None:
            return False
        tail = candidate[1:]
        for suffix in suffixes:
            if not any(
                all(equal_value(left, right, tolerance) for left, right in zip(tail[start:], suffix))
                for start in range(len(tail) - len(suffix) + 1)
            ):
                return False
    return True


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", type=Path, default=Path("benchmarks/v2/query-suite-v2.json"))
    parser.add_argument("--endpoint", default="http://127.0.0.1:8020/v1")
    parser.add_argument("--model", required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--timeout-seconds", type=int, default=120)
    parser.add_argument("--api-key-file", type=Path)
    parser.add_argument("--max-tokens", type=int, default=1024)
    parser.add_argument("--reasoning-effort", choices=("none", "low", "medium", "high", "xhigh"))
    parser.add_argument("--temperature", type=float, default=0.0)
    parser.add_argument("--no-think-template", action="store_true", help="send chat_template_kwargs.enable_thinking=false (local llama.cpp/vLLM servers)")
    parser.add_argument("--only-task", action="append", default=[])
    parser.add_argument("--only-sample", action="append", default=[])
    parser.add_argument("--max-samples", type=int)
    parser.add_argument("--prompt-style", choices=("gate", "shell", "shell-plus", "io", "io-quote"), default="gate",
                        help="gate: frozen v2 SQLite-expert prompt (default); shell: the live shell's DuckDB prompt with known categorical values; shell-plus: shell plus one generic aggregate-row rule; io: shell plus the io desktop app's two rules (quoting/UNION alignment, jaro_winkler hint); io-quote: shell plus the quoting/UNION rule only")
    args = parser.parse_args()
    api_key = None
    if args.api_key_file:
        api_key = json.loads(args.api_key_file.expanduser().read_text())["api_key"]

    manifest = json.loads(args.manifest.read_text())
    db, data_evidence = create_database(manifest)
    table_specs = {table["name"]: table for table in manifest["tables"]}
    samples = [
        {"id": f"{task['id']}-p{index + 1:02d}", "task": task, "question": question}
        for task in manifest["tasks"]
        if not args.only_task or task["id"] in args.only_task
        for index, question in enumerate(task["phrasings"])
        if not args.only_sample or f"{task['id']}-p{index + 1:02d}" in args.only_sample
    ]
    if args.max_samples:
        samples = samples[:args.max_samples]

    args.output.mkdir(parents=True, exist_ok=True)
    write_json(args.output / "run-config.json", {
        "schema_version": 1,
        "manifest": str(args.manifest),
        "manifest_sha256": sha256(args.manifest),
        "endpoint": args.endpoint,
        "requested_model": args.model,
        "credential_source": str(args.api_key_file) if args.api_key_file else None,
        "timeout_seconds": args.timeout_seconds,
        "max_tokens": args.max_tokens,
        "reasoning_effort": args.reasoning_effort,
        "temperature": args.temperature,
        "prompt_style": args.prompt_style,
        "samples": len(samples),
        "data": data_evidence,
    })
    records = []
    tolerance = manifest_tolerance(manifest)
    for sample in samples:
        sample_dir = args.output / "samples" / sample["id"]
        expected = rows(db, sample["task"]["gold_sql"])
        scoped_manifest = {"tables": [table_specs[name] for name in sample["task"]["tables"]]}
        schema = ddl(db, scoped_manifest)
        reference = sample["task"].get(
            "reference",
            "Percent means 100.0 times numerator divided by the stated denominator. Percentage-point change means later rate minus earlier rate.",
        )
        if args.prompt_style.startswith("shell") or args.prompt_style.startswith("io"):
            categories = known_categories(db, scoped_manifest)
            prompt = make_shell_prompt(sample["question"], schema, categories, plus=args.prompt_style == "shell-plus", io={"io": "full", "io-quote": True}.get(args.prompt_style, False))
        else:
            prompt = make_prompt(sample["question"], schema, reference)
        attempts = []
        actual: list[list[Any]] | None = None
        final_sql = None
        for attempt in range(1, 3):
            request_record = {"attempt": attempt, "prompt": prompt}
            attempt_started = time.monotonic()
            try:
                response, api = call_model(
                    args.endpoint,
                    args.model,
                    prompt,
                    args.timeout_seconds,
                    api_key,
                    args.max_tokens,
                    args.reasoning_effort,
                    args.temperature,
                    args.no_think_template,
                )
                request_record.update({"response": response, "api": api})
                final_sql = extract_sql(response)
                actual = rows(db, final_sql)
                request_record["duration_seconds"] = round(time.monotonic() - attempt_started, 3)
                request_record.update({"sql_duckdb": final_sql, "rows": actual})
                attempts.append(request_record)
                break
            except Exception as error:
                request_record["duration_seconds"] = round(time.monotonic() - attempt_started, 3)
                request_record["error"] = f"{type(error).__name__}: {error}"
                attempts.append(request_record)
                if attempt == 2:
                    break
                if not isinstance(error, TimeoutError):
                    if args.prompt_style.startswith("shell") or args.prompt_style.startswith("io"):
                        prompt = make_shell_prompt(sample["question"], schema, categories, request_record["error"],
                                                   plus=args.prompt_style == "shell-plus", io={"io": "full", "io-quote": True}.get(args.prompt_style, False))
                    else:
                        prompt = make_prompt(sample["question"], schema, reference, request_record["error"])
        order_matters = bool(sample["task"].get("order_matters", False))
        passed = actual is not None and equal_rows(actual, expected, tolerance, order_matters)
        record = {
            "id": sample["id"],
            "task_id": sample["task"]["id"],
            "pattern": sample["task"]["pattern"],
            "question": sample["question"],
            "gold_sql": sample["task"]["gold_sql"],
            "expected_rows": expected,
            "actual_rows": actual,
            "row_order_required": order_matters,
            "sql_duckdb": final_sql,
            "attempts": attempts,
            "passed": passed,
            "failure_class": None if passed else ("request_or_execution" if actual is None else "semantic_result_mismatch"),
        }
        write_json(sample_dir / "record.json", record)
        records.append(record)
        print(json.dumps({"id": record["id"], "passed": passed, "attempts": len(attempts)}), flush=True)

    passed = sum(record["passed"] for record in records)
    summary = {
        "schema_version": 1,
        "model": args.model,
        "samples": len(records),
        "passed": passed,
        "execution_accuracy": passed / len(records) if records else 0,
        "first_attempt_executable": sum(bool(record["attempts"] and record["attempts"][0].get("rows") is not None) for record in records),
        "semantic_mismatches": sum(record["failure_class"] == "semantic_result_mismatch" for record in records),
        "request_or_execution_failures": sum(record["failure_class"] == "request_or_execution" for record in records),
        "successful_model_calls": sum(
            bool(attempt.get("api")) for record in records for attempt in record["attempts"]
        ),
        "timeout_count": sum(
            str(attempt.get("error", "")).startswith("TimeoutError:")
            for record in records for attempt in record["attempts"]
        ),
        "total_tokens": sum(
            int(attempt.get("api", {}).get("usage", {}).get("total_tokens") or 0)
            for record in records for attempt in record["attempts"]
        ),
        "cost_usd": round(sum(
            float(attempt.get("api", {}).get("usage", {}).get("cost") or 0)
            for record in records for attempt in record["attempts"]
        ), 8),
        "duration_seconds": round(sum(
            attempt.get("duration_seconds", attempt.get("api", {}).get("duration_seconds", 0))
            for record in records for attempt in record["attempts"]
        ), 3),
        "records": [{key: record[key] for key in ("id", "task_id", "passed", "failure_class")} for record in records],
    }
    write_json(args.output / "result.json", summary)
    print(json.dumps({key: summary[key] for key in (
        "model", "samples", "passed", "execution_accuracy", "duration_seconds"
    )}))
    db.close()
    return 0 if summary["execution_accuracy"] >= 0.85 else 1


if __name__ == "__main__":
    raise SystemExit(main())
