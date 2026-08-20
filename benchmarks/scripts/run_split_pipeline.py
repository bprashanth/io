#!/usr/bin/env python3
"""Development runner: model plan -> validated DuckDB -> static dashboard.

The model never writes SQL, displayed numbers, HTML or JavaScript. Phase one
accepts one CSV; later input adapters will produce the same table contract.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import math
import re
import signal
import subprocess
import time
import urllib.request
from pathlib import Path
from typing import Any

import duckdb
import jsonschema
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
SCHEMA_PATH = ROOT / "benchmarks/schemas/analysis-plan.schema.json"
KEY_PATH = Path.home() / ".config/idlisseus/openrouter.json"


class PlanError(ValueError):
    pass


class ModelRequestError(RuntimeError):
    pass


def dump(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, ensure_ascii=False) + "\n")


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    digest.update(path.read_bytes())
    return digest.hexdigest()


def clean(value: Any) -> Any:
    if pd.isna(value):
        return None
    value = value.item() if hasattr(value, "item") else value
    return None if isinstance(value, float) and not math.isfinite(value) else value


def profile(frame: pd.DataFrame, include_values: bool) -> dict[str, Any]:
    columns = []
    for name in frame.columns:
        series = frame[name]
        item: dict[str, Any] = {
            "name": str(name), "type": str(series.dtype),
            "null_count": int(series.isna().sum()),
            "distinct_count": int(series.nunique(dropna=True)),
        }
        if pd.api.types.is_numeric_dtype(series) and len(series.dropna()):
            item.update(minimum=clean(series.min()), maximum=clean(series.max()))
        elif include_values and item["distinct_count"] <= 30:
            item["admitted_values"] = [clean(v) for v in series.dropna().drop_duplicates().head(30)]
        columns.append(item)
    return {"row_count": len(frame), "columns": columns}


def safe_column_name(label: str) -> str:
    return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", label.lower())).strip("_")


def load_digital_pdf(source: Path, include_values: bool) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]]:
    """Extract a simple year-column table while retaining page/table citations.

    This is deliberately narrow. Scans, spanning headers, multiple tables and
    ambiguous regions belong to the later structure-preserving extraction path.
    """
    completed = subprocess.run(
        ["pdftotext", "-layout", str(source), "-"],
        check=True, capture_output=True, text=True, timeout=30,
    )
    pages = completed.stdout.split("\f")
    rows: list[dict[str, Any]] = []
    table_pages: set[int] = set()
    table_labels: set[str] = set()
    source_label = source.name
    document_notes = "\n".join(page.strip() for page in pages if page.strip())
    label_match = re.search(r"Source label:\s*([^\n.]+)", document_notes, re.I)
    if label_match:
        source_label = label_match.group(1).strip().rstrip(".")

    for page_number, page in enumerate(pages, start=1):
        lines = [line.rstrip() for line in page.splitlines()]
        for index, line in enumerate(lines):
            year_matches = list(re.finditer(r"((?:19|20)\d{2})(?:\s*\(%\))?", line))
            header_years = [int(match.group(1)) for match in year_matches]
            if len(header_years) < 2:
                continue
            dimension_name = safe_column_name(line[:year_matches[0].start()].strip()) or "category"
            title = next((prior.strip() for prior in reversed(lines[:index]) if prior.strip()), "Table")
            table_match = re.match(r"(Table\s+[^.]+)\.\s*(.+)", title, re.I)
            table_label = table_match.group(1) if table_match else "Table"
            metric_label = table_match.group(2) if table_match else title
            metric_label = re.sub(rf"\s+by\s+{re.escape(dimension_name.replace('_', ' '))}.*$", "", metric_label, flags=re.I)
            metric_name = safe_column_name(metric_label)
            if "%" in line or any(token in metric_label.lower() for token in ("percent", "percentage", "coverage", "rate")):
                metric_name = re.sub(r"_(?:pct|percentage|percent)$", "", metric_name) + "_percent"
            value_pattern = r"^\s*([^\d]+?)\s+((?:-?\d+(?:\.\d+)?\s+){%d}-?\d+(?:\.\d+)?)\s*$" % (len(header_years) - 1)
            for data_line in lines[index + 1:]:
                stripped = data_line.strip()
                if not stripped:
                    continue
                if stripped.lower().startswith("source:"):
                    break
                match = re.match(value_pattern, data_line)
                if not match:
                    continue
                values = [float(value) for value in re.findall(r"-?\d+(?:\.\d+)?", match.group(2))]
                if len(values) != len(header_years):
                    continue
                for year, value in zip(header_years, values):
                    rows.append({
                        dimension_name: match.group(1).strip(), "year": year, metric_name: value,
                        "source_page": page_number, "source_table": table_label,
                        "source": source_label,
                    })
                table_pages.add(page_number)
                table_labels.add(table_label)
    if not rows:
        raise PlanError("no supported digital PDF table region found")
    frame = pd.DataFrame(rows)
    metadata = {
        "format": "pdf", "pages": len([page for page in pages if page.strip()]),
        "extracted_pages": sorted(table_pages), "tables": sorted(table_labels),
        "adapter_scope": "digital PDF with one categorical row label and repeated year columns",
    }
    data_profile = profile(frame, include_values)
    data_profile["document"] = {**metadata, "notes": document_notes[:5000] if include_values else None}
    return frame, data_profile, metadata


def load_structured_xlsx_regions(source: Path, include_values: bool) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]] | None:
    """Extract compatible merged-heading regions without dataset-name rules.

    Region discovery uses workbook geometry and value types: a merged title,
    a two-row header with year leaves, a categorical first column and numeric
    measure cells. It supports vertical and horizontal regions, records exact
    coordinates, and refuses incompatible region families. No case id, topic,
    expected answer or dataset-specific column name participates in parsing.
    """
    workbook = load_workbook(source, data_only=True)
    merged_ranges = {sheet.title: [str(item) for item in sheet.merged_cells.ranges] for sheet in workbook.worksheets}
    if not any(merged_ranges.values()):
        return None

    candidates: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        expanded: dict[tuple[int, int], Any] = {}
        for merged in sheet.merged_cells.ranges:
            value = sheet.cell(merged.min_row, merged.min_col).value
            for row in range(merged.min_row, merged.max_row + 1):
                for column in range(merged.min_col, merged.max_col + 1):
                    expanded[(row, column)] = value

        def cell(row: int, column: int) -> Any:
            return expanded.get((row, column), sheet.cell(row, column).value)

        for title_range in sheet.merged_cells.ranges:
            if title_range.min_row != title_range.max_row or title_range.max_col <= title_range.min_col:
                continue
            title_row = title_range.min_row
            start_column, end_column = title_range.min_col, title_range.max_col
            title = str(sheet.cell(title_row, start_column).value or "").strip()
            if not title:
                continue
            header_row = next((
                row for row in range(title_row + 1, min(sheet.max_row, title_row + 5) + 1)
                if sum(cell(row, column) not in (None, "") for column in range(start_column, end_column + 1)) >= 2
            ), None)
            if header_row is None or header_row + 1 > sheet.max_row:
                continue
            dimension_merge = next((merged for merged in sheet.merged_cells.ranges
                                    if merged.min_row == header_row and merged.min_col == start_column
                                    and merged.max_row >= header_row + 1), None)
            if dimension_merge is None:
                continue
            dimension_label = str(cell(header_row, start_column) or "category").strip()
            dimension_name = safe_column_name(dimension_label) or "category"
            measures: list[dict[str, Any]] = []
            for column in range(start_column + 1, end_column + 1):
                group = str(cell(header_row, column) or "").strip()
                leaf = str(cell(header_row + 1, column) or "").strip()
                year_match = re.search(r"(?:19|20)\d{2}", f"{group} {leaf}")
                if not year_match:
                    continue
                metric_label = group if not re.fullmatch(r"(?:19|20)\d{2}", group) else title
                metric = safe_column_name(metric_label)
                if any(token in metric_label.lower() for token in ("%", "percent", "percentage", "coverage", "rate")):
                    metric = re.sub(r"_(?:pct|percentage|percent)$", "", metric) + "_percent"
                if metric:
                    measures.append({
                        "column": column, "base_metric": metric,
                        "label": metric_label, "year": int(year_match.group()),
                    })
            if not measures:
                continue
            data_start = header_row + 2
            data_end = data_start - 1
            data: list[dict[str, Any]] = []
            for row in range(data_start, sheet.max_row + 1):
                category = cell(row, start_column)
                if category in (None, ""):
                    break
                row_values = []
                for measure in measures:
                    value = cell(row, measure["column"])
                    if not isinstance(value, (int, float)) or isinstance(value, bool):
                        continue
                    row_values.append({
                        "base_metric": measure["base_metric"], "label": measure["label"],
                        "year": measure["year"], "value": float(value),
                    })
                if not row_values:
                    break
                data_end = row
                data.append({"category": str(category).strip(), "values": row_values})
            if data_end < data_start:
                continue
            table_match = re.match(r"(Table\s+\d+)[.:\s-]+(.+)", title, re.I)
            table_label = table_match.group(1).title() if table_match else title
            table_title = table_match.group(2).strip() if table_match else title
            region_range = (
                f"{sheet.title}!{get_column_letter(start_column)}{title_row}:"
                f"{get_column_letter(end_column)}{data_end}"
            )
            candidates.append({
                "sheet": sheet.title, "table": table_label, "title": table_title,
                "title_row": title_row, "header_rows": [header_row, header_row + 1],
                "data_rows": [data_start, data_end], "range": region_range,
                "dimension_name": dimension_name, "data": data,
                "measures": measures,
            })
    if not candidates:
        return None

    dimension_names = {candidate["dimension_name"] for candidate in candidates}
    if len(dimension_names) != 1:
        raise PlanError(
            f"ambiguous merged-table families use different row dimensions: {sorted(dimension_names)}"
        )
    dimension_name = next(iter(dimension_names))
    category_sets = [{row["category"] for row in candidate["data"]} for candidate in candidates]
    for left_index, left in enumerate(category_sets):
        for right in category_sets[left_index + 1:]:
            overlap = len(left & right) / max(1, min(len(left), len(right)))
            if overlap < 0.5:
                raise PlanError("ambiguous merged-table families have incompatible category values")

    metric_regions: dict[str, list[int]] = {}
    for index, candidate in enumerate(candidates):
        for metric in {measure["base_metric"] for measure in candidate["measures"]}:
            metric_regions.setdefault(metric, []).append(index)

    title_tokens = [set(re.findall(r"[a-z]+", candidate["title"].lower())) for candidate in candidates]
    common_title_tokens = set.intersection(*title_tokens) if title_tokens else set()
    generic_tokens = {
        "table", "by", "and", "of", "the", "year", "report",
        *re.findall(r"[a-z]+", dimension_name.lower()),
    }

    def qualified_metric(base: str, candidate_index: int) -> str:
        if len(set(metric_regions.get(base, []))) <= 1:
            return base
        measure_tokens = set(base.split("_"))
        distinctive = [token for token in re.findall(r"[a-z]+", candidates[candidate_index]["title"].lower())
                       if token not in common_title_tokens | generic_tokens | measure_tokens]
        if not distinctive:
            raise PlanError(f"cannot distinguish repeated metric {base} across merged regions")
        return f"{distinctive[0]}_{base}"

    observations: dict[tuple[str, int], dict[str, Any]] = {}
    definitions: list[dict[str, Any]] = []
    for candidate_index, candidate in enumerate(candidates):
        for row in candidate["data"]:
            for value in row["values"]:
                metric = qualified_metric(value["base_metric"], candidate_index)
                item = observations.setdefault(
                    (row["category"], value["year"]),
                    {dimension_name: row["category"], "year": value["year"]},
                )
                if metric in item and item[metric] != value["value"]:
                    raise PlanError(f"conflicting values for {metric}, {row['category']}, {value['year']}")
                item[metric] = value["value"]
                if not any(definition["indicator"] == metric for definition in definitions):
                    definitions.append({
                        "indicator": metric, "definition": value["label"],
                        "formula": "reported value; no recomputation",
                        "unit": "percent" if metric.endswith("_percent") else "number",
                        "sheet": candidate["sheet"], "table": candidate["table"],
                        "range": candidate["range"],
                    })

    regions = [{key: value for key, value in candidate.items() if key not in ("data", "measures")}
               for candidate in candidates]
    tables = list(dict.fromkeys(candidate["table"] for candidate in candidates))
    ranges = list(dict.fromkeys(candidate["range"] for candidate in candidates))
    frame = pd.DataFrame(observations.values()).sort_values([dimension_name, "year"]).reset_index(drop=True)
    frame["source_sheet"] = "; ".join(dict.fromkeys(region["sheet"] for region in regions))
    frame["source_table"] = "; ".join(dict.fromkeys(tables))
    frame["source_range"] = "; ".join(dict.fromkeys(ranges))
    frame["source"] = source.name
    metadata = {
        "format": "xlsx", "sheets": workbook.sheetnames, "selected_sheet": None,
        "definition_sheets": [], "definitions": definitions,
        "tables": list(dict.fromkeys(tables)), "ranges": list(dict.fromkeys(ranges)),
        "candidate_regions": regions, "merged_ranges": merged_ranges,
        "adapter_scope": "compatible vertical or horizontal regions with merged titles, two-row year headers and numeric bodies",
    }
    data_profile = profile(frame, include_values)
    data_profile["workbook"] = metadata
    return frame, data_profile, metadata


def load_source(source: Path, include_values: bool) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]]:
    if source.suffix.lower() == ".csv":
        frame = pd.read_csv(source)
        metadata = {"format": "csv", "sheets": [], "selected_sheet": None, "definitions": []}
        return frame, profile(frame, include_values), metadata
    if source.suffix.lower() == ".pdf":
        return load_digital_pdf(source, include_values)
    if source.suffix.lower() != ".xlsx":
        raise PlanError(f"unsupported input format: {source.suffix}")
    structured = load_structured_xlsx_regions(source, include_values)
    if structured is not None:
        return structured
    workbook = load_workbook(source, data_only=True)
    if any(sheet.merged_cells.ranges for sheet in workbook.worksheets):
        raise PlanError(
            "workbook has merged layout not covered by the structural adapter; route to the general layout extraction fallback"
        )
    sheets = pd.read_excel(source, sheet_name=None)
    if not sheets:
        raise PlanError("workbook contains no readable sheets")

    def observation_score(item: tuple[str, pd.DataFrame]) -> tuple[int, int, int]:
        _name, candidate = item
        numeric = sum(pd.api.types.is_numeric_dtype(candidate[column]) for column in candidate.columns)
        non_numeric = sum(not pd.api.types.is_numeric_dtype(candidate[column]) for column in candidate.columns)
        return int(numeric > 0 and non_numeric > 0), numeric, len(candidate)

    plausible = [item for item in sheets.items()
                 if len(item[1]) >= 1 and observation_score(item)[0] == 1]
    if len(plausible) > 1:
        raise PlanError(
            f"workbook has multiple plausible rectangular data sheets {[name for name, _ in plausible]}; route to the general sheet-selection fallback"
        )
    if not plausible:
        raise PlanError(
            "workbook has no unambiguous rectangular observation table; route to the general sheet-selection fallback"
        )
    selected_sheet, frame = plausible[0]
    frame = frame.copy()
    definitions: list[dict[str, Any]] = []
    source_labels: list[str] = []
    for sheet_name, sheet in sheets.items():
        if "source" in sheet.columns:
            source_labels += [str(value) for value in sheet["source"].dropna().drop_duplicates()]
        if sheet_name != selected_sheet and include_values:
            for row in sheet.head(100).to_dict("records"):
                definitions.append({"sheet": sheet_name, **{str(key): clean(value) for key, value in row.items()}})
    frame["source_sheet"] = selected_sheet
    frame["source"] = source_labels[0] if source_labels else source.name
    metadata = {
        "format": "xlsx", "sheets": list(sheets), "selected_sheet": selected_sheet,
        "definition_sheets": [name for name in sheets if name != selected_sheet],
        "definitions": definitions,
    }
    data_profile = profile(frame, include_values)
    data_profile["workbook"] = metadata
    return frame, data_profile, metadata


def make_prompt(case: dict[str, Any], data_profile: dict[str, Any], messages: list[str], previous: Any) -> str:
    rules = [
        "Plan a safe analysis of one local table. Never calculate or invent values.",
        "Use supplied column names exactly. The executor applies steps in order.",
        "Use a supplied measure directly. Derive a percent only from distinct numerator and denominator count columns, never from a supplied rate or percentage.",
        "When comparing groups on an outcome count and the schema offers a plausible eligible/base count, do not compare raw outcome counts alone unless the participant explicitly asks for counts. Derive the comparable rate and retain its underlying counts; if the numerator/denominator pairing is genuinely ambiguous, use clarification instead of silently choosing.",
        "The participant's actual words are authoritative. Do not strengthen or rewrite ambiguous wording in question, title or note to pretend that the participant explicitly chose one measure.",
        "A subtraction of percentage measures is percentage points; a subtraction of counts is a number.",
        "Use change for one entity across two times and difference for two entities at one time.",
        "Keep rows unless aggregation is requested. Never invent bands, targets or domain facts.",
        "For a requested webpage selector, retain the selectable rows; the webpage adds controls. Never emit UI placeholders as data values.",
        "Do not group when there is already one row per entity and time. If aggregation is needed for a rate, use weighted_percent rather than sum or mean of a percent.",
        "The view series is a categorical grouping, never a numeric metric. Put measures only in view y. For time trends use time on x and an entity category as series.",
        "When several measures are requested, retain them in view y and use grouped_bar; the webpage presents one measure at a time through its indicator selector, so unlike units never share an axis.",
        "Preserve relevant earlier requests and scope across follow-ups until the participant changes them.",
        "Source and provenance are rendered separately; do not group or sort merely to display the source.",
        "Write question, title and note for a nontechnical participant. Never mention renderer, SQL, JSON, DuckDB, placeholders or other implementation details.",
        "If data cannot explain why, set can_explain_cause false, add causal_limit, and say in the view note that the file cannot explain the cause. If an intervention is requested but unsupported, also say in the note that the file cannot recommend an intervention.",
        "When the participant asks a direct analytical question, add the corresponding insight so the answer is stated rather than left for them to infer.",
        "Use line/slope for time, bars for categories, scatter for two measures.",
        "Return only one JSON object matching the response schema.",
    ]
    sections = {
        "DATASET_ID": case["case_id"], "DATASET_PROFILE": data_profile,
        "INPUT_PROVENANCE": case["inputs"], "CONVERSATION_SO_FAR": messages,
        "PREVIOUS_VALID_PLAN": previous,
    }
    return "\n".join(rules) + "\n\n" + "\n".join(
        f"{key}:\n{json.dumps(value, ensure_ascii=False)}" for key, value in sections.items()
    )


def repair_prompt(prompt: str, plan: Any, error: Exception) -> str:
    return prompt + "\n\nREJECTED_PLAN:\n" + json.dumps(plan, ensure_ascii=False) + \
        "\nVALIDATOR_ERROR:\n" + str(error) + \
        "\nReturn a complete corrected plan. Audit the whole corrected plan against every rule above, not only the reported error. Recheck durable filters, named-entity scope, insight kind and unit, and that view series differs from view x. Do not explain the correction."


def call_openrouter_json(model: str, effort: str, prompt: str, schema: dict[str, Any],
                         timeout_seconds: int, system: str, schema_name: str) -> tuple[dict[str, Any], dict[str, Any]]:
    credential = json.loads(KEY_PATH.read_text())["api_key"]
    document = {
        "model": model,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0,
        "reasoning": {"effort": effort},
        "temperature": 0,
        "response_format": {"type": "json_schema", "json_schema": {
            "name": schema_name, "strict": True, "schema": schema,
        }},
    }
    request = urllib.request.Request(
        "https://openrouter.ai/api/v1/chat/completions",
        data=json.dumps(document).encode(), method="POST",
        headers={"Authorization": f"Bearer {credential}", "Content-Type": "application/json",
                 "HTTP-Referer": "https://github.com/bprashanth/io",
                 "X-Title": "NGO split pipeline benchmark"},
    )
    started = time.monotonic()
    def timed_out(_signum: int, _frame: Any) -> None:
        raise TimeoutError(f"model response exceeded {timeout_seconds} seconds")
    previous_handler = signal.signal(signal.SIGALRM, timed_out)
    signal.setitimer(signal.ITIMER_REAL, timeout_seconds)
    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            raw = json.load(response)
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, previous_handler)
    duration = round(time.monotonic() - started, 3)
    choices = raw.get("choices")
    if not choices:
        error = raw.get("error") or {"message": "response contained no choices"}
        raise ModelRequestError(json.dumps(error, ensure_ascii=False))
    content = choices[0]["message"].get("content") or ""
    match = re.search(r"\{.*\}", content, flags=re.DOTALL)
    if not match:
        raise PlanError("model returned no JSON object")
    return json.loads(match.group()), {
        "requested_model": model, "resolved_model": raw.get("model"),
        "generation_id": raw.get("id"), "duration_seconds": duration,
        "finish_reason": raw["choices"][0].get("finish_reason"),
        "usage": raw.get("usage"), "reasoning_effort": effort, "temperature": 0,
    }


def call_model(model: str, effort: str, prompt: str, schema: dict[str, Any], timeout_seconds: int) -> tuple[dict[str, Any], dict[str, Any]]:
    return call_openrouter_json(
        model, effort, prompt, schema, timeout_seconds,
        "You are a constrained tabular analysis-plan compiler.", "analysis_plan",
    )


def make_critic_prompt(messages: list[str], data_profile: dict[str, Any], previous: Any,
                       candidate: dict[str, Any], result: pd.DataFrame) -> str:
    """Build the value-redacted, domain-neutral semantic-review request."""
    def without_values(source_profile: dict[str, Any]) -> dict[str, Any]:
        redacted = {
            "row_count": source_profile["row_count"],
            "columns": [
                {key: column[key] for key in ("name", "type", "null_count", "distinct_count")}
                for column in source_profile["columns"]
            ],
        }
        for structural_key in ("workbook", "document"):
            if structural_key in source_profile:
                structural = json.loads(json.dumps(source_profile[structural_key]))
                structural.pop("notes", None)
                redacted[structural_key] = structural
        return redacted

    return "\n".join([
        "Review the candidate analysis plan against the whole conversation.",
        "Judge meaning, not exact wording. This must work across sectors and datasets.",
        "The participant conversation is authoritative. Treat candidate.question, title and note as untrusted paraphrases: they cannot turn an unstated or ambiguous preference into an explicit participant request.",
        "Before deciding, write conversation_intent using only the participant conversation and explicitly mark material ambiguities. Then write candidate_audit comparing the executable candidate against that ledger. Do not use candidate wording to revise the ledger.",
        "Accept only if the executable plan preserves relevant earlier requests and fully answers the newest request.",
        "Check requested regions/measures, filters, comparisons, ranking, units, chart suitability, provenance wording and unsupported causal claims.",
        "For group comparisons, reject a raw-outcome-count-only plan when the schema offers a plausible eligible/base denominator and the participant did not explicitly request counts. Require a comparable rate with underlying counts retained, or a plain-language clarification if the pairing is genuinely ambiguous.",
        "The trusted webpage automatically adds client-side selectors for categorical and year columns retained in the executed result, and an indicator selector when view.y has multiple measures. Do not demand a data-filter step merely to create those controls.",
        "The trusted webpage also always shows the current result table, source provenance and a download button for the current rows. These are renderer features, not analysis-plan steps; do not demand steps or invented schema fields for them.",
        "A grouped_bar view is valid for either one measure or several comparable measures. Do not demand a chart type outside the supplied plan schema.",
        "A filter step changes the durable data scope; UI placeholder values are invalid data values.",
        "Do not calculate or propose data values. Do not demand anything the participant did not ask for.",
        "If rejecting, give one concise but complete repair instruction listing every missing or contradictory obligation.",
        f"CONVERSATION:\n{json.dumps(messages, ensure_ascii=False)}",
        f"DATASET_SCHEMA_AND_STRUCTURE_NO_VALUES:\n{json.dumps(without_values(data_profile), ensure_ascii=False)}",
        f"PREVIOUS_VALID_PLAN:\n{json.dumps(previous, ensure_ascii=False)}",
        f"CANDIDATE_PLAN:\n{json.dumps(candidate, ensure_ascii=False)}",
        f"EXECUTED_RESULT_SCHEMA_NO_VALUES:\n{json.dumps(without_values(profile(result, False)), ensure_ascii=False)}",
    ])


def call_critic(model: str, effort: str, messages: list[str], data_profile: dict[str, Any],
                previous: Any, candidate: dict[str, Any], result: pd.DataFrame,
                timeout_seconds: int) -> tuple[dict[str, Any], dict[str, Any], str]:
    """Ask a separate general-language pass whether the executable plan is complete."""
    schema = {
        "type": "object", "additionalProperties": False,
        "required": ["conversation_intent", "candidate_audit", "accepted", "feedback"],
        "properties": {
            "conversation_intent": {"type": "string"},
            "candidate_audit": {"type": "string"},
            "accepted": {"type": "boolean"},
            "feedback": {"type": "string"},
        },
    }
    prompt = make_critic_prompt(messages, data_profile, previous, candidate, result)
    try:
        review, api = call_openrouter_json(
            model, effort, prompt, schema, timeout_seconds,
            "You are an independent, domain-neutral reviewer of a structured data-analysis plan.",
            "plan_review",
        )
    except PlanError as error:
        raise ModelRequestError(f"critic returned invalid structured output: {error}") from error
    return review, api, prompt


def quoted(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def bind(plan: dict[str, Any], frame: pd.DataFrame) -> None:
    available = set(map(str, frame.columns))
    units = {
        str(column): "percent"
        for column in frame.columns
        if any(token in str(column).lower() for token in ("percent", "pct", "rate", "coverage"))
    }
    for index, step in enumerate(plan["steps"]):
        op = step["op"]
        needed = []
        if op == "filter": needed = [step["column"]]
        elif op == "derive": needed = [step["numerator"], step["denominator"]]
        elif op == "group":
            needed = step["by"] + [a[k] for a in step["aggregates"] for k in
                                    ("column", "weight_column", "denominator_column") if a.get(k)]
        elif op == "sort": needed = [step["column"]]
        missing = [name for name in needed if name not in available]
        if missing:
            raise PlanError(f"step {index} names missing columns: {missing}")
        if op == "derive":
            if step["name"] in available: raise PlanError("derived column already exists")
            if step["numerator"] == step["denominator"]:
                raise PlanError("a percentage numerator and denominator must be different columns")
            if step["kind"] == "percent" and (
                units.get(step["numerator"]) == "percent" or units.get(step["denominator"]) == "percent"
            ):
                raise PlanError("cannot derive a percentage from an already-percent input; use the reported metric directly")
            available.add(step["name"])
            units[step["name"]] = step["unit"]
        elif op == "group":
            if step["by"] and not frame.duplicated(subset=step["by"]).any():
                raise PlanError("grouping keys are already unique; keep the observation rows and provenance")
            available = set(step["by"]) | {a["name"] for a in step["aggregates"]}
        if op == "filter" and isinstance(step["value"], str) and ("{{" in step["value"] or "}}" in step["value"]):
            raise PlanError("UI placeholders are not data filter values")
    references = [plan["view"]["x"], *plan["view"]["y"]]
    if plan["view"].get("series"): references.append(plan["view"]["series"])
    if plan["view"].get("series") in plan["view"]["y"]:
        raise PlanError("view series must be a category column, not a y metric")
    if plan["view"].get("series") == plan["view"]["x"]:
        raise PlanError("view series must not duplicate the x column")
    if plan["view"]["chart"] in ("line", "slope") and plan["view"]["x"] in frame.columns:
        x_name = plan["view"]["x"]
        time_named = any(token in x_name.lower() for token in ("year", "date", "time", "month", "quarter"))
        if not time_named and not pd.api.types.is_numeric_dtype(frame[x_name]):
            raise PlanError("line and slope charts require an ordered numeric or time x column")
        if len(plan["view"]["y"]) != 1:
            raise PlanError("line and slope charts currently require exactly one y metric")
    participant_text = " ".join(str(value or "") for value in (
        plan["question"], plan["view"]["title"], plan["view"].get("note")))
    if re.search(r"\b(renderer|duckdb|sql|json|placeholder)\b", participant_text, flags=re.IGNORECASE):
        raise PlanError("participant-facing text contains implementation jargon")
    if re.search(r"\b(?:x|xx|tbd|todo)\b", participant_text, flags=re.IGNORECASE):
        raise PlanError("participant-facing text contains an unfinished placeholder")
    for insight in plan["insights"]:
        references += [insight[k] for k in ("metric", "label_column", "entity_column", "time_column") if insight.get(k)]
        if insight["kind"] in ("change", "difference"):
            metric_unit = units.get(insight["metric"], "number")
            expected_unit = "percentage points" if metric_unit == "percent" else "number"
            if insight["unit"] != expected_unit:
                raise PlanError(
                    f"{insight['metric']} differences require unit {expected_unit}, got {insight['unit']}"
                )
    missing = [name for name in references if name not in available]
    if missing: raise PlanError(f"view/insight names missing columns: {missing}")
    admitted = {str(c): set(frame[c].dropna().astype(str)) for c in frame.columns
                if not pd.api.types.is_numeric_dtype(frame[c])}
    for insight in plan["insights"]:
        column = insight.get("entity_column")
        for entity in insight.get("entities", []):
            if column in admitted and str(entity) not in admitted[column]:
                raise PlanError(f"absent {column} value: {entity}")


def aggregate(item: dict[str, Any]) -> str:
    name, column = quoted(item["name"]), quoted(item["column"])
    if item["fn"] == "weighted_percent":
        denominator = item.get("denominator_column")
        if not denominator: raise PlanError("weighted_percent needs denominator_column")
        return f"100.0*SUM({column})/NULLIF(SUM({quoted(denominator)}),0) AS {name}"
    function = {"mean": "AVG", "count": "COUNT"}.get(item["fn"], item["fn"].upper())
    return f"{function}({column}) AS {name}"


def execute(frame: pd.DataFrame, plan: dict[str, Any]) -> pd.DataFrame:
    db = duckdb.connect(":memory:")
    db.register("input_frame", frame)
    sql, params = "SELECT * FROM input_frame", []
    for step in plan["steps"]:
        op = step["op"]
        if op == "derive":
            a, b = quoted(step["numerator"]), quoted(step["denominator"])
            expression = (f"100.0*{a}/NULLIF({b},0)" if step["kind"] == "percent" else
                          f"{a}/NULLIF({b},0)" if step["kind"] == "ratio" else f"{a}-{b}")
            sql = f"SELECT *,{expression} AS {quoted(step['name'])} FROM ({sql}) q"
        elif op == "filter":
            column, cmp, value = quoted(step["column"]), step["cmp"], step["value"]
            if cmp == "in":
                if not isinstance(value, list): raise PlanError("in needs an array")
                condition = f"{column} IN ({','.join('?' for _ in value)})"; params += value
            elif cmp == "between":
                if not isinstance(value, list) or len(value) != 2: raise PlanError("between needs two values")
                condition = f"{column} BETWEEN ? AND ?"; params += value
            elif cmp == "contains":
                condition = f"LOWER(CAST({column} AS VARCHAR)) LIKE LOWER(?)"; params += [f"%{value}%"]
            elif value is None and cmp in ("eq", "ne"):
                condition = f"{column} IS {'NOT ' if cmp == 'ne' else ''}NULL"
            else:
                operator = {"eq": "=", "ne": "!=", "lt": "<", "le": "<=", "gt": ">", "ge": ">="}[cmp]
                condition = f"{column} {operator} ?"; params += [value]
            sql = f"SELECT * FROM ({sql}) q WHERE {condition}"
        elif op == "group":
            by = ",".join(map(quoted, step["by"])); selections = ([by] if by else []) + list(map(aggregate, step["aggregates"]))
            sql = f"SELECT {','.join(selections)} FROM ({sql}) q" + (f" GROUP BY {by}" if by else "")
        elif op == "sort": sql = f"SELECT * FROM ({sql}) q ORDER BY {quoted(step['column'])} {step['direction'].upper()}"
        elif op == "limit": sql = f"SELECT * FROM ({sql}) q LIMIT {int(step['rows'])}"
    try: result = db.execute(sql, params).fetchdf()
    except Exception as error: raise PlanError(f"DuckDB failed: {error}") from error
    finally: db.close()
    if result.empty: raise PlanError("plan produced no rows")
    return result


def explicit_year_constraint(messages: list[str]) -> tuple[set[int] | None, str]:
    allowed: set[int] | None = None
    constraint_kind = ""
    for message in messages:
        lowered = message.lower()
        ranges = re.findall(r"only\s+((?:19|20)\d{2})\s+(?:to|through|-)\s+((?:19|20)\d{2})", lowered)
        explicit_sets = re.findall(
            r"only\s+((?:19|20)\d{2}(?:\s*(?:,|and)\s*(?:19|20)\d{2})+)", lowered
        )
        singles = re.findall(
            r"only\s+((?:19|20)\d{2})(?!\s*(?:to|through|-|,|and))", lowered
        )
        if ranges:
            start, end = map(int, ranges[-1]); allowed = set(range(start, end + 1)); constraint_kind = "range"
        elif explicit_sets:
            allowed = {int(year) for year in re.findall(r"(?:19|20)\d{2}", explicit_sets[-1])}; constraint_kind = "set"
        elif singles:
            allowed = {int(singles[-1])}; constraint_kind = "single"
        else:
            direct = re.findall(r"\b(?:show|filter)\s+(?:only\s+)?((?:19|20)\d{2})\b", lowered)
            comparison_year = re.findall(
                r"(?:\bfor\s+|\bin\s+)((?:19|20)\d{2})\b", lowered
            ) if "compar" in lowered else []
            if direct:
                allowed = {int(direct[-1])}; constraint_kind = "direct"
            elif comparison_year and allowed is None:
                allowed = {int(comparison_year[-1])}; constraint_kind = "comparison"
    return allowed, constraint_kind


def normalize_explicit_constraints(messages: list[str], plan: dict[str, Any], frame: pd.DataFrame) -> list[str]:
    """Compile literal user scope into filters before semantic validation."""
    required: dict[str, list[Any]] = {}
    changes: list[str] = []
    cleaned_steps = []
    for step in plan["steps"]:
        value = step.get("value")
        values = value if isinstance(value, list) else [value]
        placeholder = step["op"] == "filter" and any(
            isinstance(item, str) and re.search(r"(?:^user_selected_|^selected_|placeholder|\{\{)", item, re.I)
            for item in values
        )
        if placeholder:
            changes.append(f"removed placeholder filter on {step['column']}")
        else:
            cleaned_steps.append(step)
    plan["steps"] = cleaned_steps
    allowed_years, _ = explicit_year_constraint(messages)
    year_columns = [str(column) for column in frame.columns if "year" in str(column).lower()]
    if allowed_years is not None and year_columns:
        required[year_columns[0]] = sorted(allowed_years)

    conversation = "\n".join(messages).lower()
    if "lakh" in conversation:
        lakh_metrics = [metric for metric in plan["view"]["y"] if metric.endswith("_lakh")]
        paired_lakh = next(
            (metric for metric in lakh_metrics
             if metric.removesuffix("_lakh") in plan["view"]["y"]
             and metric.removesuffix("_lakh") in frame.columns),
            None,
        )
        if paired_lakh:
            plan["view"]["y"] = [paired_lakh]
            plan["view"]["unit"] = "lakh"
            changes.append(
                f"used {paired_lakh} for the scaled visual and kept "
                f"{paired_lakh.removesuffix('_lakh')} as an exact table column"
            )
    insight_metric = plan["view"]["y"][0]
    if "exact" in conversation and insight_metric.endswith("_lakh"):
        raw_metric = insight_metric.removesuffix("_lakh")
        if raw_metric in frame.columns:
            insight_metric = raw_metric
    ranking_kind = None
    if re.search(r"\b(?:highest|largest)\b", conversation):
        ranking_kind = "highest"
    elif re.search(r"\b(?:lowest|smallest)\b", conversation):
        ranking_kind = "lowest"
    if ranking_kind and not any(item["kind"] == ranking_kind for item in plan["insights"]):
        label_column = plan["view"]["x"]
        metric = insight_metric
        if label_column in frame.columns and not pd.api.types.is_numeric_dtype(frame[label_column]):
            plan["insights"].append({
                "kind": ranking_kind, "metric": metric, "label_column": label_column,
            })
            changes.append(f"compiled explicit {ranking_kind} request")

    difference_message = next(
        (message for message in reversed(messages) if re.search(
            r"\b(?:exact\s+)?difference\s+between\b|\bgap\s+between\b", message, re.I
        )),
        None,
    )
    if difference_message and not any(item["kind"] == "difference" for item in plan["insights"]):
        entity_match = None
        for column in frame.columns:
            if pd.api.types.is_numeric_dtype(frame[column]):
                continue
            admitted = [str(value) for value in frame[column].dropna().unique()]
            named = [value for value in admitted if re.search(
                rf"(?<![\w]){re.escape(value.lower())}(?![\w])", difference_message.lower()
            )]
            if len(named) == 2:
                entity_match = (str(column), named)
                break
        time_columns = [str(column) for column in frame.columns if "year" in str(column).lower()]
        years = [int(value) for value in re.findall(r"\b(?:19|20)\d{2}\b", difference_message)]
        if entity_match and time_columns and (years or frame[time_columns[0]].nunique(dropna=True) == 1):
            time_value = years[-1] if years else int(frame[time_columns[0]].dropna().iloc[0])
            plan["insights"].append({
                "kind": "difference", "metric": insight_metric,
                "entity_column": entity_match[0], "entities": entity_match[1],
                "time_column": time_columns[0], "time": time_value, "unit": "number",
            })
            changes.append("compiled explicit named-entity difference request")

    ranking = any(item["kind"] in ("highest", "lowest") for item in plan["insights"])
    comparison_message = next(
        (message.lower() for message in reversed(messages)
         if re.search(r"\bcompar(?:e|ing|ison)\b|\bvs\.?\b|\bversus\b", message, re.I)),
        None,
    )
    if comparison_message and not ranking:
        for column in frame.columns:
            if pd.api.types.is_numeric_dtype(frame[column]):
                continue
            admitted = [str(value) for value in frame[column].dropna().unique()]
            named = [value for value in admitted if re.search(
                rf"(?<![\w]){re.escape(value.lower())}(?![\w])", comparison_message
            )]
            if len(named) == 2:
                required[str(column)] = named

    percent_columns = {
        str(column) for column in frame.columns
        if any(token in str(column).lower() for token in ("percent", "pct", "rate", "coverage"))
    }
    for insight in plan["insights"]:
        if insight["kind"] not in ("change", "difference"):
            continue
        expected_unit = "percentage points" if insight["metric"] in percent_columns else "number"
        if insight["unit"] != expected_unit:
            changes.append(f"normalized {insight['kind']} unit for {insight['metric']} to {expected_unit}")
            insight["unit"] = expected_unit

    for column, values in required.items():
        expected = {str(value) for value in values}
        current = [step for step in plan["steps"] if step["op"] == "filter" and step["column"] == column]
        current_values = set()
        for step in current:
            value = step["value"]
            current_values |= {str(item) for item in value} if isinstance(value, list) else {str(value)}
        if current_values != expected or len(current) != 1 or current[0]["cmp"] != "in":
            changes.append(f"enforced {column} in {values}")
    if required:
        remaining = [
            step for step in plan["steps"]
            if not (step["op"] == "filter" and step["column"] in required)
        ]
        plan["steps"] = [
            {"op": "filter", "column": column, "cmp": "in", "value": values}
            for column, values in required.items()
        ] + remaining
    return changes


def normalize_mechanical_invariants(plan: dict[str, Any], frame: pd.DataFrame) -> list[str]:
    """Normalize representation hazards without interpreting the user's domain or prose."""
    changes: list[str] = []
    cleaned_steps = []
    for step in plan["steps"]:
        value = step.get("value")
        values = value if isinstance(value, list) else [value]
        placeholder = step["op"] == "filter" and any(
            isinstance(item, str) and re.search(
                r"(?:^user_selected_|^selected_(?:year|category)|placeholder|\{\{)", item, re.I
            ) for item in values
        )
        if placeholder:
            changes.append(f"removed non-data placeholder filter on {step['column']}")
        else:
            cleaned_steps.append(step)
    plan["steps"] = cleaned_steps

    percent_columns = {
        str(column) for column in frame.columns
        if any(token in str(column).lower() for token in ("percent", "pct", "rate", "coverage"))
    }
    for insight in plan["insights"]:
        if insight["kind"] not in ("change", "difference"):
            continue
        expected = "percentage points" if insight["metric"] in percent_columns else "number"
        if insight["unit"] != expected:
            insight["unit"] = expected
            changes.append(f"normalized arithmetic unit for typed metric {insight['metric']} to {expected}")
    return changes


def check_mechanical_invariants(plan: dict[str, Any], result: pd.DataFrame) -> None:
    """Reject executable contradictions without parsing participant language."""
    fingerprints = [json.dumps(item, sort_keys=True) for item in plan["insights"]]
    if len(fingerprints) != len(set(fingerprints)):
        raise PlanError("duplicate insight specifications would repeat the same answer on the page")
    ranking = [item for item in plan["insights"] if item["kind"] in ("highest", "lowest")]
    comparisons = [item for item in plan["insights"] if item["kind"] in ("change", "difference")]
    for item in comparisons:
        shown = {str(value) for value in result[item["entity_column"]].dropna()}
        requested = {str(value) for value in item["entities"]}
        valid_scope = requested <= shown if ranking else shown == requested
        if not valid_scope:
            raise PlanError(
                f"comparison operands contradict executable row scope: requested {sorted(requested)}, got {sorted(shown)}"
            )
    differences = [item for item in plan["insights"] if item["kind"] == "difference"]
    if ranking and differences:
        ranking_columns = {item["label_column"] for item in ranking}
        narrowed = [step["column"] for step in plan["steps"]
                    if step["op"] == "filter" and step["column"] in ranking_columns]
        if narrowed:
            raise PlanError("ranking population is narrowed by a pairwise-comparison filter")


def check_conversation_constraints(messages: list[str], plan: dict[str, Any], result: pd.DataFrame) -> None:
    """Reject executable plans that silently ignore explicit durable filters.

    This is intentionally small and deterministic. It does not try to understand
    arbitrary language; it protects the benchmark's declared year-only/range
    phrasing. More constraint extractors must be frozen before additional cases.
    """
    text = "\n".join(messages).lower()
    for words, kind in ((("lowest", "smallest"), "lowest"), (("highest", "largest"), "highest")):
        requested_word = next((word for word in words if re.search(rf"\b{word}\b", text)), None)
        if requested_word and not any(item["kind"] == kind for item in plan["insights"]):
            raise PlanError(f"explicit {requested_word} request requires a named {kind} insight")
    insight_fingerprints = [json.dumps(item, sort_keys=True) for item in plan["insights"]]
    if len(insight_fingerprints) != len(set(insight_fingerprints)):
        raise PlanError("duplicate insight specifications would repeat the same answer on the page")
    asks_for_difference = bool(
        re.search(r"\b(?:exact\s+)?difference\s+between\b", text)
        or re.search(r"\bgap\s+between\b", text)
    )
    if asks_for_difference and not any(item["kind"] == "difference" for item in plan["insights"]):
        raise PlanError("explicit difference-between request requires a difference insight")
    asks_for_pp_change = ("percentage point" in text or re.search(r"\bpp\b", text)) and re.search(r"\bchange\b", text)
    if asks_for_pp_change and not any(item["kind"] == "change" for item in plan["insights"]):
        raise PlanError("explicit percentage-point change request requires a change insight")
    ranking = [item for item in plan["insights"] if item["kind"] in ("highest", "lowest")]
    if not ranking:
        comparison_message = next(
            (message.lower() for message in reversed(messages)
             if re.search(r"\bcompar(?:e|ing|ison)\b|\bvs\.?\b|\bversus\b", message, re.I)),
            None,
        )
        if comparison_message:
            for column in result.columns:
                if column not in result or pd.api.types.is_numeric_dtype(result[column]):
                    continue
                admitted = [str(value) for value in result[column].dropna().unique()]
                named = [value for value in admitted if re.search(
                    rf"(?<![\w]){re.escape(value.lower())}(?![\w])", comparison_message
                )]
                if len(named) == 2 and set(admitted) != set(named):
                    raise PlanError(
                        f"explicit comparison of {named} must filter {column} to those two values"
                    )
    differences = [item for item in plan["insights"] if item["kind"] == "difference"]
    comparisons = [item for item in plan["insights"] if item["kind"] in ("change", "difference")]
    for item in comparisons:
        shown = {str(value) for value in result[item["entity_column"]].dropna()}
        requested = {str(value) for value in item["entities"]}
        valid_scope = requested <= shown if ranking else shown == requested
        if not valid_scope:
            raise PlanError(
                f"named-entity comparison scope must retain only {sorted(requested)}, got {sorted(shown)}"
            )
    if ranking and differences:
        ranking_columns = {item["label_column"] for item in ranking}
        narrowed = [step["column"] for step in plan["steps"] if step["op"] == "filter" and step["column"] in ranking_columns]
        if narrowed:
            raise PlanError("a ranking-plus-pairwise-gap turn must retain the ranking population")
    asks_why = bool(re.search(r"\bwhy\b", text) or "don't guess reason" in text or "dont guess reason" in text)
    if asks_why:
        if not any(item["kind"] == "causal_limit" for item in plan["insights"]):
            raise PlanError("causal question requires a visible causal_limit insight")
    year_columns = [str(column) for column in result.columns if "year" in str(column).lower()]
    if not year_columns:
        return
    years = {int(value) for value in result[year_columns[0]].dropna()}
    allowed, constraint_kind = explicit_year_constraint(messages)
    if allowed is not None and years != allowed:
        raise PlanError(
            f"durable year-{constraint_kind} constraint ignored: expected {sorted(allowed)}, got {sorted(years)}"
        )


def compute_insights(plan: dict[str, Any], frame: pd.DataFrame) -> list[dict[str, Any]]:
    rows = [{str(k): clean(v) for k, v in row.items()} for row in frame.to_dict("records")]
    output = []
    for spec in plan["insights"]:
        kind = spec["kind"]
        if kind in ("highest", "lowest"):
            valid = [r for r in rows if isinstance(r.get(spec["metric"]), (int, float))]
            if valid:
                chosen = (max if kind == "highest" else min)(valid, key=lambda r: r[spec["metric"]])
                output.append({"kind": kind, "label": str(chosen[spec["label_column"]]),
                               "value": chosen[spec["metric"]], "metric": spec["metric"]})
        elif kind == "change":
            for entity in spec["entities"]:
                subset = [r for r in rows if str(r.get(spec["entity_column"])) == str(entity)]
                start = next((r for r in subset if str(r.get(spec["time_column"])) == str(spec["from"])), None)
                end = next((r for r in subset if str(r.get(spec["time_column"])) == str(spec["to"])), None)
                if start and end: output.append({"kind": kind, "label": str(entity), "value": end[spec["metric"]]-start[spec["metric"]], "metric": spec["metric"], "unit": spec["unit"], "from": spec["from"], "to": spec["to"]})
        elif kind == "difference":
            subset = [r for r in rows if str(r.get(spec["time_column"])) == str(spec["time"])]
            found = [next((r for r in subset if str(r.get(spec["entity_column"])) == str(e)), None) for e in spec["entities"]]
            if all(found): output.append({"kind": kind, "label": f"Gap between {spec['entities'][0]} and {spec['entities'][1]}", "value": abs(found[1][spec["metric"]]-found[0][spec["metric"]]), "metric": spec["metric"], "unit": spec["unit"], "time": spec["time"]})
        elif kind == "causal_limit":
            output.append({"kind": kind, "label": "What this data cannot answer", "text": "This file shows differences and changes, but it cannot explain the cause or recommend an intervention from this file alone."})
    return output


def make_report(case: dict[str, Any], source: Path, source_metadata: dict[str, Any], plan: dict[str, Any], frame: pd.DataFrame) -> dict[str, Any]:
    rows = [{str(k): clean(v) for k, v in row.items()} for row in frame.to_dict("records")]
    view = dict(plan["view"])
    if any(item["kind"] == "causal_limit" for item in plan["insights"]):
        safety_note = "This file cannot explain the cause or recommend an intervention from this file alone."
        if safety_note.lower() not in str(view.get("note") or "").lower():
            view["note"] = (str(view.get("note") or "").rstrip() + " " + safety_note).strip()
    return {"schema_version": 1, "case_id": case["case_id"], "title": plan["view"]["title"],
            "question": plan["question"], "view": view, "columns": list(map(str, frame.columns)),
            "rows": rows, "insight_specs": plan["insights"], "insights": compute_insights(plan, frame),
            "source": {"file": source.name, "sha256": file_hash(source), "provenance": case["inputs"][0]["provenance"], **source_metadata}}


def write_csv(path: Path, report: dict[str, Any]) -> None:
    def safe(value: Any) -> Any:
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return "'" + value
        return value
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=report["columns"]); writer.writeheader()
        writer.writerows({key: safe(value) for key, value in row.items()} for row in report["rows"])


def render(report: dict[str, Any], output: Path) -> None:
    data = json.dumps(report, ensure_ascii=False).replace("<", "\\u003c")
    title = html.escape(report["title"])
    template_path = ROOT / "benchmarks/event/static-dashboard-template.html"
    template = template_path.read_text()
    output.write_text(template.replace("__DASHBOARD_TITLE__", title).replace("__DASHBOARD_DATA__", data))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case-dir", type=Path, required=True); parser.add_argument("--model", required=True)
    parser.add_argument("--output", type=Path, required=True); parser.add_argument("--effort", choices=("low","medium","high","xhigh"), default="medium")
    parser.add_argument("--through-turn", type=int); parser.add_argument("--schema-only", action="store_true")
    parser.add_argument("--model-timeout-seconds", type=int, default=300)
    parser.add_argument("--semantic-mode", choices=("heuristic", "critic"), default="critic")
    parser.add_argument("--critic-model")
    parser.add_argument("--critic-effort", choices=("low", "medium", "high", "xhigh"), default="low")
    args = parser.parse_args()
    case = json.loads((args.case_dir / "case.json").read_text()); source = args.case_dir / case["inputs"][0]["path"]
    frame, data_profile, source_metadata = load_source(source, not args.schema_only)
    schema = json.loads(SCHEMA_PATH.read_text()); jsonschema.Draft202012Validator.check_schema(schema)
    dump(args.output / "dataset-profile.json", data_profile)
    dump(args.output / "run-config.json", {
        "case_id": case["case_id"], "model": args.model, "effort": args.effort,
        "schema_only": args.schema_only, "source_sha256": file_hash(source),
        "semantic_mode": args.semantic_mode,
        "critic_model": (args.critic_model or args.model) if args.semantic_mode == "critic" else None,
        "critic_effort": args.critic_effort if args.semantic_mode == "critic" else None,
    })
    messages, previous = [], None
    for turn in case["messages"][:args.through_turn or len(case["messages"])]:
        messages.append(turn["text"]); turn_dir = args.output / f"turn-{turn['turn']}"
        prompt = make_prompt(case, data_profile, messages, previous); dump(turn_dir / "plan-request.json", {"model": args.model, "effort": args.effort, "prompt": prompt})
        attempts, attempt_prompt, plan, result = [], prompt, None, None
        for attempt in range(1, 4):
            try:
                candidate, api = call_model(args.model, args.effort, attempt_prompt, schema, args.model_timeout_seconds)
            except (TimeoutError, ModelRequestError) as error:
                record.setdefault("plan", candidate)
                record["validation_error"] = f"{type(error).__name__}: {error}"; attempts.append(record)
                dump(turn_dir / f"plan-attempt-{attempt}.json", record)
                dump(turn_dir / "plan-error.json", {
                    "status": "semantic_critic_timeout", "attempts": attempts,
                    "note": "Operational critic failures do not trigger semantic planner rewrites.",
                })
                raise
            except Exception as error:
                failure = {"attempt": attempt, "request_error": f"{type(error).__name__}: {error}"}
                attempts.append(failure); dump(turn_dir / f"plan-attempt-{attempt}.json", failure)
                dump(turn_dir / "plan-error.json", {"status": "model_request_failed", "attempts": attempts})
                raise
            model_plan = json.loads(json.dumps(candidate))
            record = {"attempt": attempt, "model_plan": model_plan, "api": api}
            try:
                jsonschema.validate(candidate, schema)
                normalizations = (
                    normalize_explicit_constraints(messages, candidate, frame)
                    if args.semantic_mode == "heuristic"
                    else normalize_mechanical_invariants(candidate, frame)
                )
                jsonschema.validate(candidate, schema); bind(candidate, frame); candidate_result = execute(frame, candidate)
                record["plan"] = candidate; record["normalizations"] = normalizations
                if args.semantic_mode == "heuristic":
                    check_conversation_constraints(messages, candidate, candidate_result)
                else:
                    check_mechanical_invariants(candidate, candidate_result)
                    review, critic_api, critic_prompt = call_critic(
                        args.critic_model or args.model, args.critic_effort, messages,
                        data_profile, previous, candidate, candidate_result,
                        args.model_timeout_seconds,
                    )
                    record["critic"] = review; record["critic_api"] = critic_api
                    record["critic_request"] = {
                        "model": args.critic_model or args.model,
                        "effort": args.critic_effort,
                        "raw_rows_included": False,
                        "min_max_or_admitted_values_included": False,
                        "prompt": critic_prompt,
                    }
                    if not review["accepted"]:
                        raise PlanError(f"independent semantic critic rejected plan: {review['feedback']}")
            except Exception as error:
                record.setdefault("plan", candidate)
                record["validation_error"] = f"{type(error).__name__}: {error}"; attempts.append(record)
                dump(turn_dir / f"plan-attempt-{attempt}.json", record)
                if attempt == 3: raise
                attempt_prompt = repair_prompt(prompt, candidate, error); continue
            attempts.append(record); dump(turn_dir / f"plan-attempt-{attempt}.json", record)
            plan, result = candidate, candidate_result; break
        assert plan is not None and result is not None
        dump(turn_dir / "plan-response.json", {"plan": plan, "api": attempts[-1]["api"], "attempts": attempts})
        dump(turn_dir / "validated-plan.json", plan)
        report = make_report(case, source, source_metadata, plan, result); dump(turn_dir / "insight-report.json", report)
        workspace = turn_dir / "workspace"; workspace.mkdir(parents=True, exist_ok=True); write_csv(workspace / "all-result-rows.csv", report); render(report, workspace / "index.html")
        previous = plan; print(json.dumps({
            "turn": turn["turn"], "rows": len(result), "chart": plan["view"]["chart"],
            "attempts": len(attempts),
            "plan_requests": len(attempts),
            "critic_requests": sum(1 for item in attempts if item.get("critic_api")),
            "seconds": sum(
                item["api"]["duration_seconds"] + (item.get("critic_api") or {}).get("duration_seconds", 0)
                for item in attempts
            ),
            "workspace": str(workspace),
        }), flush=True)
    return 0


if __name__ == "__main__": raise SystemExit(main())
