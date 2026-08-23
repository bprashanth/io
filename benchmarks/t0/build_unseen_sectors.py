#!/usr/bin/env python3
"""Build MESSY, realistic spreadsheet fixtures for three sectors an NGO data
tool has not been tested on (agriculture, rural water/WASH, SHG microfinance),
plus pandas-computed ground truth for a fixed set of questions per sector.

Deterministic: same seed (20260824) always produces byte-identical fixtures
and gold.json. Ground truth is computed from the CLEAN logical data (after
stripping title/footer rows, merging inconsistent crop case, and parsing
Indian-formatted numbers/dates) using pandas -- the point is that a loader
which reads the messy on-disk files must reproduce these numbers.

Run with the v2 project's venv:
    .venv-v2/bin/python benchmarks/t0/build_unseen_sectors.py

Outputs:
    benchmarks/t0/unseen/agri/kharif_2025_yield.xlsx  (sheets: Yield, Inputs)
    benchmarks/t0/unseen/wash/handpump_survey.csv
    benchmarks/t0/unseen/mfi/shg_repayment_ledger.csv
    benchmarks/t0/unseen/gold.json
"""

from __future__ import annotations

import csv
import json
import random
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

SEED = 20260824
rng = random.Random(SEED)

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "benchmarks" / "t0" / "unseen"
AGRI_DIR = OUT / "agri"
WASH_DIR = OUT / "wash"
MFI_DIR = OUT / "mfi"
for d in (AGRI_DIR, WASH_DIR, MFI_DIR):
    d.mkdir(parents=True, exist_ok=True)

GOLD_PATH = OUT / "gold.json"


def r2(x) -> float:
    return round(float(x), 2)


def native(v):
    """Convert a pandas/numpy scalar to a plain JSON-safe python value."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, pd.Timestamp):
        return v.strftime("%Y-%m-%d")
    if hasattr(v, "item"):  # numpy/pandas scalar (incl. pandas nullable Int64)
        v = v.item()
    if isinstance(v, float):
        return r2(v)
    return v


def rows_of(df: pd.DataFrame) -> list[list]:
    return [[native(v) for v in row] for row in df.itertuples(index=False, name=None)]


def indian_format(n: int) -> str:
    """Indian digit-grouping: last 3 digits, then groups of 2. e.g. 1250 -> '1,250'."""
    s = str(int(n))
    if len(s) <= 3:
        return s
    head, tail = s[:-3], s[-3:]
    parts = []
    while len(head) > 2:
        parts.insert(0, head[-2:])
        head = head[:-2]
    if head:
        parts.insert(0, head)
    return ",".join(parts + [tail])


def dd_mm_yyyy(d: date) -> str:
    return d.strftime("%d-%m-%Y")


def dd_slash_mm_yyyy(d: date) -> str:
    return d.strftime("%d/%m/%Y")


def random_date(start: date, end: date) -> date:
    span = (end - start).days
    return start + timedelta(days=rng.randint(0, span))


# ======================================================================
# SECTOR 1: agri/kharif_2025_yield.xlsx  (sheets "Yield", "Inputs")
# ======================================================================

AGRI_N = 220

VILLAGES = [
    "Toranmal", "Molgi", "Bilgaon", "Manibeli", "Roshmal",
    "Sindhi", "Dab", "Kathi", "Waki", "Dhanora",
]

CROPS = ["Soyabean", "Cotton", "Maize", "Tur", "Groundnut"]

VARIETIES = {
    "Soyabean": ["JS 335", "JS 9560", "MAUS 71"],
    "Cotton": ["Bt Cotton", "RCH 2 Bt", "Suraj"],
    "Maize": ["African Tall", "DHM 117", "Ganga Safed 2"],
    "Tur": ["ICPL 87", "BSMR 736", "Vipula"],
    "Groundnut": ["TAG 24", "JL 24", "TG 37A"],
}

YIELD_RANGE = {  # qtl/acre
    "Soyabean": (6.0, 14.0),
    "Cotton": (3.0, 7.0),
    "Maize": (12.0, 28.0),
    "Tur": (3.0, 7.0),
    "Groundnut": (5.0, 10.0),
}

SEED_COST_RANGE = {
    "Soyabean": (1200, 3200),
    "Cotton": (2200, 4500),
    "Maize": (900, 2400),
    "Tur": (700, 1800),
    "Groundnut": (1500, 3600),
}

FIRST_NAMES = [
    "Ramesh", "Sunil", "Kailas", "Ganesh", "Bhagwan", "Tulsiram", "Somnath",
    "Dilip", "Ashok", "Vinod", "Sarla", "Kamla", "Suman", "Laxmi", "Radha",
    "Sunita", "Manisha", "Kavita", "Anita", "Shobha",
]
LAST_NAMES = ["Pawara", "Vasave", "Valvi", "Gavit", "Padvi", "Dhangar"]

REMARK_POOL = ["pest damage", "delayed sowing", "drought stress", "good rainfall", "replanted"]


def full_farmer_name() -> str:
    return f"{rng.choice(FIRST_NAMES)} {rng.choice(LAST_NAMES)}"


def cased_crop(canonical: str) -> str:
    style = rng.choice(["title", "upper", "lower"])
    if style == "title":
        return canonical
    if style == "upper":
        return canonical.upper()
    return canonical.lower()


def build_agri():
    farmer_ids = [f"NDB-{i + 1:04d}" for i in range(AGRI_N)]
    yield_rows_clean = []
    yield_rows_display = []  # for the messy xlsx (cased crop, text dates)
    for fid in farmer_ids:
        name = full_farmer_name()
        village = rng.choice(VILLAGES)
        crop = rng.choice(CROPS)
        variety = rng.choice(VARIETIES[crop])
        area = r2(rng.uniform(0.4, 5.5) if rng.random() < 0.85 else rng.uniform(5.5, 8.0))
        lo, hi = YIELD_RANGE[crop]
        yld = r2(rng.uniform(lo, hi))
        sow = random_date(date(2025, 6, 15), date(2025, 7, 15))
        harvest = random_date(date(2025, 10, 15), date(2025, 11, 30))
        irrigated = "Y" if rng.random() < 0.4 else "N"
        remark = rng.choice(REMARK_POOL) if rng.random() < 0.15 else ""

        yield_rows_clean.append(
            [fid, name, village, crop, variety, area, yld, sow, harvest, irrigated, remark]
        )
        yield_rows_display.append(
            [fid, name, village, cased_crop(crop), variety, area, yld,
             dd_mm_yyyy(sow), dd_mm_yyyy(harvest), irrigated, remark]
        )

    yield_clean = pd.DataFrame(
        yield_rows_clean,
        columns=["Farmer ID", "Farmer Name", "Village", "Crop", "Variety",
                  "Area (acre)", "Yield (qtl/acre)", "Sowing Date", "Harvest Date",
                  "Irrigated", "Remarks"],
    )

    inputs_rows_clean = []
    for fid, crop in zip(farmer_ids, yield_clean["Crop"]):
        slo, shi = SEED_COST_RANGE[crop]
        seed_cost = rng.randint(slo, shi)
        fert_cost = rng.randint(1500, 6000)
        labour_days = rng.randint(5, 40)
        loan = "Yes" if rng.random() < 0.45 else "No"
        inputs_rows_clean.append([fid, seed_cost, fert_cost, labour_days, loan])

    inputs_clean = pd.DataFrame(
        inputs_rows_clean,
        columns=["Farmer ID", "Seed Cost Rs", "Fertiliser Cost Rs", "Labour Days", "Loan Taken"],
    )

    # ---- write the messy workbook ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Yield"

    ws["A1"] = "Kharif 2025 – Plot wise yield, FPO Nandurbar"
    ws.merge_cells("A1:K1")
    ws.append([])  # row 2 left blank deliberately
    header = ["Farmer ID", "Farmer Name", "Village", "Crop", "Variety",
              "Area (acre)", "Yield (qtl/acre)", "Sowing Date (dd-mm-yyyy)",
              "Harvest Date", "Irrigated (Y/N)", "Remarks"]
    ws.append(header)  # row 3
    for row in yield_rows_display:
        ws.append(row)
    total_area = r2(sum(r[5] for r in yield_rows_display))
    ws.append(["Total", "", "", "", "", total_area, "", "", "", "", ""])
    ws.append(["Source: field survey", "", "", "", "", "", "", "", "", "", ""])

    ws2 = wb.create_sheet("Inputs")
    ws2.append(["Farmer ID", "Seed Cost Rs", "Fertiliser Cost Rs", "Labour Days", "Loan Taken (Yes/No)"])
    for row in inputs_rows_clean:
        ws2.append(row)

    wb.save(AGRI_DIR / "kharif_2025_yield.xlsx")

    return yield_clean, inputs_clean


def gold_agri(yield_clean: pd.DataFrame, inputs_clean: pd.DataFrame) -> list[dict]:
    items = []

    g1 = (
        yield_clean.groupby("Crop", as_index=False)["Yield (qtl/acre)"]
        .mean()
        .round(2)
        .rename(columns={"Yield (qtl/acre)": "Avg Yield (qtl/acre)"})
        .sort_values("Crop")
        .reset_index(drop=True)
    )
    items.append({
        "sector": "agri",
        "file": "agri/kharif_2025_yield.xlsx#Yield",
        "question": "average yield per acre by crop",
        "columns": ["Crop", "Avg Yield (qtl/acre)"],
        "expected_rows": rows_of(g1),
        "notes": (
            "Crop names are inconsistently cased in the file "
            "(Soyabean/soyabean/SOYABEAN); must case-fold before grouping. "
            "No ranking word was given, so rows are ordered alphabetically by "
            "crop (pandas groupby default), not by value."
        ),
    })

    soy = yield_clean[yield_clean["Crop"] == "Soyabean"]
    g2 = (
        soy.groupby("Village", as_index=False)["Area (acre)"]
        .sum()
        .round(2)
        .sort_values(["Area (acre)", "Village"], ascending=[False, True])
        .head(5)
        .reset_index(drop=True)
    )
    items.append({
        "sector": "agri",
        "file": "agri/kharif_2025_yield.xlsx#Yield",
        "question": "top 5 villages by total soyabean area",
        "columns": ["Village", "Area (acre)"],
        "expected_rows": rows_of(g2),
        "notes": "Case-insensitive crop match on 'Soyabean'; ties broken by village name ascending.",
    })

    crop_avg = yield_clean.groupby("Crop")["Yield (qtl/acre)"].transform("mean")
    mask = (yield_clean["Irrigated"] == "Y") & (yield_clean["Yield (qtl/acre)"] < crop_avg)
    g3 = yield_clean.loc[mask, ["Farmer ID", "Farmer Name", "Village", "Crop", "Yield (qtl/acre)"]].copy()
    g3["Crop Avg Yield (qtl/acre)"] = crop_avg[mask].round(2)
    g3 = g3.sort_values("Farmer ID").reset_index(drop=True)
    items.append({
        "sector": "agri",
        "file": "agri/kharif_2025_yield.xlsx#Yield",
        "question": "which irrigated plots yielded below the crop average",
        "columns": ["Farmer ID", "Farmer Name", "Village", "Crop", "Yield (qtl/acre)", "Crop Avg Yield (qtl/acre)"],
        "expected_rows": rows_of(g3),
        "notes": (
            "Crop average is computed across ALL plots of that crop (irrigated "
            "and rainfed) using merged case, then compared per-row using the "
            "unrounded mean (rounding only for display). No ranking word given, "
            "so rows are sorted by Farmer ID ascending."
        ),
    })

    merged = yield_clean[["Farmer ID", "Farmer Name"]].merge(inputs_clean, on="Farmer ID")
    merged["Total Input Cost (Rs)"] = merged["Seed Cost Rs"] + merged["Fertiliser Cost Rs"]
    g4 = (
        merged[["Farmer ID", "Farmer Name", "Total Input Cost (Rs)"]]
        .sort_values(["Total Input Cost (Rs)", "Farmer ID"], ascending=[False, True])
        .reset_index(drop=True)
    )
    items.append({
        "sector": "agri",
        "file": ["agri/kharif_2025_yield.xlsx#Yield", "agri/kharif_2025_yield.xlsx#Inputs"],
        "question": "total input cost per farmer, highest first (join the two sheets)",
        "columns": ["Farmer ID", "Farmer Name", "Total Input Cost (Rs)"],
        "expected_rows": rows_of(g4),
        "notes": (
            "Total input cost = Seed Cost Rs + Fertiliser Cost Rs (Labour Days "
            "and Loan Taken excluded, they are not a cost). Farmer ID is 1:1 "
            "between the two sheets in this fixture. Full ranking, all 220 "
            "farmers, ties broken by Farmer ID ascending."
        ),
    })

    return items


# ======================================================================
# SECTOR 2: wash/handpump_survey.csv
# ======================================================================

WASH_N = 260

DISTRICT_BLOCKS = {
    "Koraput": ["Koraput Sadar", "Semiliguda", "Boipariguda"],
    "Kalahandi": ["Bhawanipatna", "Junagarh", "Narla"],
    "Rayagada": ["Rayagada Sadar", "Gunupur", "Kashipur"],
    "Nabarangpur": ["Nabarangpur Sadar", "Umerkote", "Dabugaon"],
}
DISTRICTS = list(DISTRICT_BLOCKS.keys())

HAB_PREFIXES = ["Kendu", "Padma", "Jhola", "Simili", "Baman", "Gopal", "Ambo",
                 "Ranga", "Tenta", "Kumuli", "Duduma", "Chandi", "Manik", "Sano"]
HAB_SUFFIXES = ["pur", "guda", "padar", "basti", "sahi", "gaon", "palli"]

PUMP_TYPES = ["India Mark II", "India Mark III", "Tara Pump", "Shallow Hand Pump"]


def random_habitation() -> str:
    return f"{rng.choice(HAB_PREFIXES)}{rng.choice(HAB_SUFFIXES)}"


def build_wash():
    rows_clean = []
    rows_display = []
    for _ in range(WASH_N):
        district = rng.choice(DISTRICTS)
        block = rng.choice(DISTRICT_BLOCKS[district])
        habitation = random_habitation()
        pump_type = rng.choice(PUMP_TYPES)
        installed_year = rng.randint(1998, 2022)
        users = rng.randint(15, 1200)
        depth = rng.randint(40, 350)

        repaired_ever = rng.random() < 0.75
        if repaired_ever:
            last_repair = random_date(date(installed_year + 1, 1, 1), date(2026, 8, 1))
        else:
            last_repair = None

        functional = "Functional" if rng.random() < 0.75 else "Not Functional"

        quality_roll = rng.random()
        if quality_roll < 0.15:
            quality = "Not Tested"
            tds = None
        elif quality_roll < 0.15 + 0.25:
            quality = "Unsafe"
            tds = rng.randint(500, 2500)
        else:
            quality = "Safe"
            tds = rng.randint(150, 500)

        lat = r2(rng.uniform(18.0, 19.9))
        lon = r2(rng.uniform(82.0, 83.9))

        rows_clean.append([
            habitation, block, district, pump_type, installed_year, users, depth,
            pd.Timestamp(last_repair) if last_repair else pd.NaT,
            functional, quality, tds,
        ])
        rows_display.append([
            habitation, block, district, pump_type, installed_year,
            indian_format(users), depth,
            dd_slash_mm_yyyy(last_repair) if last_repair else "",
            functional, quality,
            "" if tds is None else tds,
            f"{lat}, {lon}",
        ])

    clean = pd.DataFrame(
        rows_clean,
        columns=["Habitation", "Block", "District", "Pump Type", "Installed Year",
                  "Users (HH)", "Depth (ft)", "Last Repair", "Status Functional",
                  "Status Quality", "TDS (ppm)"],
    )

    header = ["Habitation ", "Block ", "District ", "Pump Type ", "Installed Year ",
              "Users (HH) ", "Depth (ft) ", "Last Repair ", "Status ", "Status ",
              "TDS (ppm) ", "Lat, Long "]
    with (WASH_DIR / "handpump_survey.csv").open("w", newline="") as fh:
        writer = csv.writer(fh, lineterminator="\n")
        writer.writerow(header)
        writer.writerows(rows_display)

    return clean


def gold_wash(clean: pd.DataFrame) -> list[dict]:
    items = []

    g1 = (
        clean[clean["Status Functional"] == "Not Functional"]
        .groupby("Block", as_index=False)
        .size()
        .rename(columns={"size": "Not Functional Count"})
        .sort_values("Block")
        .reset_index(drop=True)
    )
    items.append({
        "sector": "wash",
        "file": "wash/handpump_survey.csv",
        "question": "how many handpumps are not functional per block",
        "columns": ["Block", "Not Functional Count"],
        "expected_rows": rows_of(g1),
        "notes": (
            "The functional-status column is the FIRST of the two literal "
            "'Status ' columns in the file. Blocks that have zero not-functional "
            "pumps are absent (groupby of the filtered rows), not zero-filled. "
            "No ranking word, so alphabetical by block."
        ),
    })

    func = clean[clean["Status Functional"] == "Functional"]
    g2 = (
        func.groupby("District", as_index=False)["TDS (ppm)"]
        .mean()
        .round(2)
        .sort_values("District")
        .reset_index(drop=True)
    )
    items.append({
        "sector": "wash",
        "file": "wash/handpump_survey.csv",
        "question": "average TDS by district for functional pumps",
        "columns": ["District", "Avg TDS (ppm)"],
        "expected_rows": rows_of(g2.rename(columns={"TDS (ppm)": "Avg TDS (ppm)"})),
        "notes": (
            "TDS is the SECOND numeric-looking column after the two 'Status ' "
            "columns; rows with quality 'Not Tested' have blank TDS and are "
            "excluded by pandas mean() skipna, not treated as zero."
        ),
    })

    mask = (clean["Depth (ft)"] > 200) & (
        clean["Last Repair"].isna() | (clean["Last Repair"].dt.year < 2023)
    )
    g3 = clean.loc[mask, ["Habitation", "Block", "District", "Depth (ft)", "Last Repair"]].copy()
    g3["Last Repair"] = g3["Last Repair"].apply(lambda v: v.strftime("%Y-%m-%d") if pd.notna(v) else "Never repaired")
    g3 = g3.sort_values(["Depth (ft)", "Habitation"], ascending=[False, True]).reset_index(drop=True)
    items.append({
        "sector": "wash",
        "file": "wash/handpump_survey.csv",
        "question": "pumps deeper than 200 ft not repaired since 2023",
        "columns": ["Habitation", "Block", "District", "Depth (ft)", "Last Repair"],
        "expected_rows": rows_of(g3),
        "notes": (
            "'Last Repair' text 'dd/mm/yyyy' must be parsed to a date. A blank "
            "Last Repair means the pump has never been professionally repaired, "
            "which is judged to satisfy 'not repaired since 2023' unconditionally "
            "(shown as 'Never repaired'), not excluded as missing data. Sorted "
            "by depth descending, ties by habitation ascending."
        ),
    })

    g4 = (
        clean.groupby("Block", as_index=False)["Users (HH)"]
        .sum()
        .sort_values("Block")
        .reset_index(drop=True)
    )
    items.append({
        "sector": "wash",
        "file": "wash/handpump_survey.csv",
        "question": "total users served by block",
        "columns": ["Block", "Total Users (HH)"],
        "expected_rows": rows_of(g4.rename(columns={"Users (HH)": "Total Users (HH)"})),
        "notes": (
            "Users (HH) is stored Indian-comma-formatted as text (e.g. '1,250') "
            "and must be parsed back to an integer before summing. No ranking "
            "word, so alphabetical by block."
        ),
    })

    return items


# ======================================================================
# SECTOR 3: mfi/shg_repayment_ledger.csv
# ======================================================================

MFI_N = 180

MFI_BLOCKS = ["Nawada Sadar", "Rajauli", "Hisua", "Warisaliganj", "Pakribarawan", "Gobindpur"]

SHG_THEME = ["Jai Durga", "Maa Saraswati", "Jai Bhawani", "Shanti", "Ekta", "Pragati",
             "Ganga", "Jyoti", "Sakhi", "Vikas", "Nari Shakti", "Ambika", "Annapurna",
             "Krishna", "Lakshmi", "Sarojini", "Jai Kali", "Maa Durga", "Radha Krishna",
             "Vaishnavi", "Gayatri", "Tulsi", "Chandani", "Bhagirathi", "Ganga Jamuna",
             "Jai Hanuman", "Navjyoti", "Uday", "Samriddhi", "Sanjeevani", "Shakti",
             "Kalyani", "Mangalam", "Utthan", "Nav Nirman", "Aastha", "Sahyog",
             "Jyoti Kiran", "Prerna", "Sadbhavna"]
SHG_SUFFIX = ["Mahila SHG", "Swayam Sahayata Samuh", "Mahila Sangh", "SHG", "Mahila Samuh"]

MONTH_COLS = ["Apr-25", "May-25", "Jun-25", "Jul-25", "Aug-25", "Sep-25",
              "Oct-25", "Nov-25", "Dec-25", "Jan-26", "Feb-26", "Mar-26"]


def unique_shg_names(n: int) -> list[str]:
    combos = [f"{theme} {suffix}" for theme in SHG_THEME for suffix in SHG_SUFFIX]
    if n > len(combos):
        raise ValueError(f"need {n} unique SHG names but only {len(combos)} theme/suffix combos exist")
    rng.shuffle(combos)
    return combos[:n]


def build_mfi():
    names = unique_shg_names(MFI_N)
    mismatch_idx = set(rng.sample(range(MFI_N), 5))

    records = []
    for i, name in enumerate(names):
        block = rng.choice(MFI_BLOCKS)
        members = rng.randint(8, 15)
        loan_amount = rng.randint(20000, 150000)

        target_rate = rng.uniform(0.25, 1.05)
        target_total = loan_amount * target_rate

        active_months = sorted(rng.sample(range(12), k=rng.randint(6, 12)))
        if active_months:
            weights = [rng.random() + 0.2 for _ in active_months]
            wsum = sum(weights)
            amounts = [max(0, round((w / wsum) * target_total / 10) * 10) for w in weights]
        else:
            amounts = []

        month_values = [None] * 12
        for pos, amt in zip(active_months, amounts):
            if amt > 0:
                month_values[pos] = int(amt)

        sum_of_months = sum(v for v in month_values if v)

        if i in mismatch_idx:
            error = rng.choice([-5000, -2500, -1000, 1000, 2500, 5000, 300, -300])
            typed_total = sum_of_months + error
        else:
            typed_total = sum_of_months

        rate = sum_of_months / loan_amount if loan_amount else 0.0
        overdue = "Y" if rate < 0.5 else "N"

        records.append({
            "SHG Name": name,
            "Block": block,
            "Members": members,
            "Loan Amount": loan_amount,
            **{MONTH_COLS[j]: month_values[j] for j in range(12)},
            "Total Repaid": typed_total,
            "Overdue (Y/N)": overdue,
            "_Sum Of Months": sum_of_months,
        })

    clean = pd.DataFrame(records)
    clean[MONTH_COLS] = clean[MONTH_COLS].astype("Int64")

    display = clean.drop(columns=["_Sum Of Months"]).copy()
    display.to_csv(MFI_DIR / "shg_repayment_ledger.csv", index=False, na_rep="")

    return clean


def gold_mfi(clean: pd.DataFrame) -> list[dict]:
    items = []

    g1 = (
        clean.groupby("Block", as_index=False)["_Sum Of Months"]
        .sum()
        .rename(columns={"_Sum Of Months": "Total Repaid (Rs)"})
        .sort_values("Block")
        .reset_index(drop=True)
    )
    items.append({
        "sector": "mfi",
        "file": "mfi/shg_repayment_ledger.csv",
        "question": "total repaid per block from the monthly columns",
        "columns": ["Block", "Total Repaid (Rs)"],
        "expected_rows": rows_of(g1),
        "notes": (
            "Sum the 12 month columns (Apr-25..Mar-26), treating blank cells as "
            "0; the typed 'Total Repaid' column is deliberately ignored per the "
            "question. No ranking word, so alphabetical by block."
        ),
    })

    mismatch = clean[clean["Total Repaid"] != clean["_Sum Of Months"]].copy()
    mismatch["Difference (Typed - Computed)"] = mismatch["Total Repaid"] - mismatch["_Sum Of Months"]
    g2 = (
        mismatch[["SHG Name", "Block", "Total Repaid", "_Sum Of Months", "Difference (Typed - Computed)"]]
        .rename(columns={"Total Repaid": "Typed Total", "_Sum Of Months": "Computed Sum Of Months"})
        .sort_values("SHG Name")
        .reset_index(drop=True)
    )
    items.append({
        "sector": "mfi",
        "file": "mfi/shg_repayment_ledger.csv",
        "question": "SHGs whose typed total does not match the sum of months",
        "columns": ["SHG Name", "Block", "Typed Total", "Computed Sum Of Months", "Difference (Typed - Computed)"],
        "expected_rows": rows_of(g2),
        "notes": "Exactly 5 SHGs are constructed with a mismatched typed total. Sorted by SHG Name ascending.",
    })

    month_totals = clean[MONTH_COLS].sum()
    g3 = (
        month_totals.reset_index()
        .rename(columns={"index": "Month", 0: "Total Collected (Rs)"})
        .sort_values("Total Collected (Rs)")
        .reset_index(drop=True)
    )
    items.append({
        "sector": "mfi",
        "file": "mfi/shg_repayment_ledger.csv",
        "question": "months with the lowest collections overall",
        "columns": ["Month", "Total Collected (Rs)"],
        "expected_rows": rows_of(g3),
        "notes": (
            "All 12 months returned, sorted ascending by total collections "
            "(blank month cells treated as 0); the caller can take head(n) for "
            "'the lowest N'."
        ),
    })

    rate_df = clean[["SHG Name", "Block", "_Sum Of Months", "Loan Amount"]].copy()
    rate_df["Repayment Rate"] = (rate_df["_Sum Of Months"] / rate_df["Loan Amount"]).round(2)
    g4 = (
        rate_df[["SHG Name", "Block", "Repayment Rate"]]
        .sort_values(["Repayment Rate", "SHG Name"])
        .head(5)
        .reset_index(drop=True)
    )
    items.append({
        "sector": "mfi",
        "file": "mfi/shg_repayment_ledger.csv",
        "question": "repayment rate = repaid/loan amount per SHG, lowest 5",
        "columns": ["SHG Name", "Block", "Repayment Rate"],
        "expected_rows": rows_of(g4),
        "notes": (
            "Repaid = computed sum of the 12 month columns (not the typed "
            "total). Rate rounded to 2 dp; ties broken by SHG Name ascending."
        ),
    })

    return items


# ======================================================================
# main
# ======================================================================

def main():
    yield_clean, inputs_clean = build_agri()
    wash_clean = build_wash()
    mfi_clean = build_mfi()

    gold = []
    gold += gold_agri(yield_clean, inputs_clean)
    gold += gold_wash(wash_clean)
    gold += gold_mfi(mfi_clean)

    with GOLD_PATH.open("w") as fh:
        json.dump(gold, fh, indent=2)
        fh.write("\n")

    print(f"agri Yield rows: {len(yield_clean)}, Inputs rows: {len(inputs_clean)}")
    print(f"wash rows: {len(wash_clean)}")
    print(f"mfi rows: {len(mfi_clean)}")
    print(f"gold questions: {len(gold)} -> {GOLD_PATH}")
    for item in gold:
        print(f"  [{item['sector']}] {item['question']} -> {len(item['expected_rows'])} row(s)")


if __name__ == "__main__":
    main()
