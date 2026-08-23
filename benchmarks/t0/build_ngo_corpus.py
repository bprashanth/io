#!/usr/bin/env python3
"""Deterministic synthetic Indian small-NGO test corpus generator.

Produces messy, realistic-shaped files for 8 fictional NGOs under
benchmarks/t0/ngo-corpus/<org-slug>/, a README.md documenting the data
shapes, and a cases.json of 40 ask/build/page cases (with expected_rows
computed from the generated data for the ask cases where cheap).

Everything here is synthetic: no real people, no real organisations. PII
formats (Aadhaar-shaped 12-digit numbers, PAN-shaped strings, phone numbers,
UDID-shaped strings, bank account/IFSC-shaped strings) are realistic in
*shape* only, generated from a seeded RNG.

Seed: 20260825. Re-running this script reproduces byte-identical output
(modulo openpyxl's own non-determinism, which we do not rely on for any
case's expected values -- all expected values are recomputed from the
written files in the same run).

Usage:
    .venv-v2/bin/python benchmarks/t0/build_ngo_corpus.py
"""
from __future__ import annotations

import csv
import json
import os
import random
import shutil
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

SEED = 20260825
HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "ngo-corpus")

RNG = random.Random(SEED)

# ---------------------------------------------------------------------------
# Name / place pools (Maharashtra-weighted, matching the io.xlsx sourcing
# region, plus a spread of other Indian regions/communities since real NGO
# beneficiary lists are rarely single-community).
# ---------------------------------------------------------------------------

MALE_FIRST = [
    "Rahul", "Amit", "Sandeep", "Vijay", "Sunil", "Ganesh", "Prakash", "Ravi",
    "Suresh", "Mahesh", "Ramesh", "Sachin", "Vishal", "Rohan", "Nitin",
    "Sanjay", "Ashok", "Dinesh", "Kiran", "Santosh", "Pravin", "Yogesh",
    "Anil", "Manoj", "Rajesh", "Imran", "Salman", "Aslam", "Firoz", "Iqbal",
    "Gurpreet", "Harpreet", "Jaspal", "Balwinder", "Om", "Akash", "Karan",
    "Abhijit", "Tushar", "Prashant",
]
FEMALE_FIRST = [
    "Sunita", "Priya", "Anjali", "Kavita", "Rekha", "Meena", "Pooja",
    "Nisha", "Deepa", "Savita", "Vaishali", "Snehal", "Manisha", "Jyoti",
    "Asha", "Rani", "Kalpana", "Shobha", "Vandana", "Archana", "Neha",
    "Shabana", "Rukhsar", "Fatima", "Ayesha", "Zarina", "Simran",
    "Gurmeet", "Harpreet", "Kajal", "Sarita", "Lata", "Usha", "Radha",
    "Swati", "Pallavi", "Divya", "Komal", "Rupali", "Mangal",
]
SURNAMES = [
    "Pawar", "Jadhav", "Shinde", "Kale", "More", "Gaikwad", "Deshmukh",
    "Patil", "Sawant", "Kadam", "Bhosale", "Chavan", "Kulkarni", "Joshi",
    "Bagul", "Nikam", "Gawde", "Waghmare", "Salunkhe", "Thorat", "Khan",
    "Shaikh", "Ansari", "Pathan", "Sheikh", "Sayyed", "Kaur", "Singh",
    "Chowdhury", "Yadav", "Mane", "Bhoir", "Dhumal", "Rathod", "Suryawanshi",
]

VILLAGES = [
    "Wadgaon", "Shirur", "Baramati", "Velhe", "Mulshi", "Junnar", "Ambegaon",
    "Purandar", "Bhor", "Daund", "Indapur", "Haveli", "Khed", "Maval",
    "Shirval", "Saswad", "Rajgurunagar", "Otur", "Manchar", "Narayangaon",
]
DISTRICTS = ["Pune", "Satara", "Ahmednagar", "Solapur", "Raigad"]
BLOCKS = ["Block A", "Block B", "Block C", "Block D"]

BANKS = [
    "Bank of Maharashtra", "State Bank of India", "Union Bank of India",
    "Punjab National Bank", "Bank of India", "Central Bank of India",
    "IDBI Bank", "Canara Bank",
]
IFSC_PREFIX = {
    "Bank of Maharashtra": "MAHB",
    "State Bank of India": "SBIN",
    "Union Bank of India": "UBIN",
    "Punjab National Bank": "PUNB",
    "Bank of India": "BKID",
    "Central Bank of India": "CBIN",
    "IDBI Bank": "IBKL",
    "Canara Bank": "CNRB",
}

SCHOOL_NAMES = [
    "Zilla Parishad Primary School, {v}", "Municipal School No. {n}, {v}",
    "Anganwadi Kendra {v}-{n}", "Adarsh Vidyalaya {v}",
]

HINGLISH_NOTES = [
    "sab thik hai", "bacche khush the", "ration late aaya", "rasta kharab tha",
    "gaon mein pani ki problem hai", "school band tha aaj", "sab present the",
    "mother absent thi, बच्चा akela aaya", "coach ne extra session liya",
    "material kam pada", "next visit 2 hafte baad", "sarpanch se baat hui",
    "no issues reported", "van transport delay hua", "team ne feedback liya",
]

CAMPAIGN_NAMES = ["MonsoonRelief2025", "GirlChildEducation", "AnnualGala2025",
                   "EmergencyMedicalFund", "SponsorAChild", "WinterDrive2025"]

LEAD_SOURCES = ["Website", "Referral", "Event", "Social Media", "Cold Call",
                "Corporate CSR", "WalkIn"]
LEAD_STATUS = ["New", "Contacted", "Qualified", "Converted", "Not Interested",
               "Junk Lead"]

EXPENSE_HEADS = {
    "Programme": ["Training materials", "Field travel", "Volunteer stipend",
                  "Beneficiary kits"],
    "Admin": ["Office rent", "Stationery", "Internet & phone", "Bank charges"],
    "Salaries": ["Field staff salary", "Programme staff salary"],
    "Logistics": ["Fuel", "Vehicle maintenance", "Courier"],
}

DEVICE_TYPES = ["Wheelchair", "Hearing aid", "Walking stick", "Tricycle",
                 "Crutches", "Braille kit", "Calipers"]
DISABILITY_TYPES = ["Locomotor", "Hearing Impairment", "Visual Impairment",
                     "Intellectual Disability", "Multiple Disability",
                     "Speech and Language Disability"]

MONTHS_2025 = ["April", "May", "June", "July", "August", "September"]
MONTH_ABBR_2025 = ["Apr-25", "May-25", "Jun-25", "Jul-25", "Aug-25", "Sep-25"]


def gendered_name(rng: random.Random, gender: str) -> tuple[str, str]:
    pool = FEMALE_FIRST if gender == "F" else MALE_FIRST
    return rng.choice(pool), rng.choice(SURNAMES)


def full_name(rng: random.Random) -> tuple[str, str]:
    return gendered_name(rng, "F" if rng.random() < 0.5 else "M")


def phone(rng: random.Random) -> str:
    return f"{rng.choice('6789')}{rng.randint(10**8, 10**9 - 1)}"


def aadhaar(rng: random.Random) -> str:
    n = rng.randint(10**11, 10**12 - 1)
    s = str(n)
    return f"{s[0:4]} {s[4:8]} {s[8:12]}"


def pan(rng: random.Random) -> str:
    letters = "".join(rng.choice("ABCDEFGHIJKLMNOPQRSTUVWXYZ") for _ in range(5))
    digits = "".join(rng.choice("0123456789") for _ in range(4))
    last = rng.choice("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
    return f"{letters}{digits}{last}"


def bank_account(rng: random.Random) -> str:
    return str(rng.randint(10**10, 10**12 - 1))


def ifsc(rng: random.Random, bank: str) -> str:
    branch = rng.randint(100000, 999999)
    return f"{IFSC_PREFIX[bank]}0{branch}"


def udid(rng: random.Random, dob_year: int) -> str:
    state = "27"  # Maharashtra state code
    district = f"{rng.randint(1, 35):02d}"
    cmo = str(rng.randint(1, 9))
    dis_type = f"{rng.randint(1, 21):02d}"
    running = f"{rng.randint(1, 999999):06d}"
    check = str(rng.randint(0, 9))
    return f"{state}{district}{cmo}{dis_type}{dob_year}{running}{check}"


def dd_mm_yyyy(d: datetime) -> str:
    return d.strftime("%d/%m/%Y")


def yyyy_mm_dd(d: datetime) -> str:
    return d.strftime("%Y-%m-%d")


def rand_date(rng: random.Random, start: datetime, end: datetime) -> datetime:
    delta = (end - start).days
    return start + timedelta(days=rng.randint(0, max(delta, 0)))


def inr(amount: float) -> str:
    return f"₹{amount:,.2f}"


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


# ---------------------------------------------------------------------------
# Registry used to assemble cases.json expected values after generation.
# ---------------------------------------------------------------------------
GENERATED = {}  # org_slug -> {filename: {sheet_or_none: DataFrame}}


def record(org: str, fname: str, df: pd.DataFrame, sheet: str | None = None):
    GENERATED.setdefault(org, {}).setdefault(fname, {})[sheet] = df


# ===========================================================================
# ORG 1: education NGO
# ===========================================================================

def gen_education():
    org = "sunrise-shiksha"
    d = os.path.join(OUT, org)
    ensure_dir(d)
    rng = random.Random(f"{SEED}-education")

    n = 320
    classes = [f"{g}{s}" for g in range(1, 8) for s in "AB"]
    enrol_rows = []
    names = []
    for i in range(1, n + 1):
        gender = rng.choice(["M", "F"])
        fn, sn = gendered_name(rng, gender)
        names.append(f"{fn} {sn}")
        dob = rand_date(rng, datetime(2011, 1, 1), datetime(2019, 6, 1))
        pfn, psn = rng.choice(MALE_FIRST), sn
        village = rng.choice(VILLAGES)
        enrol_rows.append({
            "sno": i,
            "student_id": f"STU-25-{i:04d}",
            "name": f"{fn} {sn}",
            "gender": gender,
            "dob": dd_mm_yyyy(dob),
            "class": rng.choice(classes),
            "parent_name": f"{pfn} {psn}",
            "parent_occupation": rng.choice(
                ["Farmer", "Daily wage labour", "Auto driver", "Shopkeeper",
                 "Domestic worker", "Farm labour", "Mason", "Tailor"]),
            "aadhaar": aadhaar(rng),
            "phone": phone(rng),
            "village": village,
            "enrolment_date": dd_mm_yyyy(rand_date(rng, datetime(2025, 5, 15), datetime(2025, 6, 20))),
        })
    enrol_df = pd.DataFrame(enrol_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Enrolment"

    # two-row merged group header + sub-header, matching the messy shape
    group_spans = [
        ("S.No", 1, 1),
        ("Student Details", 2, 6),
        ("Parent/Guardian", 7, 8),
        ("Identification & Contact", 9, 12),
    ]
    for label, c1, c2 in group_spans:
        ws.cell(row=1, column=c1, value=label).font = Font(bold=True)
        if c2 > c1:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        else:
            ws.merge_cells(start_row=1, start_column=c1, end_row=2, end_column=c1)
    # Actual column layout: S.No | Student ID Name Gender DOB Class Village | Parent Name Occupation | Aadhaar Phone EnrolDate
    headers_row2 = ["Student ID", "Name", "Gender", "DOB", "Class",
                     "Village", "Parent Name", "Parent Occupation",
                     "Aadhaar Number", "Phone Number", "Enrolment Date"]
    for c, h in enumerate(headers_row2, start=2):
        ws.cell(row=2, column=c, value=h).font = Font(bold=True)

    for r, row in enumerate(enrol_rows, start=3):
        ws.cell(row=r, column=1, value=row["sno"])
        ws.cell(row=r, column=2, value=row["student_id"])
        ws.cell(row=r, column=3, value=row["name"])
        ws.cell(row=r, column=4, value=row["gender"])
        ws.cell(row=r, column=5, value=row["dob"])
        ws.cell(row=r, column=6, value=row["class"])
        ws.cell(row=r, column=7, value=row["village"])
        ws.cell(row=r, column=8, value=row["parent_name"])
        ws.cell(row=r, column=9, value=row["parent_occupation"])
        ws.cell(row=r, column=10, value=row["aadhaar"])
        ws.cell(row=r, column=11, value=row["phone"])
        ws.cell(row=r, column=12, value=row["enrolment_date"])

    # ---- Attendance sheet: wide by month, different ID format ----
    ws2 = wb.create_sheet("Attendance")
    att_months = MONTH_ABBR_2025
    working_days = {"Apr-25": 22, "May-25": 20, "Jun-25": 21, "Jul-25": 24,
                     "Aug-25": 22, "Sep-25": 23}
    headers = ["Attendance ID", "Student Name", "Class"] + att_months + ["Total Present"]
    ws2.append(headers)
    for c in ws2[1]:
        c.font = Font(bold=True)
    att_rows = []
    for i, row in enumerate(enrol_rows, start=1):
        att_id = f"ATT/2025/{i:04d}"
        present = []
        base_rate = rng.uniform(0.72, 0.99)
        for m in att_months:
            wd = working_days[m]
            p = max(0, min(wd, round(rng.gauss(base_rate * wd, 2))))
            present.append(p)
        total = sum(present)
        ws2.append([att_id, row["name"], row["class"], *present, total])
        att_rows.append({"attendance_id": att_id, "name": row["name"],
                          "class": row["class"],
                          **dict(zip(att_months, present)), "total_present": total})
    att_df = pd.DataFrame(att_rows)

    # ---- Assessment sheet: two stacked tables, join key = name only ----
    ws3 = wb.create_sheet("Assessment")
    ws3.cell(row=1, column=1, value="Baseline Assessment - June 2025").font = Font(bold=True, italic=True)
    base_headers = ["Roll No", "Student Name", "Class", "Reading Level", "Math Score (/20)"]
    for c, h in enumerate(base_headers, start=1):
        ws3.cell(row=2, column=c, value=h).font = Font(bold=True)
    reading_levels = ["Beginner", "Letter", "Word", "Paragraph", "Story"]
    baseline_rows = []
    r = 3
    for i, row in enumerate(enrol_rows, start=1):
        roll = f"R{i:03d}"
        rl = rng.choice(reading_levels)
        ms = rng.randint(2, 18)
        ws3.cell(row=r, column=1, value=roll)
        ws3.cell(row=r, column=2, value=row["name"])
        ws3.cell(row=r, column=3, value=row["class"])
        ws3.cell(row=r, column=4, value=rl)
        ws3.cell(row=r, column=5, value=ms)
        baseline_rows.append({"roll_no": roll, "name": row["name"], "class": row["class"],
                               "reading_level": rl, "math_score": ms})
        r += 1
    baseline_df = pd.DataFrame(baseline_rows)

    r += 1  # blank separator row
    endline_start_row = r
    ws3.cell(row=r, column=1, value="Endline Assessment - March 2026").font = Font(bold=True, italic=True)
    r += 1
    end_headers = ["Student Name", "Class", "Reading Level", "Math Score (/20)"]
    for c, h in enumerate(end_headers, start=1):
        ws3.cell(row=r, column=c, value=h).font = Font(bold=True)
    r += 1
    dropout_count = 28
    dropped = set(rng.sample(range(len(enrol_rows)), dropout_count))
    endline_rows = []
    for i, row in enumerate(enrol_rows):
        if i in dropped:
            continue
        base = baseline_rows[i]
        # most students improve; a few regress
        rl_idx = reading_levels.index(base["reading_level"])
        new_idx = min(len(reading_levels) - 1, rl_idx + rng.choice([0, 1, 1, 2]))
        rl = reading_levels[new_idx]
        ms = min(20, max(0, base["math_score"] + rng.randint(-2, 8)))
        ws3.cell(row=r, column=1, value=row["name"])
        ws3.cell(row=r, column=2, value=row["class"])
        ws3.cell(row=r, column=3, value=rl)
        ws3.cell(row=r, column=4, value=ms)
        endline_rows.append({"name": row["name"], "class": row["class"],
                              "reading_level": rl, "math_score": ms})
        r += 1
    endline_df = pd.DataFrame(endline_rows)

    for c in range(1, 8):
        ws.column_dimensions[get_column_letter(c)].width = 16
        ws2.column_dimensions[get_column_letter(c)].width = 12
        ws3.column_dimensions[get_column_letter(c)].width = 16

    fpath = os.path.join(d, "students_2025_26.xlsx")
    wb.save(fpath)
    record(org, "students_2025_26.xlsx", enrol_df, "Enrolment")
    record(org, "students_2025_26.xlsx", att_df, "Attendance")
    record(org, "students_2025_26.xlsx", baseline_df, "Assessment_Baseline")
    record(org, "students_2025_26.xlsx", endline_df, "Assessment_Endline")

    # ---- funder_report_Q1.xlsx: indicator tracker with totals row ----
    indicators = [
        ("Students enrolled", 320, 320),
        ("Students retained (no dropout)", 300, 292),
        ("Avg attendance %", 85, 81),
        ("Students at 'Story' reading level", 60, 54),
        ("Learning camps conducted", 12, 11),
        ("Parent meetings held", 6, 6),
        ("Teachers trained", 18, 16),
        ("Classrooms with library corner", 14, 12),
        ("Remedial sessions conducted", 40, 47),
        ("Students receiving scholarship", 25, 21),
        ("School infrastructure audits", 8, 8),
        ("Community awareness sessions", 10, 7),
    ]
    wb2 = Workbook()
    ws4 = wb2.active
    ws4.title = "Indicator Tracker"
    ws4.append(["Indicator", "Target", "Achieved", "% Achievement", "Remarks"])
    for c in ws4[1]:
        c.font = Font(bold=True)
    remarks_pool = ["On track", "Delay due to monsoon", "Ahead of target",
                    "Funder query pending", "Data pending from field",
                    "Revised target under discussion"]
    ind_rows = []
    tgt_sum = ach_sum = 0
    for name_, tgt, ach in indicators:
        pct = round(100 * ach / tgt, 1)
        rem = rng.choice(remarks_pool)
        ws4.append([name_, tgt, ach, pct, rem])
        ind_rows.append({"indicator": name_, "target": tgt, "achieved": ach,
                          "pct_achievement": pct, "remarks": rem})
        tgt_sum += tgt
        ach_sum += ach
    ws4.append(["TOTAL", tgt_sum, ach_sum, round(100 * ach_sum / tgt_sum, 1), ""])
    for c in ws4[ws4.max_row]:
        c.font = Font(bold=True)
    ind_df = pd.DataFrame(ind_rows)
    fpath2 = os.path.join(d, "funder_report_Q1.xlsx")
    wb2.save(fpath2)
    record(org, "funder_report_Q1.xlsx", ind_df, "Indicator Tracker")
    return org


# ===========================================================================
# ORG 2: livelihoods / SHG NGO
# ===========================================================================

def gen_livelihoods():
    org = "swayam-mahila"
    d = os.path.join(OUT, org)
    ensure_dir(d)
    rng = random.Random(f"{SEED}-livelihoods")

    n_shg = 180
    shg_rows = []
    shg_ids = []
    for i in range(1, n_shg + 1):
        shg_id = f"SHG-{i:04d}"
        shg_ids.append(shg_id)
        village = rng.choice(VILLAGES)
        bank = rng.choice(BANKS)
        n_members = rng.randint(8, 15)
        shg_rows.append({
            "shg_id": shg_id,
            "shg_name": f"{rng.choice(FEMALE_FIRST)} Mahila Bachat Gat {i}",
            "village": village,
            "block": rng.choice(BLOCKS),
            "date_formed": dd_mm_yyyy(rand_date(rng, datetime(2018, 1, 1), datetime(2024, 12, 31))),
            "bank_name": bank,
            "bank_account": bank_account(rng),
            "ifsc": ifsc(rng, bank),
            "members": n_members,
            "president_name": f"{rng.choice(FEMALE_FIRST)} {rng.choice(SURNAMES)}",
            "president_phone": phone(rng),
        })
    shg_df = pd.DataFrame(shg_rows)
    csv_path = os.path.join(d, "shg_master.csv")
    shg_df.to_csv(csv_path, index=False)
    record(org, "shg_master.csv", shg_df)

    # ---- loan_ledger.xlsx: wide by month, outstanding balances ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Ledger"
    headers = ["SHG ID", "SHG Name", "Loan Amount Disbursed", "Disbursement Date",
               "Purpose"] + MONTH_ABBR_2025 + ["Outstanding Balance"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    purposes = ["Dairy", "Tailoring unit", "Kirana shop", "Goat rearing",
                "Papad making", "Agri input purchase", "Poultry"]
    ledger_rows = []
    # only SHGs that took loans (not all)
    borrowers = rng.sample(shg_rows, 140)
    for row in borrowers:
        amt = rng.choice([20000, 30000, 40000, 50000, 60000, 75000, 100000])
        disb_date = rand_date(rng, datetime(2025, 1, 1), datetime(2025, 6, 30))
        purpose = rng.choice(purposes)
        n_emi = rng.randint(6, 18)
        emi = round(amt / n_emi, 2)
        remaining = amt
        monthly_paid = []
        for m in MONTH_ABBR_2025:
            paid_this_month = rng.random() < 0.85
            pay = emi if (paid_this_month and remaining > 0) else 0
            pay = min(pay, remaining)
            remaining = round(remaining - pay, 2)
            monthly_paid.append(round(pay, 2))
        ws.append([row["shg_id"], row["shg_name"], amt, dd_mm_yyyy(disb_date),
                   purpose, *monthly_paid, remaining])
        ledger_rows.append({
            "shg_id": row["shg_id"], "shg_name": row["shg_name"],
            "loan_amount": amt, "disbursement_date": dd_mm_yyyy(disb_date),
            "purpose": purpose, **dict(zip(MONTH_ABBR_2025, monthly_paid)),
            "outstanding_balance": remaining,
        })
    ledger_df = pd.DataFrame(ledger_rows)
    ll_path = os.path.join(d, "loan_ledger.xlsx")
    wb.save(ll_path)
    record(org, "loan_ledger.xlsx", ledger_df, "Loan Ledger")

    # ---- zoho_books_invoices_export.csv: odd Zoho column order ----
    n_inv = 260
    statuses = ["Paid", "Overdue", "Sent", "Draft", "Partially Paid"]
    status_weights = [0.55, 0.15, 0.15, 0.05, 0.10]
    inv_rows = []
    for i in range(1, n_inv + 1):
        cust = rng.choice(shg_rows)
        total = rng.choice([5000, 7500, 10000, 12500, 15000, 20000, 25000])
        status = rng.choices(statuses, weights=status_weights)[0]
        balance = 0.0 if status == "Paid" else round(total * rng.uniform(0.2, 1.0), 2)
        inv_date = rand_date(rng, datetime(2025, 1, 1), datetime(2025, 9, 15))
        due_date = inv_date + timedelta(days=30)
        inv_rows.append({
            "Invoice Number": f"INV-2025-{i:05d}",
            "Customer ID": f"CUST-{1000 + (i % n_shg)}",
            "Customer Name": cust["shg_name"],
            "Invoice Date": dd_mm_yyyy(inv_date),
            "Due Date": dd_mm_yyyy(due_date),
            "Status": status,
            "Total": inr(total),
            "Balance": inr(balance),
            "Currency Code": "INR",
            "Sales Order Number": "",
            "Notes": rng.choice(["Livelihoods material grant", "Revolving fund",
                                   "Skill training kit invoice", ""]),
        })
    inv_df = pd.DataFrame(inv_rows)
    inv_path = os.path.join(d, "zoho_books_invoices_export.csv")
    inv_df.to_csv(inv_path, index=False)
    record(org, "zoho_books_invoices_export.csv", inv_df)
    return org


# ===========================================================================
# ORG 3: health NGO (maternal health / ANC tracking)
# ===========================================================================

def gen_health():
    org = "arogya-jyoti"
    d = os.path.join(OUT, org)
    ensure_dir(d)
    rng = random.Random(f"{SEED}-health")

    n = 340
    enumerators = [f"{rng.choice(FEMALE_FIRST)} {rng.choice(SURNAMES)}" for _ in range(6)]
    kobo_rows = []
    base_dt = datetime(2025, 3, 1)
    for i in range(1, n + 1):
        fn, sn = full_name(rng)
        village = rng.choice(VILLAGES)
        enum = rng.choice(enumerators)
        sub_dt = base_dt + timedelta(days=rng.randint(0, 180), hours=rng.randint(7, 18),
                                       minutes=rng.randint(0, 59))
        start_dt = sub_dt - timedelta(minutes=rng.randint(8, 25))
        anc1 = rng.random() < 0.95
        anc2 = anc1 and rng.random() < 0.82
        anc3 = anc2 and rng.random() < 0.68
        anc4 = anc3 and rng.random() < 0.5
        kobo_rows.append({
            "_id": 900000 + i,
            "_uuid": f"a1b2c3{i:06d}-uuid-4f2a-{rng.randint(1000,9999)}",
            "start": start_dt.strftime("%Y-%m-%dT%H:%M:%S"),
            "end": sub_dt.strftime("%Y-%m-%dT%H:%M:%S"),
            "today": sub_dt.strftime("%Y-%m-%d"),
            "enumerator/name": enum,
            "hh/village": village,
            "woman/name": f"{fn} {sn}",
            "woman/age": rng.randint(18, 38),
            "woman/phone": phone(rng),
            "visit/anc1": "yes" if anc1 else "",
            "visit/anc2": "yes" if anc2 else "",
            "visit/anc3": "yes" if anc3 else "",
            "visit/anc4": "yes" if anc4 else "",
            "visit/hb_level": round(rng.uniform(8.5, 13.5), 1) if anc1 else "",
            "visit/bp_systolic": rng.randint(96, 138) if anc1 else "",
            "visit/institutional_delivery_planned": rng.choice(["yes", "no", "not_sure"]),
            "_submission_time": sub_dt.strftime("%Y-%m-%dT%H:%M:%S"),
            "_validation_status": rng.choices(
                ["validation_status_approved", "validation_status_on_hold", ""],
                weights=[0.75, 0.15, 0.10])[0],
        })
    kobo_df = pd.DataFrame(kobo_rows)
    kobo_path = os.path.join(d, "anc_tracking_kobo_export.csv")
    kobo_df.to_csv(kobo_path, index=False)
    record(org, "anc_tracking_kobo_export.csv", kobo_df)

    # ---- camp_register.xlsx: title row, merged cells, Hinglish notes ----
    n_camps = 24
    camp_rows = []
    wb = Workbook()
    ws = wb.active
    ws.title = "Camp Register"
    for camp_no in range(1, n_camps + 1):
        village = rng.choice(VILLAGES)
        camp_date = rand_date(rng, datetime(2025, 3, 1), datetime(2025, 9, 15))
        attendees = rng.randint(15, 60)
        women = rng.randint(int(attendees * 0.4), attendees)
        men = attendees - women
        for _ in range(1):
            camp_rows.append({
                "camp_no": camp_no, "village": village,
                "camp_date": dd_mm_yyyy(camp_date),
                "attendees": attendees, "women": women, "men": men,
            })
    # write as title + merged-cell blocks, one block per camp (realistic mess:
    # a "register" pasted per camp rather than one tidy table)
    r = 1
    for row in camp_rows:
        ws.cell(row=r, column=1,
                value=f"Health Camp #{row['camp_no']} - {row['village']} - {row['camp_date']}").font = Font(bold=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        ws.append(["S.No", "Name", "Age", "Gender", "Village", "Notes"])
        for c in ws[r]:
            c.font = Font(bold=True)
        r += 1
        for s in range(1, row["attendees"] + 1):
            gender = "F" if s <= row["women"] else "M"
            fn, sn = gendered_name(rng, gender)
            ws.append([s, f"{fn} {sn}", rng.randint(18, 65), gender, row["village"],
                       rng.choice(HINGLISH_NOTES) if rng.random() < 0.3 else ""])
            r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=r, column=2, value=row["attendees"])
        ws.cell(row=r, column=3, value=f"W:{row['women']} M:{row['men']}")
        r += 2  # blank spacer row between camps
    camp_df = pd.DataFrame(camp_rows)
    camp_path = os.path.join(d, "camp_register.xlsx")
    wb.save(camp_path)
    record(org, "camp_register.xlsx", camp_df, "Camp Register")
    return org


# ===========================================================================
# ORG 4: disability NGO
# ===========================================================================

def gen_disability():
    org = "saksham-kalyan"
    d = os.path.join(OUT, org)
    ensure_dir(d)
    rng = random.Random(f"{SEED}-disability")

    n = 210
    wb = Workbook()
    ws = wb.active
    ws.title = "Beneficiary Master"
    headers = ["UDID Number", "Name", "Gender", "DOB", "Disability Type",
               "Disability %", "Village", "Guardian Name", "Guardian Phone",
               "Assistive Device Issued", "Device Issue Date", "Registration Date"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    ben_rows = []
    for i in range(1, n + 1):
        gender = rng.choice(["M", "F"])
        fn, sn = gendered_name(rng, gender)
        dob = rand_date(rng, datetime(1965, 1, 1), datetime(2018, 12, 31))
        uid = udid(rng, dob.year)
        dis_type = rng.choice(DISABILITY_TYPES)
        dis_pct = rng.randint(40, 100)
        device = rng.choice(DEVICE_TYPES) if rng.random() < 0.7 else ""
        # deliberately mixed date formats between rows/columns, a common mess
        issue_date = rand_date(rng, datetime(2023, 1, 1), datetime(2025, 8, 1)) if device else None
        reg_date = rand_date(rng, datetime(2019, 1, 1), datetime(2025, 6, 1))
        issue_str = ""
        if device:
            issue_str = yyyy_mm_dd(issue_date) if i % 3 == 0 else dd_mm_yyyy(issue_date)
        reg_str = dd_mm_yyyy(reg_date) if i % 4 != 0 else yyyy_mm_dd(reg_date)
        row = {
            "udid": uid, "name": f"{fn} {sn}", "gender": gender,
            "dob": dd_mm_yyyy(dob) if i % 5 else yyyy_mm_dd(dob),
            "disability_type": dis_type, "disability_pct": dis_pct,
            "village": rng.choice(VILLAGES),
            "guardian_name": f"{rng.choice(MALE_FIRST)} {sn}",
            "guardian_phone": phone(rng),
            "device_issued": device, "device_issue_date": issue_str,
            "registration_date": reg_str,
        }
        ws.append([row["udid"], row["name"], row["gender"], row["dob"],
                   row["disability_type"], row["disability_pct"], row["village"],
                   row["guardian_name"], row["guardian_phone"],
                   row["device_issued"], row["device_issue_date"],
                   row["registration_date"]])
        ben_rows.append(row)
    ben_df = pd.DataFrame(ben_rows)
    ben_path = os.path.join(d, "beneficiary_master.xlsx")
    wb.save(ben_path)
    record(org, "beneficiary_master.xlsx", ben_df, "Beneficiary Master")

    # ---- device_inventory.csv: stock in/out ledger ----
    n_tx = 260
    tx_rows = []
    stock = {t: rng.randint(20, 60) for t in DEVICE_TYPES}
    start = datetime(2025, 1, 5)
    for i in range(1, n_tx + 1):
        device = rng.choice(DEVICE_TYPES)
        tx_type = rng.choices(["IN", "OUT"], weights=[0.35, 0.65])[0]
        qty = rng.randint(1, 5) if tx_type == "OUT" else rng.randint(5, 30)
        tx_date = start + timedelta(days=rng.randint(0, 230))
        stock[device] += qty if tx_type == "IN" else -qty
        tx_rows.append({
            "txn_id": f"DEV-{i:05d}",
            "date": dd_mm_yyyy(tx_date),
            "device_type": device,
            "txn_type": tx_type,
            "qty": qty,
            "balance_after": stock[device],
            "reference": (f"Issued to {full_name(rng)[0]} {rng.choice(SURNAMES)}"
                          if tx_type == "OUT" else
                          f"Received from {rng.choice(['State disability office','Donor grant','Rotary club donation','Purchase order'])}"),
        })
    tx_df = pd.DataFrame(tx_rows)
    tx_path = os.path.join(d, "device_inventory.csv")
    tx_df.to_csv(tx_path, index=False)
    record(org, "device_inventory.csv", tx_df)
    return org


# ===========================================================================
# ORG 5: funder-facing fundraising team
# ===========================================================================

def gen_fundraising():
    org = "asha-kiran"
    d = os.path.join(OUT, org)
    ensure_dir(d)
    rng = random.Random(f"{SEED}-fundraising")

    # ---- donor_crm_zoho_export.csv: Zoho CRM lead/contact export ----
    n_leads = 300
    donor_pool = []  # (first, last, email, phone) reused for repeat-donor case
    crm_rows = []
    owners = [f"{rng.choice(MALE_FIRST + FEMALE_FIRST)} {rng.choice(SURNAMES)}" for _ in range(4)]
    for i in range(1, n_leads + 1):
        fn, sn = full_name(rng)
        email = f"{fn.lower()}.{sn.lower()}{rng.randint(1,99)}@{rng.choice(['gmail.com','yahoo.com','outlook.com'])}"
        ph = phone(rng)
        donor_pool.append((fn, sn, email, ph))
        created = rand_date(rng, datetime(2025, 1, 1), datetime(2025, 8, 20))
        modified = created + timedelta(days=rng.randint(0, 60))
        status = rng.choices(LEAD_STATUS, weights=[0.2, 0.2, 0.15, 0.25, 0.15, 0.05])[0]
        crm_rows.append({
            "Lead Owner": rng.choice(owners),
            "First Name": fn,
            "Last Name": sn,
            "Email": email,
            "Phone": ph,
            "Lead Source": rng.choice(LEAD_SOURCES),
            "Lead Status": status,
            "Annual Revenue": rng.choice(["", "", 500000, 1200000, 2500000, 8000000]),
            "Created Time": created.strftime("%Y-%m-%d %H:%M:%S"),
            "Modified Time": modified.strftime("%Y-%m-%d %H:%M:%S"),
            "Tag": rng.choice(["Individual Donor", "Corporate", "Alumni", "Recurring", ""]),
            "Description": rng.choice(["Met at annual gala", "Referred by board member",
                                          "Responded to email campaign", "", "CSR enquiry"]),
        })
    crm_df = pd.DataFrame(crm_rows)
    crm_path = os.path.join(d, "donor_crm_zoho_export.csv")
    crm_df.to_csv(crm_path, index=False)
    record(org, "donor_crm_zoho_export.csv", crm_df)

    # ---- razorpay_payments.csv: some donors give more than once ----
    n_pay = 420
    repeat_donors = rng.sample(donor_pool, 60)
    pay_rows = []
    epoch0 = datetime(2025, 1, 1)
    for i in range(1, n_pay + 1):
        # 30% of payments come from the repeat-donor pool (so several donors
        # end up with 2+ payments); rest are one-off first-time donors
        if rng.random() < 0.4:
            fn, sn, email, ph = rng.choice(repeat_donors)
        else:
            fn, sn = full_name(rng)
            email = f"{fn.lower()}.{sn.lower()}{rng.randint(1,999)}@{rng.choice(['gmail.com','yahoo.com','outlook.com'])}"
            ph = phone(rng)
        amount_inr = rng.choice([500, 1000, 1500, 2000, 2500, 5000, 10000, 25000])
        created_dt = rand_date(rng, datetime(2025, 1, 1), datetime(2025, 8, 20))
        status = rng.choices(["captured", "failed", "refunded"], weights=[0.9, 0.07, 0.03])[0]
        pay_rows.append({
            "id": f"pay_{rng.randint(10**12, 10**13-1):x}",
            "amount": amount_inr * 100,  # paise
            "currency": "INR",
            "status": status,
            "method": rng.choice(["upi", "card", "netbanking", "wallet"]),
            "email": email,
            "contact": f"+91{ph}",
            "created_at": int(created_dt.timestamp()),
            "notes.campaign": rng.choice(CAMPAIGN_NAMES),
            "fee": round(amount_inr * 0.02 * 100),
            "tax": round(amount_inr * 0.0036 * 100),
        })
    pay_df = pd.DataFrame(pay_rows)
    pay_path = os.path.join(d, "razorpay_payments.csv")
    pay_df.to_csv(pay_path, index=False)
    record(org, "razorpay_payments.csv", pay_df)

    # ---- 80g_receipts_2025.xlsx ----
    n_receipts = 200
    captured = pay_df[pay_df["status"] == "captured"].sample(
        n=min(n_receipts, (pay_df["status"] == "captured").sum()), random_state=SEED)
    wb = Workbook()
    ws = wb.active
    ws.title = "80G Receipts"
    ws.append(["Receipt No", "Donor Name", "PAN", "Amount", "Date", "Payment Mode",
               "80G Reg No", "Campaign"])
    for c in ws[1]:
        c.font = Font(bold=True)
    rec_rows = []
    for i, (_, prow) in enumerate(captured.iterrows(), start=1):
        rno = f"AK/80G/2025/{i:04d}"
        donor_name = None
        # find matching name for this email from donor_pool/repeat_donors
        for fn, sn, em, ph in donor_pool + repeat_donors:
            if em == prow["email"]:
                donor_name = f"{fn} {sn}"
                break
        if donor_name is None:
            donor_name = "Donor " + prow["email"].split("@")[0]
        amt = prow["amount"] / 100
        dt = datetime.fromtimestamp(prow["created_at"])
        row = {
            "receipt_no": rno, "donor_name": donor_name, "pan": pan(rng),
            "amount": amt, "date": dd_mm_yyyy(dt), "mode": prow["method"].upper(),
            "reg_no": "AAATA1234B/80G/2021-22", "campaign": prow["notes.campaign"],
        }
        ws.append([row["receipt_no"], row["donor_name"], row["pan"], row["amount"],
                   row["date"], row["mode"], row["reg_no"], row["campaign"]])
        rec_rows.append(row)
    rec_df = pd.DataFrame(rec_rows)
    rec_path = os.path.join(d, "80g_receipts_2025.xlsx")
    wb.save(rec_path)
    record(org, "80g_receipts_2025.xlsx", rec_df, "80G Receipts")
    return org


# ===========================================================================
# ORG 6: ops / admin
# ===========================================================================

def gen_ops():
    org = "prayas-seva-sangh"
    d = os.path.join(OUT, org)
    ensure_dir(d)
    rng = random.Random(f"{SEED}-ops")

    # ---- expense_vouchers_2025.xlsx: vouchers + totals per month + pivot sheet ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Vouchers"
    ws.append(["Voucher No", "Date", "Head", "Sub-head", "Amount", "Paid To",
               "Mode", "Project Code", "Approved By"])
    for c in ws[1]:
        c.font = Font(bold=True)
    approvers = [f"{rng.choice(MALE_FIRST+FEMALE_FIRST)} {rng.choice(SURNAMES)}" for _ in range(3)]
    project_codes = ["PRJ-ED-01", "PRJ-HL-02", "PRJ-LV-03", "PRJ-RD-04", "PRJ-ADM-00"]
    n_v = 480
    v_rows = []
    monthly_totals = {m: 0.0 for m in MONTHS_2025}
    head_actuals = {h: 0.0 for h in EXPENSE_HEADS}
    vno = 1
    for i in range(n_v):
        head = rng.choice(list(EXPENSE_HEADS.keys()))
        sub = rng.choice(EXPENSE_HEADS[head])
        amt = round(rng.uniform(300, 18000), 2)
        vdate = rand_date(rng, datetime(2025, 4, 1), datetime(2025, 9, 30))
        month_name = vdate.strftime("%B")
        mode = rng.choice(["Cash", "Bank Transfer", "Cheque", "UPI"])
        vid = f"V-{vno:05d}"
        vno += 1
        row = {"voucher_no": vid, "date": dd_mm_yyyy(vdate), "head": head,
               "sub_head": sub, "amount": amt,
               "paid_to": f"{rng.choice(MALE_FIRST+FEMALE_FIRST)} {rng.choice(SURNAMES)}",
               "mode": mode, "project_code": rng.choice(project_codes),
               "approved_by": rng.choice(approvers)}
        ws.append([row["voucher_no"], row["date"], row["head"], row["sub_head"],
                   row["amount"], row["paid_to"], row["mode"], row["project_code"],
                   row["approved_by"]])
        v_rows.append(row)
        if month_name in monthly_totals:
            monthly_totals[month_name] += amt
        head_actuals[head] += amt
    ws.append(["", "", "", "TOTAL", round(sum(monthly_totals.values()), 2), "", "", "", ""])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    v_df = pd.DataFrame(v_rows)
    record(org, "expense_vouchers_2025.xlsx", v_df, "Vouchers")

    # second sheet: "Budget vs Actual" pivot pasted as values (no formulas)
    ws2 = wb.create_sheet("Budget vs Actual")
    ws2.append(["Head", "Budget (FY25-26)", "Actual (Apr-Sep)", "Variance", "% Utilised"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    budgets = {"Programme": 350000, "Admin": 120000, "Salaries": 900000, "Logistics": 150000}
    bva_rows = []
    for head, budget in budgets.items():
        actual = round(head_actuals[head], 2)
        variance = round(actual - budget, 2)
        pct = round(100 * actual / budget, 1)
        ws2.append([head, budget, actual, variance, pct])
        bva_rows.append({"head": head, "budget": budget, "actual": actual,
                          "variance": variance, "pct_utilised": pct})
    bva_df = pd.DataFrame(bva_rows)
    record(org, "expense_vouchers_2025.xlsx", bva_df, "Budget vs Actual")
    ev_path = os.path.join(d, "expense_vouchers_2025.xlsx")
    wb.save(ev_path)

    # ---- staff_attendance_aug.xlsx: wide days 1..31, P/A/L/H codes ----
    n_staff = 42
    staff = [{"staff_id": f"STF-{i:03d}", "name": f"{rng.choice(MALE_FIRST+FEMALE_FIRST)} {rng.choice(SURNAMES)}",
              "designation": rng.choice(["Field Officer", "Programme Coordinator", "Driver",
                                           "Accountant", "Community Mobiliser", "Data Entry Operator"])}
             for i in range(1, n_staff + 1)]
    wb2 = Workbook()
    ws3 = wb2.active
    ws3.title = "Attendance Aug 2025"
    days = list(range(1, 32))
    ws3.append(["Staff ID", "Name", "Designation"] + [str(dd) for dd in days] + ["P", "A", "L", "H"])
    for c in ws3[1]:
        c.font = Font(bold=True)
    att_rows = []
    for s in staff:
        codes = []
        for dd in days:
            wd = datetime(2025, 8, dd).weekday()
            if wd == 6:
                codes.append("H")
            else:
                codes.append(rng.choices(["P", "A", "L"], weights=[0.88, 0.07, 0.05])[0])
        counts = {c: codes.count(c) for c in "PALH"}
        ws3.append([s["staff_id"], s["name"], s["designation"], *codes,
                    counts["P"], counts["A"], counts["L"], counts["H"]])
        att_rows.append({**s, **{f"d{dd}": codes[dd-1] for dd in days}, **counts})
    att2_df = pd.DataFrame(att_rows)
    sa_path = os.path.join(d, "staff_attendance_aug.xlsx")
    wb2.save(sa_path)
    record(org, "staff_attendance_aug.xlsx", att2_df, "Attendance Aug 2025")

    # ---- vehicle_log.csv ----
    n_trips = 300
    vehicles = ["MH12AB1234 (Bolero)", "MH14CD5678 (Scorpio)", "MH12EF9012 (Activa)"]
    trip_rows = []
    for i in range(1, n_trips + 1):
        vdate = rand_date(rng, datetime(2025, 4, 1), datetime(2025, 9, 30))
        start_km = rng.randint(10000, 80000)
        km = rng.randint(15, 180)
        trip_rows.append({
            "trip_id": f"TRP-{i:05d}", "date": dd_mm_yyyy(vdate),
            "vehicle": rng.choice(vehicles),
            "driver": rng.choice([s["name"] for s in staff if s["designation"] == "Driver"] or ["Ramesh Pawar"]),
            "purpose": rng.choice(["Field visit", "Material delivery", "Office errand",
                                     "Camp transport", "Meeting"]),
            "start_km": start_km, "end_km": start_km + km, "km_travelled": km,
            "fuel_cost": round(km * rng.uniform(6.5, 8.5), 2),
        })
    trip_df = pd.DataFrame(trip_rows)
    vl_path = os.path.join(d, "vehicle_log.csv")
    trip_df.to_csv(vl_path, index=False)
    record(org, "vehicle_log.csv", trip_df)
    return org


# ===========================================================================
# ORG 7: field team (WhatsApp export + village visit log)
# ===========================================================================

def gen_field():
    org = "gram-sudhar"
    d = os.path.join(OUT, org)
    ensure_dir(d)
    rng = random.Random(f"{SEED}-field")

    members = [f"{rng.choice(MALE_FIRST+FEMALE_FIRST)} {rng.choice(SURNAMES)}" for _ in range(9)]
    admin = members[0]
    chatter = [
        "kal meeting hai sarpanch office mein", "photo bhejta hu thodi der mein",
        "route confirm ho gaya", "gaadi 9 baje niklegi", "material load ho gaya",
        "sab volunteers ready hain", "field visit start", "lunch break le rahe hain",
        "signal weak hai yaha", "next stop {v}", "session complete ho gaya",
        "attendance sheet upload kar diya", "rasta kharab tha aaj",
        "parents se baat hui achi rahi", "ok noted", "thik hai", "great work team",
        "kal 10 baje milte hain", "van mein jagah nahi bachi", "backup plan chahiye",
    ]
    daily_count_templates = [
        "Aaj {v} mein {n} bacche aaye session ke liye",
        "{v} update: total {n} bacche aaye aaj",
        "Aaj ki attendance - {n} bacche, sab active the",
        "{n} bacche the aaj {v} centre par",
        "Session summary {v}: {n} children present today",
    ]

    start_date = datetime(2025, 5, 1)
    lines = []
    daily_counts = []  # (date_str_ddmmyy, village, n)
    d_cursor = start_date
    n_days = 0
    while n_days < 95:  # ~3+ months of field days
        wd = d_cursor.weekday()
        if wd != 6:  # skip Sundays
            village = rng.choice(VILLAGES)
            n_children = rng.randint(8, 45)
            n_msgs = rng.randint(3, 9)
            day_msgs = []
            count_msg_idx = rng.randint(0, n_msgs - 1)
            for mi in range(n_msgs):
                sender = rng.choice(members)
                hh = rng.randint(9, 19)
                mm = rng.randint(0, 59)
                ts = d_cursor.replace(hour=hh, minute=mm)
                if mi == count_msg_idx:
                    tmpl = rng.choice(daily_count_templates)
                    text = tmpl.format(v=village, n=n_children)
                    daily_counts.append((ts.strftime("%d/%m/%y"), village, n_children))
                elif rng.random() < 0.12:
                    text = f"<Media omitted>"
                elif rng.random() < 0.08:
                    text = f"call me on {phone(rng)} if any issue"
                else:
                    text = rng.choice(chatter).format(v=village)
                day_msgs.append((ts, sender, text))
            day_msgs.sort()
            for ts, sender, text in day_msgs:
                lines.append(f"{ts.strftime('%d/%m/%y')}, {ts.strftime('%H:%M')} - {sender}: {text}")
            n_days += 1
        d_cursor += timedelta(days=1)

    header = (f"{start_date.strftime('%d/%m/%y')}, 08:00 - Messages and calls are end-to-end "
              f"encrypted. No one outside of this chat, not even WhatsApp, can read or listen to them.\n"
              f"{start_date.strftime('%d/%m/%y')}, 08:00 - {admin} created group \"Gram Sudhar Field Team\"\n")
    txt_path = os.path.join(d, "whatsapp_field_group_export.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(header)
        f.write("\n".join(lines))
        f.write("\n")

    counts_df = pd.DataFrame(daily_counts, columns=["date_ddmmyy", "village", "children_count"])
    record(org, "whatsapp_field_group_export.txt", counts_df, "daily_counts_extracted")

    # ---- village_visit_log.csv ----
    n_visits = 320
    purposes = ["Household survey", "SHG meeting", "School follow-up",
                "Water source inspection", "Health camp coordination",
                "Grievance redressal", "Ration distribution check"]
    visit_rows = []
    for i in range(1, n_visits + 1):
        vdate = rand_date(rng, datetime(2025, 4, 1), datetime(2025, 9, 15))
        village = rng.choice(VILLAGES)
        visit_rows.append({
            "visit_id": f"VV-{i:05d}", "date": dd_mm_yyyy(vdate), "village": village,
            "staff_name": rng.choice(members), "purpose": rng.choice(purposes),
            "households_met": rng.randint(2, 25),
            "issues_raised": rng.choice(["None", "Water shortage", "Road damage",
                                            "Ration delay", "School infra", "None", "None"]),
        })
    visit_df = pd.DataFrame(visit_rows)
    vv_path = os.path.join(d, "village_visit_log.csv")
    visit_df.to_csv(vv_path, index=False)
    record(org, "village_visit_log.csv", visit_df)
    return org


# ===========================================================================
# ORG 8: rural development NGO - monthly MIS workbook with header drift
# ===========================================================================

def gen_rural_mis():
    org = "krishi-jal-vikas"
    d = os.path.join(OUT, org)
    ensure_dir(d)
    rng = random.Random(f"{SEED}-ruralmis")

    villages = rng.sample(VILLAGES, 16)
    wb = Workbook()
    wb.remove(wb.active)

    # header labels drift month to month for the same indicator (realistic
    # mess: whoever filled the sheet that month used their own phrasing)
    hh_labels = {"April": "No. of HH", "May": "No. of HH", "June": "Households reached",
                 "July": "Households reached", "August": "HH Covered", "September": "Households reached"}
    shg_labels = {"April": "SHGs formed", "May": "SHGs formed", "June": "SHGs formed",
                  "July": "New SHGs", "August": "New SHGs", "September": "SHGs formed (cumulative)"}
    wc_labels = {"April": "Water structures built", "May": "Water structures built",
                 "June": "Water conservation structures", "July": "Water conservation structures",
                 "August": "WC Structures", "September": "Water conservation structures"}
    farmer_labels = {m: "Farmers trained" for m in MONTHS_2025}

    monthly_frames = {}
    base_hh = {v: rng.randint(80, 240) for v in villages}
    for m in MONTHS_2025:
        ws = wb.create_sheet(m[:3])
        cols = ["Village", hh_labels[m], shg_labels[m], wc_labels[m], farmer_labels[m], "Remarks"]
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True)
        month_rows = []
        for v in villages:
            growth = MONTHS_2025.index(m)
            hh = base_hh[v] + rng.randint(0, 15) * growth
            shg = rng.randint(0, 3)
            wc = rng.randint(0, 2)
            farmers = rng.randint(5, 40)
            remarks = rng.choice(["", "", "Monsoon delay", "On track", "Target revised",
                                    "Data pending"])
            ws.append([v, hh, shg, wc, farmers, remarks])
            month_rows.append({"village": v, "households_reached": hh,
                                 "shgs_formed": shg, "water_structures": wc,
                                 "farmers_trained": farmers, "remarks": remarks,
                                 "month": m})
        # totals row per sheet
        mdf = pd.DataFrame(month_rows)
        ws.append(["TOTAL", int(mdf["households_reached"].sum()),
                    int(mdf["shgs_formed"].sum()), int(mdf["water_structures"].sum()),
                    int(mdf["farmers_trained"].sum()), ""])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
        monthly_frames[m] = mdf
        record(org, "mis_apr_sep_2025.xlsx", mdf, m)

    # Consolidated sheet: formulas computed in python and pasted as plain
    # values (mirrors "copy > paste special > values" real-world habit)
    ws_c = wb.create_sheet("Consolidated")
    ws_c.append(["Village", "Total HH Reached (Apr-Sep)", "Total SHGs Formed",
                 "Total Water Structures", "Total Farmers Trained"])
    for c in ws_c[1]:
        c.font = Font(bold=True)
    consolidated_rows = []
    for v in villages:
        tot_hh = sum(monthly_frames[m].set_index("village").loc[v, "households_reached"] for m in MONTHS_2025)
        tot_shg = sum(monthly_frames[m].set_index("village").loc[v, "shgs_formed"] for m in MONTHS_2025)
        tot_wc = sum(monthly_frames[m].set_index("village").loc[v, "water_structures"] for m in MONTHS_2025)
        tot_f = sum(monthly_frames[m].set_index("village").loc[v, "farmers_trained"] for m in MONTHS_2025)
        ws_c.append([v, int(tot_hh), int(tot_shg), int(tot_wc), int(tot_f)])
        consolidated_rows.append({"village": v, "total_hh_reached": int(tot_hh),
                                    "total_shgs_formed": int(tot_shg),
                                    "total_water_structures": int(tot_wc),
                                    "total_farmers_trained": int(tot_f)})
    cons_df = pd.DataFrame(consolidated_rows)
    record(org, "mis_apr_sep_2025.xlsx", cons_df, "Consolidated")

    # Dashboard sheet: a pasted pivot (indicator x month grid), typical of
    # "someone built a PivotTable, then pasted values into a new tab"
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash.cell(row=1, column=1, value="Programme Dashboard (pasted from PivotTable)").font = Font(bold=True, italic=True)
    ws_dash.append(["Indicator"] + MONTHS_2025)
    for c in ws_dash[2]:
        c.font = Font(bold=True)
    dash_rows = []
    for label, key in [("Households Reached", "households_reached"),
                        ("SHGs Formed", "shgs_formed"),
                        ("Water Structures", "water_structures"),
                        ("Farmers Trained", "farmers_trained")]:
        vals = [int(monthly_frames[m][key].sum()) for m in MONTHS_2025]
        ws_dash.append([label] + vals)
        dash_rows.append({"indicator": label, **dict(zip(MONTHS_2025, vals))})
    dash_df = pd.DataFrame(dash_rows)
    record(org, "mis_apr_sep_2025.xlsx", dash_df, "Dashboard")

    mis_path = os.path.join(d, "mis_apr_sep_2025.xlsx")
    wb.save(mis_path)
    return org


# ===========================================================================
# cases.json construction
# ===========================================================================

def _g(org, fname, sheet=None):
    return GENERATED[org][fname][sheet]


def _parse_inr(s: str) -> float:
    return float(str(s).replace("₹", "").replace(",", ""))


def build_cases() -> list[dict]:
    cases = []

    def ask(cid, org, files, prompt, expect, expected_rows=None, order_matters=None):
        c = {"id": cid, "lane": "ask", "org": org, "files": files,
             "prompt": prompt, "expect": expect}
        if expected_rows is not None:
            c["expected_rows"] = expected_rows
        if order_matters is not None:
            c["order_matters"] = order_matters
        cases.append(c)

    def build(cid, org, files, prompt, expect):
        cases.append({"id": cid, "lane": "build", "org": org, "files": files,
                       "prompt": prompt, "expect": expect})

    def page(cid, org, files, prompt, expect):
        cases.append({"id": cid, "lane": "page", "org": org, "files": files,
                       "prompt": prompt, "expect": expect})

    # ---- education ----
    enrol = _g("sunrise-shiksha", "students_2025_26.xlsx", "Enrolment")
    baseline = _g("sunrise-shiksha", "students_2025_26.xlsx", "Assessment_Baseline")
    endline = _g("sunrise-shiksha", "students_2025_26.xlsx", "Assessment_Endline")
    att = _g("sunrise-shiksha", "students_2025_26.xlsx", "Attendance")
    ind = _g("sunrise-shiksha", "funder_report_Q1.xlsx", "Indicator Tracker")

    ask("ask-01", "sunrise-shiksha", ["students_2025_26.xlsx#Enrolment"],
        "How many students are enrolled at Sunrise Shiksha for 2025-26?",
        "A single count matching the Enrolment sheet row count (the 2-row "
        "merged group header above the real column headers must not be "
        "counted as a data row).",
        expected_rows=[{"count": int(len(enrol))}], order_matters=False)

    dropped = sorted(set(baseline["name"]) - set(endline["name"]))
    ask("ask-02", "sunrise-shiksha", ["students_2025_26.xlsx#Assessment_Baseline",
                                        "students_2025_26.xlsx#Assessment_Endline"],
        "Which students dropped out between the baseline and endline assessment?",
        "Students present in the Baseline table but absent from the Endline "
        "table below it. The two blocks share no ID scheme (Roll No vs none) "
        "so a correct answer must join on student name, not ID. "
        f"Total dropouts: {len(dropped)}.",
        expected_rows=[{"name": n} for n in dropped[:3]], order_matters=False)

    ind_behind = ind[ind["achieved"] < ind["target"]].copy()
    ind_behind["shortfall"] = ind_behind["target"] - ind_behind["achieved"]
    ind_behind = ind_behind.sort_values("shortfall", ascending=False)
    ask("ask-03", "sunrise-shiksha", ["funder_report_Q1.xlsx#Indicator Tracker"],
        "Which indicators are behind target in the Q1 funder report?",
        "Rows where Achieved < Target, excluding the bottom TOTAL row. "
        "Top rows ranked by shortfall (Target - Achieved), descending.",
        expected_rows=[{"indicator": r["indicator"], "target": int(r["target"]),
                          "achieved": int(r["achieved"])}
                         for _, r in ind_behind.head(3).iterrows()],
        order_matters=True)

    ask("ask-04", "sunrise-shiksha", ["students_2025_26.xlsx#Attendance"],
        "What was total student attendance (days present) across all classes in June 2025?",
        "Sum of the 'Jun-25' column in the wide-by-month Attendance sheet "
        "(Attendance ID differs from the Enrolment sheet's Student ID; join "
        "by name/class if the tool needs to combine the two sheets).",
        expected_rows=[{"month": "Jun-25", "total_days_present": int(att["Jun-25"].sum())}],
        order_matters=False)

    # ---- livelihoods / SHG ----
    shg = _g("swayam-mahila", "shg_master.csv")
    ledger = _g("swayam-mahila", "loan_ledger.xlsx", "Loan Ledger")
    inv = _g("swayam-mahila", "zoho_books_invoices_export.csv")

    top_village = shg["village"].value_counts().idxmax()
    ask("ask-05", "swayam-mahila", ["shg_master.csv"],
        f"How many SHGs are registered in {top_village} village?",
        "Count of shg_master.csv rows where village equals the named village.",
        expected_rows=[{"village": top_village,
                          "count": int((shg["village"] == top_village).sum())}],
        order_matters=False)

    top_outstanding = ledger.sort_values("outstanding_balance", ascending=False).head(3)
    ask("ask-06", "swayam-mahila", ["loan_ledger.xlsx#Loan Ledger"],
        "Which SHGs have the highest outstanding loan balance as of September?",
        "Top rows of loan_ledger.xlsx ranked by 'Outstanding Balance' "
        "descending (the sheet is wide-by-month; balance is the last column).",
        expected_rows=[{"shg_id": r["shg_id"], "shg_name": r["shg_name"],
                          "outstanding_balance": float(r["outstanding_balance"])}
                         for _, r in top_outstanding.iterrows()],
        order_matters=True)

    inv2 = inv.copy()
    inv2["balance_num"] = inv2["Balance"].map(_parse_inr)
    overdue = inv2[inv2["Status"] == "Overdue"].sort_values("balance_num", ascending=False).head(3)
    ask("ask-07", "swayam-mahila", ["zoho_books_invoices_export.csv"],
        "Which invoices are overdue and still have a balance pending?",
        "Rows with Status == 'Overdue' from the Zoho Books export, ranked by "
        "parsed Balance (the column is a currency string like '₹12,500.00', "
        "not a number) descending.",
        expected_rows=[{"invoice_number": r["Invoice Number"],
                          "customer_name": r["Customer Name"],
                          "balance": r["Balance"]}
                         for _, r in overdue.iterrows()],
        order_matters=True)

    # ---- health ----
    kobo = _g("arogya-jyoti", "anc_tracking_kobo_export.csv")
    camp = _g("arogya-jyoti", "camp_register.xlsx", "Camp Register")

    ask("ask-08", "arogya-jyoti", ["anc_tracking_kobo_export.csv"],
        "How many women completed at least their first ANC visit (ANC1)?",
        "Count of rows where visit/anc1 == 'yes' in the KoboToolbox export.",
        expected_rows=[{"count": int((kobo["visit/anc1"] == "yes").sum())}],
        order_matters=False)

    top_camps = camp.sort_values("women", ascending=False).head(3)
    ask("ask-09", "arogya-jyoti", ["camp_register.xlsx#Camp Register"],
        "Which health camps had the most women attendees?",
        "camp_register.xlsx has one title-row + attendee-table block per "
        "camp (not one tidy table); a correct answer aggregates women count "
        "per camp from those blocks and ranks descending.",
        expected_rows=[{"camp_no": int(r["camp_no"]), "village": r["village"],
                          "camp_date": r["camp_date"], "women": int(r["women"])}
                         for _, r in top_camps.iterrows()],
        order_matters=True)

    # ---- disability ----
    ben = _g("saksham-kalyan", "beneficiary_master.xlsx", "Beneficiary Master")
    devtx = _g("saksham-kalyan", "device_inventory.csv")

    ask("ask-10", "saksham-kalyan", ["beneficiary_master.xlsx#Beneficiary Master"],
        "How many beneficiaries have received an assistive device?",
        "Count of rows where Assistive Device Issued is non-empty. Dates in "
        "this sheet are mixed dd/mm/yyyy and yyyy-mm-dd; a correct tool "
        "should not mis-parse or drop rows because of that.",
        expected_rows=[{"count": int((ben["device_issued"] != "").sum())}],
        order_matters=False)

    devtx2 = devtx.copy()
    devtx2["_d"] = pd.to_datetime(devtx2["date"], dayfirst=True)
    last_balance = (devtx2.sort_values("_d")
                     .groupby("device_type").tail(1)
                     .sort_values("balance_after"))
    ask("ask-11", "saksham-kalyan", ["device_inventory.csv"],
        "Which device type has the lowest stock remaining?",
        "The running 'balance_after' column is a stock ledger, not a stock "
        "table; a correct answer takes the most recent (by date) balance "
        "per device_type and ranks ascending.",
        expected_rows=[{"device_type": r["device_type"], "balance_after": int(r["balance_after"])}
                         for _, r in last_balance.head(3).iterrows()],
        order_matters=True)

    # ---- fundraising ----
    crm = _g("asha-kiran", "donor_crm_zoho_export.csv")
    pay = _g("asha-kiran", "razorpay_payments.csv")
    rec80g = _g("asha-kiran", "80g_receipts_2025.xlsx", "80G Receipts")

    ask("ask-12", "asha-kiran", ["donor_crm_zoho_export.csv"],
        "How many leads in the donor CRM export are marked 'Converted'?",
        "Count of rows where Lead Status == 'Converted' in the Zoho CRM export.",
        expected_rows=[{"count": int((crm["Lead Status"] == "Converted").sum())}],
        order_matters=False)

    cap = pay[pay["status"] == "captured"].copy()
    camp_sum = cap.groupby("notes.campaign")["amount"].sum().sort_values(ascending=False)
    top_campaign = camp_sum.index[0]
    ask("ask-13", "asha-kiran", ["razorpay_payments.csv"],
        f"What was the total captured amount raised for the {top_campaign} campaign?",
        "Sum of 'amount' (paise, divide by 100 for INR) where status == "
        "'captured' and notes.campaign matches, in razorpay_payments.csv.",
        expected_rows=[{"campaign": top_campaign,
                          "total_inr": round(float(camp_sum.iloc[0]) / 100, 2)}],
        order_matters=False)

    repeat = (cap.groupby("email").size().sort_values(ascending=False))
    repeat = repeat[repeat > 1].head(3)
    ask("ask-14", "asha-kiran", ["razorpay_payments.csv"],
        "Which donors made more than one captured payment (i.e. gave twice or more)?",
        "Group captured rows in razorpay_payments.csv by email/contact; "
        "donors with count > 1, ranked by count descending.",
        expected_rows=[{"email": e, "payment_count": int(c)} for e, c in repeat.items()],
        order_matters=True)

    ask("ask-15", "asha-kiran", ["80g_receipts_2025.xlsx#80G Receipts"],
        "How many 80G receipts were issued in total for 2025?",
        "Row count of 80g_receipts_2025.xlsx (one receipt per captured "
        "donation sampled for receipting).",
        expected_rows=[{"count": int(len(rec80g))}], order_matters=False)

    # ---- ops / admin ----
    bva = _g("prayas-seva-sangh", "expense_vouchers_2025.xlsx", "Budget vs Actual")
    staffatt = _g("prayas-seva-sangh", "staff_attendance_aug.xlsx", "Attendance Aug 2025")
    vehicle = _g("prayas-seva-sangh", "vehicle_log.csv")

    over_budget = bva[bva["variance"] > 0].sort_values("variance", ascending=False)
    ask("ask-16", "prayas-seva-sangh", ["expense_vouchers_2025.xlsx#Budget vs Actual"],
        "Which expense heads are over their FY25-26 budget based on Apr-Sep actuals?",
        "Rows of the pre-pivoted 'Budget vs Actual' sheet where "
        "Actual > Budget (variance > 0), ranked by variance descending.",
        expected_rows=[{"head": r["head"], "budget": float(r["budget"]),
                          "actual": float(r["actual"]), "variance": float(r["variance"])}
                         for _, r in over_budget.iterrows()],
        order_matters=True)

    most_absent = staffatt.sort_values("A", ascending=False).head(3)
    ask("ask-17", "prayas-seva-sangh", ["staff_attendance_aug.xlsx#Attendance Aug 2025"],
        "Which staff members had the most absent (A) days in August?",
        "From the wide days-1..31 attendance grid, count 'A' codes per row "
        "and rank descending (the sheet already includes a pre-summed 'A' "
        "column that should agree with a fresh count of the day columns).",
        expected_rows=[{"name": r["name"], "designation": r["designation"], "A": int(r["A"])}
                         for _, r in most_absent.iterrows()],
        order_matters=True)

    veh2 = vehicle.copy()
    veh2["_d"] = pd.to_datetime(veh2["date"], dayfirst=True)
    june_km = veh2[(veh2["_d"].dt.month == 6) & (veh2["_d"].dt.year == 2025)]["km_travelled"].sum()
    ask("ask-18", "prayas-seva-sangh", ["vehicle_log.csv"],
        "What was the total distance travelled by the fleet in June 2025?",
        "Sum of km_travelled in vehicle_log.csv for trips dated in June 2025 "
        "(dates are dd/mm/yyyy strings).",
        expected_rows=[{"month": "June 2025", "total_km": int(june_km)}],
        order_matters=False)

    # ---- field team ----
    counts = _g("gram-sudhar", "whatsapp_field_group_export.txt", "daily_counts_extracted")
    visits = _g("gram-sudhar", "village_visit_log.csv")

    top_days = counts.sort_values("children_count", ascending=False).head(3)
    ask("ask-19", "gram-sudhar", ["whatsapp_field_group_export.txt"],
        "From the WhatsApp field group chat, which days had the highest reported "
        "number of children attending, and how many?",
        "The chat is a raw 'dd/mm/yy, HH:MM - Name: message' export with "
        "ordinary chatter, <Media omitted> lines and shared phone numbers "
        "mixed in; a correct answer extracts only the daily count messages "
        "(e.g. 'Aaj <village> mein N bacche aaye') and ranks by N descending.",
        expected_rows=[{"date": r["date_ddmmyy"], "village": r["village"],
                          "children_count": int(r["children_count"])}
                         for _, r in top_days.iterrows()],
        order_matters=True)

    top_visited = visits["village"].value_counts().head(3)
    ask("ask-20", "gram-sudhar", ["village_visit_log.csv"],
        "Which villages received the most field visits?",
        "Count of rows per village in village_visit_log.csv, ranked descending.",
        expected_rows=[{"village": v, "visit_count": int(c)} for v, c in top_visited.items()],
        order_matters=True)

    # ---- rural development MIS ----
    june = _g("krishi-jal-vikas", "mis_apr_sep_2025.xlsx", "June")
    cons = _g("krishi-jal-vikas", "mis_apr_sep_2025.xlsx", "Consolidated")

    ask("ask-21", "krishi-jal-vikas", ["mis_apr_sep_2025.xlsx#Jun"],
        "How many households were reached in June 2025 across all villages?",
        "Sum of the households column in the June sheet. The column header "
        "text drifts across the workbook's month sheets ('No. of HH' in "
        "April/May, 'Households reached' in June/July/September, 'HH Covered' "
        "in August) even though it is the same indicator; a correct tool "
        "must reconcile that before summing, not treat them as different metrics.",
        expected_rows=[{"month": "June", "total_hh_reached": int(june["households_reached"].sum())}],
        order_matters=False)

    top_hh = cons.sort_values("total_hh_reached", ascending=False).head(3)
    ask("ask-22", "krishi-jal-vikas", ["mis_apr_sep_2025.xlsx#Consolidated"],
        "Which villages had the highest total households reached across Apr-Sep 2025?",
        "Top rows of the Consolidated sheet (values pasted from formulas, "
        "not live formulas) ranked by Total HH Reached descending.",
        expected_rows=[{"village": r["village"], "total_hh_reached": int(r["total_hh_reached"])}
                         for _, r in top_hh.iterrows()],
        order_matters=True)

    # ---- build cases (10) ----
    build("build-01", "sunrise-shiksha",
          ["students_2025_26.xlsx", "funder_report_Q1.xlsx"],
          "Build a funder-facing dashboard for Sunrise Shiksha summarising "
          "Q1 enrolment, attendance trend and learning assessment progress.",
          "Must reconcile the three ID schemes across Enrolment/Attendance/"
          "Assessment sheets (join by name), show a monthly attendance "
          "trend from the wide Attendance sheet, report the dropout count "
          "from ask-02, and reflect the Indicator Tracker's target vs "
          "achieved without double-counting the TOTAL row.")
    build("build-02", "swayam-mahila",
          ["shg_master.csv", "loan_ledger.xlsx", "zoho_books_invoices_export.csv"],
          "Build a loan portfolio dashboard by village showing amount "
          "disbursed, outstanding balance, and overdue invoice exposure per SHG.",
          "Must join shg_master.csv to loan_ledger.xlsx by SHG ID, compute "
          "outstanding balance from the wide monthly repayment columns "
          "(not just read the label), and separately surface overdue/"
          "unpaid Zoho invoices with currency strings parsed to numbers.")
    build("build-03", "arogya-jyoti",
          ["anc_tracking_kobo_export.csv", "camp_register.xlsx"],
          "Build a monthly maternal-health dashboard: ANC visit funnel "
          "(ANC1 through ANC4 drop-off) from the Kobo export, plus camp "
          "attendance by village and gender from the camp register.",
          "Must compute the ANC1->ANC4 funnel drop-off correctly (each "
          "stage is only valid if the prior stage is 'yes'), and must "
          "aggregate the camp register's repeated title+table blocks per "
          "camp into one row per camp rather than treating title rows as data.")
    build("build-04", "saksham-kalyan",
          ["beneficiary_master.xlsx", "device_inventory.csv"],
          "Build a device inventory and beneficiary support dashboard: "
          "current stock per device type, beneficiaries served, and "
          "device coverage by disability type.",
          "Must derive current stock from the IN/OUT device_inventory.csv "
          "ledger (last balance per device, not a sum of all rows), and "
          "must not break on the mixed dd/mm/yyyy vs yyyy-mm-dd dates in "
          "beneficiary_master.xlsx.")
    build("build-05", "asha-kiran",
          ["donor_crm_zoho_export.csv", "razorpay_payments.csv"],
          "Build a donor funnel dashboard: CRM lead status breakdown "
          "(New -> Contacted -> Qualified -> Converted) alongside actual "
          "revenue captured per campaign from Razorpay.",
          "Must convert Razorpay's paise integers to INR, exclude "
          "failed/refunded payments from revenue, and should be able to "
          "flag repeat donors (see ask-14) as a funnel/retention signal.")
    build("build-06", "asha-kiran",
          ["razorpay_payments.csv", "80g_receipts_2025.xlsx",
           "donor_crm_zoho_export.csv"],
          "Build a fundraising board report: total raised, number of "
          "unique donors, 80G receipts issued, and top campaigns for 2025.",
          "Must reconcile that not every captured payment has a matching "
          "80G receipt (receipts are a sampled subset in this corpus) "
          "without treating the mismatch as an error, and must dedupe "
          "donors by email/contact rather than counting every payment as "
          "a new donor.")
    build("build-07", "prayas-seva-sangh",
          ["expense_vouchers_2025.xlsx"],
          "Build a Budget vs Actual dashboard by expense head for Apr-Sep "
          "2025, highlighting heads that are over budget.",
          "Must use both sheets: aggregate Vouchers by Head as a "
          "cross-check against the already-pivoted 'Budget vs Actual' "
          "sheet, and flag the heads identified in ask-16.")
    build("build-08", "prayas-seva-sangh",
          ["staff_attendance_aug.xlsx", "vehicle_log.csv"],
          "Build an ops summary dashboard for August: staff attendance "
          "(P/A/L/H) by designation, and vehicle usage (km, fuel cost) by vehicle.",
          "Must derive P/A/L/H counts itself from the 31 day-columns and "
          "cross-check against the sheet's own pre-summed P/A/L/H columns; "
          "must aggregate vehicle_log.csv by vehicle and month.")
    build("build-09", "gram-sudhar",
          ["whatsapp_field_group_export.txt", "village_visit_log.csv"],
          "Build a field activity dashboard: daily children-reached trend "
          "parsed from the WhatsApp export, and visit frequency by village "
          "from the visit log, for the field coordinator.",
          "Must parse the raw WhatsApp .txt export (dd/mm/yy, HH:MM - "
          "Name: message), correctly separate the one daily-count message "
          "per day from ordinary chatter/media-omitted/phone-number lines, "
          "and must not fabricate counts for days with no count message.")
    build("build-10", "krishi-jal-vikas",
          ["mis_apr_sep_2025.xlsx"],
          "Build a funder MIS dashboard trending households reached, SHGs "
          "formed, water structures and farmers trained from April to "
          "September 2025, by village.",
          "Must reconcile the drifting column headers across the six "
          "month sheets (see ask-21) into one consistent indicator before "
          "trending, and should treat the 'Consolidated' and 'Dashboard' "
          "sheets as pre-computed checks rather than raw source data.")

    # ---- page cases (8) ----
    page("page-01", "sunrise-shiksha", ["students_2025_26.xlsx"],
         "Build a student enrolment signup form for new admissions for "
         "2025-26, with the same fields as the Enrolment sheet.",
         "Must include Aadhaar, phone and parent-detail fields with basic "
         "format validation (Aadhaar 12 digits, phone 10 digits) and must "
         "not silently overwrite an existing Student ID format like "
         "'STU-25-0001'.")
    page("page-02", "swayam-mahila", ["shg_master.csv", "loan_ledger.xlsx"],
         "Build an internal page for field staff to record a new SHG loan "
         "repayment against the monthly ledger.",
         "Must look up the SHG by SHG ID or name, show current outstanding "
         "balance computed from the wide monthly columns, and record the "
         "new payment against the correct month without breaking the "
         "wide layout.")
    page("page-03", "arogya-jyoti", ["anc_tracking_kobo_export.csv"],
         "Build a simple PWA for ANC visit reminders that a health worker "
         "can use offline in the field to see which women are due for "
         "their next ANC visit.",
         "Must derive 'next visit due' from the ANC1-ANC4 boolean "
         "progression (not just show a static list), work offline, and "
         "handle woman/phone as sensitive data (not shown in a public view).")
    page("page-04", "saksham-kalyan", ["beneficiary_master.xlsx"],
         "Build a beneficiary lookup app for staff to search by UDID "
         "Number and see disability type, device issued, and issue date.",
         "Must parse the UDID-shaped identifier as a string (not a "
         "number, which would drop leading structure) and must handle "
         "the mixed dd/mm/yyyy / yyyy-mm-dd date formats when displaying "
         "device_issue_date consistently.")
    page("page-05", "asha-kiran", ["razorpay_payments.csv", "80g_receipts_2025.xlsx"],
         "Build a donor-facing thank-you/receipt-request page after a "
         "Razorpay payment, letting the donor request an 80G receipt.",
         "Must convert the paise amount to INR for display, and must not "
         "expose other donors' payment or contact data.")
    page("page-06", "prayas-seva-sangh", ["expense_vouchers_2025.xlsx"],
         "Build an expense voucher submission form for field staff, "
         "matching the Vouchers sheet's fields (Head, Sub-head, Amount, "
         "Paid To, Mode, Project Code, Approved By).",
         "Must constrain Head/Sub-head to the existing head-to-subhead "
         "mapping seen in the data (e.g. 'Programme' only pairs with its "
         "own sub-heads) rather than a free-text field.")
    page("page-07", "gram-sudhar", ["village_visit_log.csv"],
         "Build an offline-capable PWA for field staff to log a village "
         "visit (village, purpose, households met, issues raised) from "
         "their phone.",
         "Must work with intermittent connectivity (queue-and-sync, since "
         "this team also reports over WhatsApp when signal is weak), and "
         "the issues-raised field should match the existing category set "
         "(Water shortage, Road damage, Ration delay, School infra, None).")
    page("page-08", "krishi-jal-vikas", ["mis_apr_sep_2025.xlsx"],
         "Build a public-facing MIS summary page for donors/board showing "
         "programme reach by village, safe to share outside the organisation.",
         "Must use only aggregated/consolidated figures (no village-level "
         "PII exists in this file, but the page should still avoid "
         "exposing internal remarks like 'Target revised' or 'Data "
         "pending' verbatim to an external audience without context).")

    return cases


# ===========================================================================
# README
# ===========================================================================

DATA_SHAPES = [
    ("Multi-row / merged-cell header register",
     "students_2025_26.xlsx#Enrolment (sunrise-shiksha)",
     "A group header row ('Student Details', 'Parent/Guardian', "
     "'Identification & Contact') merged over 2 rows sits above the real "
     "column headers, so a naive header-row=0 read misparses column names. "
     "Common in enrolment/admission registers kept in Excel by field staff."),
    ("Wide month-by-month layout",
     "students_2025_26.xlsx#Attendance, loan_ledger.xlsx, "
     "staff_attendance_aug.xlsx (days 1-31), mis_apr_sep_2025.xlsx",
     "One row per entity (student/SHG/staff/village), one column per "
     "month or day, values are counts/amounts/attendance codes. Requires "
     "an unpivot before most aggregate questions can be answered; trend "
     "questions ('attendance in June') read a single column."),
    ("Cross-sheet ID drift",
     "students_2025_26.xlsx (Student ID / Attendance ID / Roll No all "
     "differ for the same student)",
     "The same real-world entity is keyed differently in each sheet of "
     "one workbook because each sheet was built by a different person/"
     "process. Joining correctly requires falling back to name matching."),
    ("Two stacked tables in one sheet",
     "students_2025_26.xlsx#Assessment (Baseline block, blank row, "
     "Endline block, each with its own header row)",
     "Baseline and Endline assessments were pasted into the same tab "
     "months apart with different columns and no shared key column; the "
     "Endline block also has fewer rows than Baseline (dropouts), which "
     "is how 'who dropped out' questions actually get answered in the wild."),
    ("Funder indicator tracker with totals row",
     "funder_report_Q1.xlsx#Indicator Tracker",
     "Target / Achieved / % Achievement / Remarks columns, a handful of "
     "indicator rows, and a bottom 'TOTAL' row that is not itself an "
     "indicator and must be excluded from ranking/filtering."),
    ("Zoho Books invoice export",
     "zoho_books_invoices_export.csv (swayam-mahila)",
     "Zoho's own export column order and naming (Invoice Number, Customer "
     "ID, Customer Name, Invoice Date, Due Date, Status, Total, Balance, "
     "Currency Code, ...), with Total/Balance as formatted currency "
     "strings ('₹12,500.00') rather than numbers."),
    ("Zoho CRM lead/contact export",
     "donor_crm_zoho_export.csv (asha-kiran)",
     "Lead Owner, Lead Source, Lead Status (funnel stages incl. 'Junk "
     "Lead'), Annual Revenue (sparse), Created/Modified Time, free-text "
     "Tag and Description columns, matching Zoho CRM's export shape."),
    ("Razorpay payments export",
     "razorpay_payments.csv (asha-kiran)",
     "amount in paise (integer), created_at as a Unix epoch, status "
     "(captured/failed/refunded), notes.campaign as a dotted custom-field "
     "column name, contact as a +91-prefixed string -- all typical of a "
     "raw Razorpay dashboard export rather than a cleaned finance sheet."),
    ("KoboToolbox / ODK survey export",
     "anc_tracking_kobo_export.csv (arogya-jyoti)",
     "_id, _uuid, start/end/today timestamps, group/field q-codes "
     "(enumerator/name, woman/phone, visit/anc1..anc4), _submission_time "
     "and _validation_status -- the standard flattened-XLSForm export shape."),
    ("Title-row + repeated block register",
     "camp_register.xlsx#Camp Register (arogya-jyoti)",
     "Rather than one tidy table, each health camp gets its own merged "
     "title row, its own header row, its own attendee rows and its own "
     "'Total' line, with a blank spacer row before the next camp -- how a "
     "field worker actually pastes one register per event into one tab."),
    ("Mixed date formats in one column",
     "beneficiary_master.xlsx (saksham-kalyan)",
     "DOB, Device Issue Date and Registration Date each switch between "
     "dd/mm/yyyy and yyyy-mm-dd within the same column, depending on who "
     "or which device entered that row."),
    ("Government-ID-shaped identifiers",
     "beneficiary_master.xlsx (UDID, 18 chars: state+district+CMO+"
     "disability-type+birth-year+running-number+checksum), Aadhaar-shaped "
     "12-digit numbers, PAN-shaped 10-character strings across several orgs",
     "Synthetic values in the correct *shape* for India-specific IDs "
     "(UDID, Aadhaar, PAN, IFSC) so a tool must recognise/redact them "
     "without a hardcoded real-ID lookup."),
    ("Stock ledger (running balance, not a stock table)",
     "device_inventory.csv (saksham-kalyan)",
     "IN/OUT transaction rows with a running balance_after column; "
     "'current stock' requires taking the latest balance per item, not "
     "summing the ledger."),
    ("Budget-vs-Actual pivot pasted as values",
     "expense_vouchers_2025.xlsx#Budget vs Actual, "
     "mis_apr_sep_2025.xlsx#Consolidated / #Dashboard",
     "A second sheet in the same workbook holds what was clearly once a "
     "PivotTable or a formula roll-up, now pasted as plain values with no "
     "live link back to the transaction-level sheet -- so it can drift "
     "from the raw data and both must be reconciled, not just read."),
    ("WhatsApp chat export",
     "whatsapp_field_group_export.txt (gram-sudhar)",
     "Raw 'dd/mm/yy, HH:MM - Name: message' lines, Hinglish, "
     "<Media omitted> placeholders, a phone number shared inline, and "
     "exactly one per-day 'count' message (e.g. 'Aaj Wadgaon mein 23 "
     "bacche aaye') buried among ordinary chatter -- the only record of "
     "some field numbers for NGOs that report over WhatsApp."),
    ("Header drift across a multi-sheet workbook",
     "mis_apr_sep_2025.xlsx (krishi-jal-vikas): 'No. of HH' (Apr, May) "
     "vs 'Households reached' (Jun, Jul, Sep) vs 'HH Covered' (Aug)",
     "The same indicator is labelled differently sheet to sheet because a "
     "different person filled in each month's tab; summing/trending "
     "requires recognising these as one indicator, not six."),
]

README_TEMPLATE = """# NGO test corpus (synthetic)

Generated by `../build_ngo_corpus.py` (seed {seed}). Do not hand-edit these
files -- change the generator and re-run it. Nothing here is committed (see
the generator's docstring); this is disposable evidence for the T0 ask/build/
page lanes.

All organisations, people, phone numbers, Aadhaar/PAN/UDID/bank-account
numbers below are **synthetic**. Formats are realistic (12-digit Aadhaar-
shaped numbers, 18-character UDID-shaped strings, dd/mm/yyyy dates, ₹
currency strings, etc.) so that a tool has to actually handle the shape, but
no value corresponds to a real person or a real government record.

## Where this comes from

Step 1 of the corpus build read the non-PII survey columns of
`/tmp/io.xlsx` (Organisation name, mission, size, data types collected, how
data is stored, the "one important question", the one workflow problem, and
"briefly describe your dataset") for the ~31 real small Indian NGOs who
responded, then looked at public material for a sample of them plus a few
other typical Indian NGOs to see what the underlying *files* actually look
like (not just what the survey said). That reconnaissance, plus the survey
itself, is where the 8 fictional orgs and the 15 data shapes below come from:

- The survey repeatedly named Google Sheets/Excel, KoboToolbox/Google Forms,
  "CRM or internal database", paper forms and "Others -> Zoho suite (Zoho
  Sheet)" as storage; several respondents explicitly described wide
  spreadsheets with "100 columns, one per form field", donor/fundraising
  data kept per-transaction, WhatsApp groups as a source of lost field
  data, and cross-file reconciliation ("applicant lists across multiple
  Excel files") as their actual pain point -- all mirrored below.
- Sector spread mirrors the survey: sports-for-development / education
  (Y-Ultimate, Lila Poonawalla Foundation, CEE), disability (PRA
  Foundation, PWID four-pillar scoring), health (Lions Club Poona Eye
  Foundation, Muktaa Charitable Foundation, PNTRS Samavedana), livelihoods/
  rural development (Krushi Vikas va Gramin Prashikshan Sanstha, Ekibeki,
  Swayam Samjik Vikas Sanstha, BAIF), and general field/ops issues (Baithak
  Foundation on resource/utilisation tracking, Foundation Without on Zoho
  Sheet + real-time visualisation).
- Public-facing format references used for realism (fetched via web
  search, August 2026): Zoho Books invoice export columns
  (https://www.zoho.com/us/books/help/invoice/,
  https://www.zoho.com/us/books/kb/invoices/exp-inv.html), Zoho CRM export
  behaviour (https://help.zoho.com/portal/en/kb/crm/faqs/data-administration/export/articles/faqs-exporting-data-from-zoho-crm),
  Razorpay payment/report fields (https://razorpay.com/docs/payments/dashboard/reports/,
  https://coefficient.io/use-cases/import-razorpay-payments-data-to-excel),
  KoboToolbox export system fields _id/_uuid/_submission_time/
  _validation_status (https://support.kobotoolbox.org/export_download.html,
  https://support.kobotoolbox.org/advanced_export.html), WhatsApp chat
  export line format (https://www.threadrecap.com/en/blog/anatomy-whatsapp-chat-export,
  https://www.threadrecap.com/en/blog/whatsapp-export-formats-explained),
  NRLM/DAY-NRLM SHG register fields -- state/district/block/GP, bank
  name/branch/account, member details
  (https://aikosh.indiaai.gov.in/home/datasets/details/national_rural_livelihood_mission_nrlm_self_help_group_shg_data.html,
  https://anantamias.com/nrlm-shg-list/), 80G donation receipt mandatory
  fields -- NGO PAN, 80G registration/URN, donor PAN, amount, mode
  (https://filingscorner.com/blogs/how-to-issue-80g-donation-receipts-format-details-2025-guide,
  https://www.genesis-foundation.net/blog/80g-donation-receipt-format-what-ngos-and-donors-need-to-know/),
  and the UDID (Unique Disability ID) 18-character structure -- state,
  district, CMO code, disability type, birth year, running number,
  checksum (https://depwd.gov.in/en/unique-disability-id-udid/,
  https://www.bankbazaar.com/govt-utility/udid-card.html).

## Data shapes in this corpus

{shapes}

## Organisations and files

{orgs}

## Using this corpus

`cases.json` in this directory has 40 cases (22 `ask`, 10 `build`, 8 `page`)
against these files. `ask` cases include `expected_rows` (top 3 rows, or a
single-row count/sum) and `order_matters`, computed directly from the
in-memory data at generation time -- not hand-typed -- so they are exact for
this seed. `build` and `page` cases carry an `expect` description of what a
correct dashboard/page must get right instead, since there is no single
correct row-level answer for those lanes.
"""


def render_readme(file_counts: dict) -> str:
    shapes_md = "\n".join(
        f"{i}. **{title}** -- `{loc}`\n   {desc}"
        for i, (title, loc, desc) in enumerate(DATA_SHAPES, start=1)
    )
    orgs_md_parts = []
    for org, files in file_counts.items():
        orgs_md_parts.append(f"### `{org}/`\n")
        for fname, info in files.items():
            orgs_md_parts.append(f"- `{fname}` -- {info}")
        orgs_md_parts.append("")
    orgs_md = "\n".join(orgs_md_parts)
    return README_TEMPLATE.format(seed=SEED, shapes=shapes_md, orgs=orgs_md)


FILE_BLURBS = {
    ("sunrise-shiksha", "students_2025_26.xlsx"):
        "3-sheet workbook: Enrolment (merged 2-row header, Aadhaar/phone/"
        "parent fields), Attendance (wide by month, own ID scheme), "
        "Assessment (two stacked Baseline/Endline tables, own ID scheme, "
        "endline shorter -- dropouts).",
    ("sunrise-shiksha", "funder_report_Q1.xlsx"):
        "Indicator tracker: Target/Achieved/%/Remarks + a TOTAL row.",
    ("swayam-mahila", "shg_master.csv"):
        "SHG register: village, bank name/account/IFSC, member count.",
    ("swayam-mahila", "loan_ledger.xlsx"):
        "Loan ledger, wide by month (repayments Apr-25..Sep-25), outstanding balance.",
    ("swayam-mahila", "zoho_books_invoices_export.csv"):
        "Zoho Books-shaped invoice export, currency strings, mixed status.",
    ("arogya-jyoti", "anc_tracking_kobo_export.csv"):
        "KoboToolbox-shaped ANC tracking export, q-codes, _submission_time.",
    ("arogya-jyoti", "camp_register.xlsx"):
        "One merged title row + attendee table + total per camp, Hinglish notes.",
    ("saksham-kalyan", "beneficiary_master.xlsx"):
        "UDID-shaped IDs, disability type/%, device issued, mixed date formats.",
    ("saksham-kalyan", "device_inventory.csv"):
        "Device stock IN/OUT ledger with running balance.",
    ("asha-kiran", "donor_crm_zoho_export.csv"):
        "Zoho CRM-shaped lead/contact export, funnel status, Tags.",
    ("asha-kiran", "razorpay_payments.csv"):
        "Razorpay-shaped payments export, paise amounts, epoch timestamps.",
    ("asha-kiran", "80g_receipts_2025.xlsx"):
        "80G receipts issued for a sample of captured Razorpay payments.",
    ("prayas-seva-sangh", "expense_vouchers_2025.xlsx"):
        "Vouchers sheet (transaction-level) + Budget vs Actual sheet (pasted pivot).",
    ("prayas-seva-sangh", "staff_attendance_aug.xlsx"):
        "Wide days 1-31, P/A/L/H codes, plus pre-summed P/A/L/H columns.",
    ("prayas-seva-sangh", "vehicle_log.csv"):
        "Trip log: vehicle, driver, odometer start/end, fuel cost.",
    ("gram-sudhar", "whatsapp_field_group_export.txt"):
        "Raw WhatsApp export, Hinglish, one daily child-count message per day.",
    ("gram-sudhar", "village_visit_log.csv"):
        "Field visit log: village, staff, purpose, households met, issues raised.",
    ("krishi-jal-vikas", "mis_apr_sep_2025.xlsx"):
        "6 month sheets (header text drifts per indicator) + Consolidated "
        "(formulas pasted as values) + Dashboard (pasted pivot grid).",
}


def row_count_desc(org: str, fname: str) -> str:
    sheets = GENERATED.get(org, {}).get(fname, {})
    parts = []
    for sheet, df in sheets.items():
        label = f"#{sheet}" if sheet else ""
        parts.append(f"{label} {len(df)} rows" if label else f"{len(df)} rows")
    return ", ".join(parts)


def main():
    if os.path.isdir(OUT):
        shutil.rmtree(OUT)
    ensure_dir(OUT)

    orgs = [gen_education(), gen_livelihoods(), gen_health(), gen_disability(),
            gen_fundraising(), gen_ops(), gen_field(), gen_rural_mis()]

    # WhatsApp txt line count (not in GENERATED as a df row count)
    wa_path = os.path.join(OUT, "gram-sudhar", "whatsapp_field_group_export.txt")
    with open(wa_path, encoding="utf-8") as f:
        wa_lines = sum(1 for _ in f)

    file_counts = {}
    for org in orgs:
        file_counts[org] = {}
        for fname in GENERATED.get(org, {}):
            counts = row_count_desc(org, fname)
            blurb = FILE_BLURBS.get((org, fname), "")
            extra = f" ({wa_lines} lines total incl. header/system lines)" if fname.endswith(".txt") else ""
            file_counts[org][fname] = f"{blurb} [{counts}{extra}]"

    readme = render_readme(file_counts)
    with open(os.path.join(OUT, "README.md"), "w", encoding="utf-8") as f:
        f.write(readme)

    cases = build_cases()
    with open(os.path.join(OUT, "cases.json"), "w", encoding="utf-8") as f:
        json.dump({"seed": SEED, "cases": cases}, f, indent=2, ensure_ascii=False, default=str)

    print(f"Wrote corpus to {OUT}")
    print(f"Orgs: {orgs}")
    print(f"Cases: {len(cases)}")
    lanes = {}
    for c in cases:
        lanes[c["lane"]] = lanes.get(c["lane"], 0) + 1
    print(f"Lane counts: {lanes}")


if __name__ == "__main__":
    main()
