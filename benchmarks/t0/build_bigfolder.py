#!/usr/bin/env python3
"""Deterministic generator for a large, messy synthetic Indian-NGO shared
drive, plus a companion question set with golds computed programmatically
from the generated files.

Produces ~120 files under benchmarks/t0/bigfolder/ (target 15-40MB total):

  - 60 per-village monthly visit logs   visits_<village>_<YYYY-MM>.csv
  - 24 attendance sheets                attendance_<center>_<YYYY-MM>.csv
  - 12 raw WhatsApp group exports       whatsapp_<village>_group.txt
  - 12 monthly narrative reports        report_<YYYY-MM>.txt
  -  8 misc admin/finance files (budget workbook, staff list, donor list,
     3 grant letters, an old duplicate donor export, a training calendar)
  -  4 junk files (readme/notes/todo clutter, no signal)

and benchmarks/t0/bigfolder-questions.json: 12 questions in three kinds
(single/aggregate/needle) whose "gold" values are computed by re-reading the
actual written files after generation (not carried over from in-memory
generation state), so the golds are exactly what a solver would find on
disk.

Everything here is synthetic: no real people, no real organisations. Village
and district names are real Maharashtra place names used only as flavour
(same convention as build_ngo_corpus.py in this directory), never as an
organisation identity.

Seed: 20260826. Re-running this script reproduces byte-identical output --
including the .xlsx budget workbook, whose openpyxl `created`/`modified`
document properties are pinned to a fixed timestamp below (openpyxl
otherwise stamps wall-clock time into every workbook, which would silently
break determinism).

Usage:
    .venv-v2/bin/python benchmarks/t0/build_bigfolder.py
"""
from __future__ import annotations

import json
import os
import random
import re
import shutil
from calendar import monthrange
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

SEED = 20260826
HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "bigfolder")
QUESTIONS_PATH = os.path.join(HERE, "bigfolder-questions.json")
XLSX_FIXED_TS = datetime(2026, 8, 26, 9, 0, 0)

# ---------------------------------------------------------------------------
# Name / place pools
# ---------------------------------------------------------------------------

MALE_FIRST = [
    "Rahul", "Amit", "Sandeep", "Vijay", "Sunil", "Ganesh", "Prakash", "Ravi",
    "Suresh", "Mahesh", "Ramesh", "Sachin", "Vishal", "Rohan", "Nitin",
    "Sanjay", "Ashok", "Dinesh", "Kiran", "Santosh", "Pravin", "Yogesh",
    "Anil", "Manoj", "Rajesh", "Imran", "Salman", "Aslam", "Firoz", "Iqbal",
    "Gurpreet", "Harpreet", "Jaspal", "Om", "Akash", "Karan", "Abhijit",
    "Tushar", "Prashant", "Deepak",
]
FEMALE_FIRST = [
    "Sunita", "Priya", "Anjali", "Kavita", "Rekha", "Meena", "Pooja",
    "Nisha", "Deepa", "Savita", "Vaishali", "Snehal", "Manisha", "Jyoti",
    "Asha", "Rani", "Kalpana", "Shobha", "Vandana", "Archana", "Neha",
    "Shabana", "Rukhsar", "Fatima", "Ayesha", "Zarina", "Simran",
    "Gurmeet", "Kajal", "Sarita", "Lata", "Usha", "Radha", "Swati",
    "Pallavi", "Divya", "Komal", "Rupali", "Mangal", "Yamini",
]
SURNAMES = [
    "Pawar", "Jadhav", "Shinde", "Kale", "More", "Gaikwad", "Deshmukh",
    "Patil", "Sawant", "Kadam", "Bhosale", "Chavan", "Kulkarni", "Joshi",
    "Bagul", "Nikam", "Gawde", "Waghmare", "Salunkhe", "Thorat", "Khan",
    "Shaikh", "Ansari", "Pathan", "Sheikh", "Sayyed", "Kaur", "Singh",
    "Chowdhury", "Yadav", "Mane", "Bhoir", "Dhumal", "Rathod", "Suryawanshi",
]
CHILD_FIRST = [
    "Aarav", "Ishaan", "Vivaan", "Aditya", "Reyansh", "Kabir", "Arjun",
    "Sai", "Krishna", "Rohan", "Ananya", "Diya", "Saanvi", "Aadhya", "Myra",
    "Anika", "Ira", "Riya", "Sara", "Zara", "Pari", "Meera", "Tanvi", "Om",
    "Yash", "Dev", "Aryan", "Kiara", "Navya", "Siya", "Advait", "Rudra",
    "Vihaan", "Aditi", "Sanvi",
]
VILLAGES = [
    "Wadgaon", "Shirur", "Velhe", "Mulshi", "Junnar", "Ambegaon", "Purandar",
    "Bhor", "Daund", "Indapur", "Khed", "Otur",
]
CITIES = [
    "Mumbai", "Pune", "Nashik", "Nagpur", "Delhi", "Bengaluru", "Hyderabad",
    "Chennai", "Kolkata", "Ahmedabad", "Jaipur", "Lucknow", "Indore",
    "Surat", "Bhopal",
]

ISSUE_POOL = (
    ["None"] * 5 + [
        "Water shortage reported", "Ration card delay",
        "School infrastructure needs repair", "Road in poor condition",
        "SHG loan installment pending", "Health camp requested by residents",
        "Electricity outage affecting households",
        "Domestic dispute referred to counsellor",
        "Anganwadi centre needs new utensils",
        "Drainage blocked near main road",
    ] + ["None", "None"]
)
NEEDLE_HANDPUMP_ISSUE = "Handpump broken, villagers without safe drinking water"
NEEDLE_LANDSLIDE_ISSUE = "Landslide blocked the approach road after heavy rain"

MEETING_TEMPLATES = [
    "Kal gram sabha meeting hai {v} mein, sab time pe aana",
    "Meeting reschedule ho gaya, ab shukrawar ko hoga",
    "Aaj ki meeting mein {n} log aaye the",
    "SHG meeting {v} mein achi rahi aaj",
    "Staff meeting 10 baje start hogi kal",
    "Sarpanch ke saath meeting fix ho gayi hai",
    "Meeting minutes bhej raha hu group mein",
]
PAYMENT_AMOUNTS = [100, 150, 200, 250, 300, 500, 750, 1000, 1500, 2000]
PAYMENT_TEMPLATES = [
    "{name} ne {amt} rupaye jama kiye aaj",
    "SHG fund mein {amt} rupaye collect hue is hafte",
    "{amt} rupaye ka payment pending hai abhi",
    "{name} ka {amt} rupaye ka installment mil gaya",
]
CHILDREN_COUNT_TEMPLATES = [
    "Aaj {n} bacche aaye session mein",
    "{v} centre par {n} bacche the aaj",
    "Attendance aaj {n} bacche, sab active the",
    "Total {n} bacche present the aaj ke session mein",
]
GENERIC_CHATTER = [
    "kal meeting hai sarpanch office mein", "photo bhejta hu thodi der mein",
    "route confirm ho gaya", "gaadi 9 baje niklegi", "material load ho gaya",
    "sab volunteers ready hain", "field visit start", "lunch break le rahe hain",
    "signal weak hai yaha", "session complete ho gaya",
    "attendance sheet upload kar diya", "rasta kharab tha aaj",
    "parents se baat hui achi rahi", "ok noted", "thik hai", "great work team",
    "kal 10 baje milte hain", "van mein jagah nahi bachi",
    "backup plan chahiye", "<Media omitted>", "good morning team",
    "weather thik nahi lag raha, umbrella le ke aana",
]

# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def dd_mm_yyyy(d: datetime) -> str:
    return d.strftime("%d/%m/%Y")


def phone(rng: random.Random) -> str:
    return f"{rng.choice('6789')}{rng.randint(10**8, 10**9 - 1)}"


def full_name(rng: random.Random) -> tuple[str, str]:
    if rng.random() < 0.5:
        first = rng.choice(MALE_FIRST)
    else:
        first = rng.choice(FEMALE_FIRST)
    return first, rng.choice(SURNAMES)


def rand_day_in_month(rng: random.Random, yyyy_mm: str) -> datetime:
    y, m = (int(x) for x in yyyy_mm.split("-"))
    ndays = monthrange(y, m)[1]
    return datetime(y, m, rng.randint(1, ndays))


def weekdays_in_month(yyyy_mm: str) -> list[datetime]:
    y, m = (int(x) for x in yyyy_mm.split("-"))
    ndays = monthrange(y, m)[1]
    out = []
    for day in range(1, ndays + 1):
        d = datetime(y, m, day)
        if d.weekday() != 6:  # skip Sunday
            out.append(d)
    return out


def new_workbook() -> Workbook:
    """Workbook with document properties pinned so output is byte-identical
    across runs (openpyxl otherwise stamps wall-clock created/modified)."""
    wb = Workbook()
    wb.properties.creator = "ngo-office-shared-drive"
    wb.properties.created = XLSX_FIXED_TS
    wb.properties.modified = XLSX_FIXED_TS
    wb.properties.lastModifiedBy = "ngo-office-shared-drive"
    return wb


def finalize_xlsx(path: str):
    """openpyxl's writer unconditionally re-stamps properties.modified with
    wall-clock time inside save() (writer/excel.py), overriding whatever we
    set beforehand; on top of that, every zip entry's own local-header
    date_time defaults to wall-clock time too. Both make the saved .xlsx
    non-deterministic across runs even when cell content is identical.
    Patch docProps/core.xml and pin every entry's date_time so the file is
    byte-identical across runs."""
    import io
    import zipfile

    with open(path, "rb") as f:
        data = f.read()
    zin = zipfile.ZipFile(io.BytesIO(data))
    fixed = XLSX_FIXED_TS.strftime("%Y-%m-%dT%H:%M:%SZ").encode()
    fixed_dt_tuple = (XLSX_FIXED_TS.year, XLSX_FIXED_TS.month, XLSX_FIXED_TS.day,
                       XLSX_FIXED_TS.hour, XLSX_FIXED_TS.minute, XLSX_FIXED_TS.second)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            content = zin.read(item.filename)
            if item.filename == "docProps/core.xml":
                content = re.sub(
                    rb"<dcterms:modified[^>]*>[^<]*</dcterms:modified>",
                    b'<dcterms:modified xsi:type="dcterms:W3CDTF">' + fixed + b"</dcterms:modified>",
                    content,
                )
            new_item = zipfile.ZipInfo(item.filename, date_time=fixed_dt_tuple)
            new_item.compress_type = item.compress_type
            new_item.external_attr = item.external_attr
            zout.writestr(new_item, content)
    with open(path, "wb") as f:
        f.write(buf.getvalue())


def months_window() -> list[str]:
    months = []
    y, m = 2025, 9
    for _ in range(12):
        months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m == 13:
            m = 1
            y += 1
    return months


ALL_MONTHS = months_window()

# ---------------------------------------------------------------------------
# Deterministic calendar / pools built from seeded sub-streams
# ---------------------------------------------------------------------------


def build_staff_pool() -> list[str]:
    rng = random.Random(f"{SEED}-staffpool")
    names = set()
    out = []
    while len(out) < 25:
        fn, sn = full_name(rng)
        name = f"{fn} {sn}"
        if name not in names:
            names.add(name)
            out.append(name)
    return out


STAFF = build_staff_pool()


def build_calendar() -> tuple[list[str], list[str]]:
    rng = random.Random(f"{SEED}-calendar")
    visit_months = sorted(rng.sample(ALL_MONTHS, 5))
    attend_months = sorted(rng.sample(visit_months, 2))
    return visit_months, attend_months


VISIT_MONTHS, ATTEND_MONTHS = build_calendar()


def build_activity_levels() -> dict:
    rng = random.Random(f"{SEED}-activity")
    return {v: rng.choice(["low", "medium", "high"]) for v in VILLAGES}


VILLAGE_ACTIVITY = build_activity_levels()
ACTIVITY_LINE_RANGE = {"low": (150, 220), "medium": (220, 320), "high": (320, 400)}


def build_needles() -> dict:
    rng = random.Random(f"{SEED}-needles")
    handpump_village = rng.choice(VILLAGES)
    remaining = [v for v in VILLAGES if v != handpump_village]
    landslide_village = rng.choice(remaining)
    return {
        "handpump_village": handpump_village,
        "handpump_month": rng.choice(VISIT_MONTHS),
        "landslide_village": landslide_village,
        "landslide_month": rng.choice(VISIT_MONTHS),
        "wa_village": rng.choice(VILLAGES),
        "wa_amount": 7373,
    }


NEEDLES = build_needles()


def center_name(village: str) -> str:
    return f"{village} Anganwadi"


def center_slug(village: str) -> str:
    return f"{village.lower()}_anganwadi"


CENTERS = [center_name(v) for v in VILLAGES]

# ===========================================================================
# Visits: visits_<village>_<YYYY-MM>.csv
# ===========================================================================


def gen_visits() -> dict:
    ensure_dir(OUT)
    filenames_by_village = {v: [] for v in VILLAGES}
    filenames_by_month = {m: [] for m in VISIT_MONTHS}
    all_filenames = []
    needle_hits = {}

    for village in VILLAGES:
        rng = random.Random(f"{SEED}-visits-{village}")
        for month in VISIT_MONTHS:
            n_rows = rng.randint(50, 400)
            rows = []
            for i in range(1, n_rows + 1):
                d = rand_day_in_month(rng, month)
                rows.append({
                    "visit_id": f"VST-{i:04d}",
                    "date": dd_mm_yyyy(d),
                    "staff_name": rng.choice(STAFF),
                    "households_met": rng.randint(1, 25),
                    "issue": rng.choice(ISSUE_POOL),
                })
            is_handpump = (village == NEEDLES["handpump_village"]
                            and month == NEEDLES["handpump_month"])
            is_landslide = (village == NEEDLES["landslide_village"]
                             and month == NEEDLES["landslide_month"])
            if is_handpump:
                idx = rng.randint(0, n_rows - 1)
                staff = rng.choice(STAFF)
                d = rand_day_in_month(rng, month)
                rows[idx] = {
                    "visit_id": rows[idx]["visit_id"], "date": dd_mm_yyyy(d),
                    "staff_name": staff, "households_met": rng.randint(1, 25),
                    "issue": NEEDLE_HANDPUMP_ISSUE,
                }
                needle_hits["handpump"] = {
                    "village": village, "month": month,
                    "staff": staff, "date": dd_mm_yyyy(d),
                }
            if is_landslide:
                idx = rng.randint(0, n_rows - 1)
                staff = rng.choice(STAFF)
                d = rand_day_in_month(rng, month)
                rows[idx] = {
                    "visit_id": rows[idx]["visit_id"], "date": dd_mm_yyyy(d),
                    "staff_name": staff, "households_met": rng.randint(1, 25),
                    "issue": NEEDLE_LANDSLIDE_ISSUE,
                }
                needle_hits["landslide"] = {
                    "village": village, "month": month,
                    "staff": staff, "date": dd_mm_yyyy(d),
                }
            df = pd.DataFrame(rows)
            fname = f"visits_{village.lower()}_{month}.csv"
            df.to_csv(os.path.join(OUT, fname), index=False)
            filenames_by_village[village].append(fname)
            filenames_by_month[month].append(fname)
            all_filenames.append(fname)

    return {
        "by_village": filenames_by_village,
        "by_month": filenames_by_month,
        "all": all_filenames,
        "needles": needle_hits,
    }


# ===========================================================================
# Attendance: attendance_<center>_<YYYY-MM>.csv
# ===========================================================================


def gen_attendance() -> dict:
    filenames_by_month = {m: [] for m in ATTEND_MONTHS}
    all_filenames = []

    for village in VILLAGES:
        center = center_name(village)
        rng = random.Random(f"{SEED}-attendance-{village}")
        for month in ATTEND_MONTHS:
            days = weekdays_in_month(month)
            target_rows = rng.randint(100, 600)
            roster_size = max(5, min(30, round(target_rows / max(len(days), 1))))
            roster = []
            pool = CHILD_FIRST[:]
            rng.shuffle(pool)
            for i in range(roster_size):
                roster.append(pool[i % len(pool)] + (f" {i // len(pool) + 1}" if i >= len(pool) else ""))
            base_rate = rng.uniform(0.65, 0.92)
            child_age = {c: rng.randint(3, 12) for c in roster}
            rows = []
            for d in days:
                for c in roster:
                    present = "Yes" if rng.random() < base_rate else "No"
                    rows.append({
                        "date": dd_mm_yyyy(d), "child_name": c,
                        "age": child_age[c], "present": present,
                    })
            df = pd.DataFrame(rows)
            fname = f"attendance_{center_slug(village)}_{month}.csv"
            df.to_csv(os.path.join(OUT, fname), index=False)
            filenames_by_month[month].append(fname)
            all_filenames.append(fname)

    return {"by_month": filenames_by_month, "all": all_filenames}


# ===========================================================================
# WhatsApp: whatsapp_<village>_group.txt
# ===========================================================================


def gen_whatsapp() -> dict:
    all_filenames = []
    for village in VILLAGES:
        rng = random.Random(f"{SEED}-whatsapp-{village}")
        level = VILLAGE_ACTIVITY[village]
        lo, hi = ACTIVITY_LINE_RANGE[level]
        n_lines = rng.randint(lo, hi)
        admin = rng.choice(STAFF)
        start_date = datetime(2025, 9, 1)

        plant_wa = (village == NEEDLES["wa_village"])
        plant_idx = rng.randint(5, n_lines - 5) if plant_wa else -1

        lines = [
            f"{start_date.strftime('%d/%m/%y')}, 08:00 - Messages and calls are "
            f"end-to-end encrypted. No one outside of this chat, not even "
            f"WhatsApp, can read or listen to them.",
            f"{start_date.strftime('%d/%m/%y')}, 08:01 - {admin} created group "
            f"\"{village} Field Group\"",
        ]
        cursor = start_date + timedelta(hours=8, minutes=5)
        for i in range(n_lines):
            cursor += timedelta(minutes=rng.randint(5, 240))
            sender = rng.choice(STAFF)
            if i == plant_idx:
                other = rng.choice([s for s in STAFF if s != sender])
                msg = (f"{other} ka {NEEDLES['wa_amount']} rupaye ka payment "
                       f"abhi tak nahi mila, delay ho raha hai")
            else:
                roll = rng.random()
                if roll < 0.12:
                    msg = rng.choice(MEETING_TEMPLATES).format(v=village, n=rng.randint(4, 30))
                elif roll < 0.27:
                    name = rng.choice(STAFF)
                    msg = rng.choice(PAYMENT_TEMPLATES).format(
                        name=name, amt=rng.choice(PAYMENT_AMOUNTS))
                elif roll < 0.47:
                    msg = rng.choice(CHILDREN_COUNT_TEMPLATES).format(
                        v=village, n=rng.randint(8, 45))
                elif roll < 0.53:
                    msg = f"call me on {phone(rng)} if any issue"
                else:
                    msg = rng.choice(GENERIC_CHATTER)
            lines.append(f"{cursor.strftime('%d/%m/%y')}, {cursor.strftime('%H:%M')} - {sender}: {msg}")

        fname = f"whatsapp_{village.lower()}_group.txt"
        with open(os.path.join(OUT, fname), "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        all_filenames.append(fname)
    return {"all": all_filenames}


# ===========================================================================
# Monthly narrative reports: report_<YYYY-MM>.txt
# (read back the CSVs already written on disk so prose totals are
# guaranteed consistent with the data)
# ===========================================================================

FILLER_PARAGRAPHS = [
    "The monsoon continued to affect road access to some of the more remote "
    "hamlets this period, and field staff adjusted their travel plans "
    "accordingly. Vehicle availability remained a recurring constraint, and "
    "two field visits were rescheduled as a result.",
    "The team held its regular internal coordination meeting this month to "
    "review pending cases and plan the following month's schedule. Staff "
    "also used part of the period for documentation and backlog data entry, "
    "which had been delayed during the busier field season.",
    "Community relationships remained positive across the operational area. "
    "Village-level contacts continued to support scheduling of household "
    "visits and flagged issues informally over phone and WhatsApp when "
    "field staff could not be present in person.",
    "Staffing remained largely stable this period, with no major changes to "
    "the field team roster. One volunteer took leave for a family event and "
    "was covered by a colleague from a neighbouring village.",
    "The organisation continued routine coordination with panchayat offices "
    "and local health functionaries. No major grievances were escalated to "
    "the district level during this period.",
    "Administrative work continued in parallel with field activity, "
    "including reconciliation of expense vouchers and preparation for the "
    "next round of donor reporting.",
    "Training and refresher sessions for field staff were discussed for the "
    "coming quarter, though no formal session was conducted this month.",
    "Overall the period was consistent with the preceding months, with no "
    "major incidents reported beyond what is noted above.",
]


def month_label(yyyy_mm: str) -> str:
    y, m = (int(x) for x in yyyy_mm.split("-"))
    return datetime(y, m, 1).strftime("%B %Y")


def gen_reports() -> list[str]:
    all_filenames = []
    for month in ALL_MONTHS:
        rng = random.Random(f"{SEED}-report-{month}")
        paragraphs = [
            f"Monthly field programme report -- {month_label(month)}.\n\n"
            f"This report summarises field activity across the twelve "
            f"villages covered by the programme during {month_label(month)}."
        ]

        if month in VISIT_MONTHS:
            total_visits = 0
            total_hh = 0
            issue_rows = 0
            village_visit_counts = {}
            for village in VILLAGES:
                fpath = os.path.join(OUT, f"visits_{village.lower()}_{month}.csv")
                vdf = pd.read_csv(fpath)
                total_visits += len(vdf)
                total_hh += int(vdf["households_met"].sum())
                # pandas read_csv treats the literal string "None" in the
                # issue column as a missing value by default, so a written
                # "None" round-trips as NaN -- notna() is the correct way
                # to count "has a specific issue" here, not != "None".
                issue_rows += int(vdf["issue"].notna().sum())
                village_visit_counts[village] = len(vdf)
            busiest_village = max(village_visit_counts, key=village_visit_counts.get)
            paragraphs.append(
                f"Field teams recorded a total of {total_visits} household "
                f"visits across all twelve villages this month, reaching "
                f"{total_hh} households in total. {issue_rows} of these "
                f"visits noted a specific issue requiring follow-up, "
                f"ranging from water shortages to school infrastructure "
                f"concerns. {busiest_village} recorded the highest number "
                f"of visits this month, with {village_visit_counts[busiest_village]} "
                f"household visits logged there alone."
            )
        else:
            paragraphs.append(
                "Detailed field visit data for this month was not fully "
                "digitised in time for this report; a consolidated update "
                "will be included in a later monthly report once the "
                "field registers have been entered."
            )

        if month in ATTEND_MONTHS:
            total_rows = 0
            total_present = 0
            for village in VILLAGES:
                fpath = os.path.join(OUT, f"attendance_{center_slug(village)}_{month}.csv")
                adf = pd.read_csv(fpath)
                total_rows += len(adf)
                total_present += int((adf["present"] == "Yes").sum())
            rate = round(100 * total_present / total_rows, 1) if total_rows else 0.0
            paragraphs.append(
                f"Across the twelve Anganwadi centres, {total_rows} child "
                f"attendance records were logged this month, with an "
                f"overall attendance rate of {rate}% of recorded sessions "
                f"marked present."
            )

        # pad with filler paragraphs (shuffled per-month, deterministically)
        pool = FILLER_PARAGRAPHS[:]
        rng.shuffle(pool)
        fi = 0
        def word_count(paras):
            return sum(len(p.split()) for p in paras)
        while word_count(paragraphs) < 400 and fi < len(pool):
            paragraphs.append(pool[fi])
            fi += 1
        # if still short (shouldn't normally happen), reuse pool
        while word_count(paragraphs) < 400:
            paragraphs.append(pool[fi % len(pool)])
            fi += 1
        # trim from the end (never touching the data-consistent paragraphs
        # at indices 0-2) if we overshot 800 words
        while word_count(paragraphs) > 800 and len(paragraphs) > 3:
            paragraphs.pop()

        fname = f"report_{month}.txt"
        with open(os.path.join(OUT, fname), "w", encoding="utf-8") as f:
            f.write("\n\n".join(paragraphs) + "\n")
        all_filenames.append(fname)
    return all_filenames


# ===========================================================================
# Misc admin/finance files (8)
# ===========================================================================

EXPENSE_CATEGORIES = [
    ("Salaries", ["Field staff salary", "Admin staff salary", "Driver salary"]),
    ("Travel", ["Field travel reimbursement", "Bus fare", "Fuel"]),
    ("Program materials", ["Anganwadi kit", "Health camp supplies", "Printed forms"]),
    ("Rent", ["Field office rent", "Storage rent"]),
    ("Utilities", ["Electricity bill", "Water bill", "Internet"]),
    ("Printing & Stationery", ["Report printing", "Office stationery"]),
    ("Training", ["Staff training venue", "Trainer honorarium"]),
    ("Audit fees", ["Annual audit fee", "Statutory compliance fee"]),
    ("Bank charges", ["NEFT charges", "Account maintenance"]),
    ("Equipment", ["Laptop purchase", "Field kit bags", "Bicycle repair"]),
    ("Communication", ["Mobile recharge", "SIM card"]),
    ("Field allowances", ["Daily field allowance", "Travel allowance"]),
    ("Repairs & maintenance", ["Vehicle repair", "Office repair"]),
    ("Miscellaneous", ["Miscellaneous expense", "Contingency"]),
]
FISCAL_YEARS = ["FY2023-24", "FY2024-25", "FY2025-26"]
VENDORS = [
    "Shree Stationery Mart", "Balaji Travels", "Om Sai Suppliers",
    "Ganesh Hardware", "Krishna Printers", "Local vendor - cash",
    "Reliable Motors", "Suraj Enterprises", "New India Traders",
]
PAYMENT_MODES = ["Bank Transfer", "Cheque", "Cash", "UPI"]
BUDGET_ROWS_PER_YEAR = 6000


def gen_budget_xlsx():
    rng = random.Random(f"{SEED}-budget")
    wb = new_workbook()
    wb.remove(wb.active)
    for fy in FISCAL_YEARS:
        y0 = int(fy[2:6])
        ws = wb.create_sheet(fy)
        ws.append(["Date", "Category", "Sub-category", "Description",
                   "Vendor/Payee", "Payment Mode", "Amount (INR)"])
        for c in ws[1]:
            c.font = Font(bold=True)
        for _ in range(BUDGET_ROWS_PER_YEAR):
            d = datetime(y0, 4, 1) + timedelta(days=rng.randint(0, 364))
            cat, subs = rng.choice(EXPENSE_CATEGORIES)
            sub = rng.choice(subs)
            amount = rng.choice([250, 500, 750, 1200, 1500, 2000, 3000, 5000,
                                  8000, 12000, 15000, 20000, 35000])
            ws.append([dd_mm_yyyy(d), cat, sub, f"{sub} - {fy}",
                       rng.choice(VENDORS), rng.choice(PAYMENT_MODES), amount])
    xlsx_path = os.path.join(OUT, "budget.xlsx")
    wb.save(xlsx_path)
    finalize_xlsx(xlsx_path)
    return "budget.xlsx"


def gen_staff_list_csv():
    roles = ["Field Coordinator", "Community Mobiliser", "Anganwadi Supervisor",
             "Program Officer", "Data Entry Operator", "Driver", "Accountant"]
    rng = random.Random(f"{SEED}-stafflist")
    rows = []
    for i, name in enumerate(STAFF, start=1):
        joined = datetime(2021, 1, 1) + timedelta(days=rng.randint(0, 1600))
        rows.append({
            "staff_id": f"STF-{i:03d}", "full_name": name,
            "role": rng.choice(roles), "village_assigned": rng.choice(VILLAGES),
            "phone": phone(rng), "join_date": dd_mm_yyyy(joined),
        })
    df = pd.DataFrame(rows)
    df.to_csv(os.path.join(OUT, "staff_list.csv"), index=False)
    return "staff_list.csv"


DONOR_ROWS = 130000
CAMPAIGNS = ["Annual Appeal 2024", "Annual Appeal 2025", "Winter Relief Drive",
             "Education Sponsorship", "Emergency Relief Fund",
             "Matching Gift Program", "Year-End Giving", "Corporate Partnership"]
DONOR_TYPES = ["Individual", "Individual", "Individual", "Corporate", "Trust"]


def gen_donor_list_csv():
    rng = random.Random(f"{SEED}-donorlist")
    domains = ["gmail.com", "yahoo.com", "outlook.com", "rediffmail.com"]
    rows = []
    for i in range(1, DONOR_ROWS + 1):
        fn, sn = full_name(rng)
        email = f"{fn.lower()}.{sn.lower()}{rng.randint(1, 9999)}@{rng.choice(domains)}"
        last_don = datetime(2020, 1, 1) + timedelta(days=rng.randint(0, 2400))
        rows.append({
            "donor_id": f"DNR-{i:06d}",
            "full_name": f"{fn} {sn}",
            "email": email,
            "phone": phone(rng),
            "city": rng.choice(CITIES),
            "donor_type": rng.choice(DONOR_TYPES),
            "total_donated_inr": rng.choice([500, 1000, 1500, 2000, 2500, 5000,
                                              10000, 15000, 25000, 50000]),
            "last_donation_date": dd_mm_yyyy(last_don),
            "campaign": rng.choice(CAMPAIGNS),
        })
    df = pd.DataFrame(rows)
    df.to_csv(os.path.join(OUT, "donor_list.csv"), index=False)
    return "donor_list.csv"


OLD_DONOR_ROWS = 18000


def gen_old_donor_export_csv():
    rng = random.Random(f"{SEED}-olddonor")
    domains = ["gmail.com", "yahoo.co.in", "hotmail.com"]
    rows = []
    for _ in range(OLD_DONOR_ROWS):
        fn, sn = full_name(rng)
        d = datetime(2016, 1, 1) + timedelta(days=rng.randint(0, 1460))
        rows.append({
            "Donor Name": f"{fn} {sn}",
            "Email Address": f"{fn.lower()}{sn.lower()}{rng.randint(1, 999)}@{rng.choice(domains)}",
            "Amount (Rs)": rng.choice([100, 250, 500, 1000, 2000, 5000]),
            "Date": d.strftime("%Y-%m-%d"),
            "Source System": "Migrated - old DonorBase export, kept for reference",
        })
    df = pd.DataFrame(rows)
    df.to_csv(os.path.join(OUT, "old_donor_export_2019.csv"), index=False)
    return "old_donor_export_2019.csv"


GRANT_DONORS = [
    ("Bright Horizon Foundation", "girls' secondary education support", 450000, 12),
    ("Sahyog Charitable Trust", "handpump and drinking-water repair", 180000, 6),
    ("Unnati CSR Initiative", "Anganwadi nutrition supplement programme", 620000, 12),
]


def inr_words_amount(amount: int) -> str:
    lakh = amount // 100000
    rem = amount % 100000
    thousand = rem // 1000
    parts = []
    if lakh:
        parts.append(f"{lakh} Lakh")
    if thousand:
        parts.append(f"{thousand} Thousand")
    if not parts:
        parts.append(str(amount))
    return " ".join(parts)


def gen_grant_letters():
    rng = random.Random(f"{SEED}-grants")
    filenames = []
    info = []
    for donor, purpose, amount, months in GRANT_DONORS:
        slug = donor.lower().replace(" ", "-").replace("'", "")
        signer = rng.choice(STAFF)
        letter_date = datetime(2025, 8, 1) + timedelta(days=rng.randint(0, 60))
        text = (
            f"{donor}\n"
            f"Date: {dd_mm_yyyy(letter_date)}\n\n"
            f"Subject: Grant confirmation letter\n\n"
            f"Dear Team,\n\n"
            f"We are pleased to confirm a grant of INR {amount:,} "
            f"(Rupees {inr_words_amount(amount)} only) to support the "
            f"{purpose} programme, to be utilised over a period of "
            f"{months} months from the date of this letter.\n\n"
            f"Please submit a utilisation report and supporting vouchers "
            f"at the end of the grant period. Funds will be disbursed in "
            f"two instalments, with the second instalment released upon "
            f"receipt of a satisfactory interim report.\n\n"
            f"We look forward to a continued partnership.\n\n"
            f"Regards,\n{signer}\nProgramme Officer, {donor}\n"
        )
        fname = f"grant_letter_{slug}.txt"
        with open(os.path.join(OUT, fname), "w", encoding="utf-8") as f:
            f.write(text)
        filenames.append(fname)
        info.append({"donor": donor, "filename": fname, "purpose": purpose,
                      "amount_inr": amount})
    return filenames, info


def gen_vehicle_log_removed_placeholder():
    # kept out: replaced by old_donor_export_2019.csv + training_calendar.txt
    # to stay within the 8-misc-file budget described in the task.
    pass


def gen_training_calendar_txt():
    rng = random.Random(f"{SEED}-training")
    topics = [
        "Child protection basics", "First aid refresher",
        "Data privacy and consent", "SHG bookkeeping", "Field safety",
        "Communicating with panchayat officials", "Report writing",
        "Grievance handling",
    ]
    lines = ["Staff Training Calendar\n"]
    d = datetime(2025, 9, 15)
    for _ in range(10):
        d += timedelta(days=rng.randint(20, 40))
        lines.append(f"{dd_mm_yyyy(d)} - {rng.choice(topics)} - venue TBC")
    fname = "training_calendar.txt"
    with open(os.path.join(OUT, fname), "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    return fname


def gen_misc() -> dict:
    budget_fname = gen_budget_xlsx()
    staff_fname = gen_staff_list_csv()
    donor_fname = gen_donor_list_csv()
    old_donor_fname = gen_old_donor_export_csv()
    grant_filenames, grant_info = gen_grant_letters()
    training_fname = gen_training_calendar_txt()
    all_filenames = ([budget_fname, staff_fname, donor_fname, old_donor_fname]
                      + grant_filenames + [training_fname])
    assert len(all_filenames) == 8, len(all_filenames)
    return {"all": all_filenames, "grants": grant_info}


# ===========================================================================
# Junk files (4)
# ===========================================================================


def gen_junk() -> list[str]:
    readme = (
        "READ ME BEFORE USING THIS FOLDER\n\n"
        "Please don't rename or move anything in here, other people's "
        "reports link to these file names. If you add a new village or "
        "centre just copy the format of an existing file.\n\n"
        "Ping the office group if something looks wrong, don't just delete "
        "it.\n"
    )
    notes = (
        "quick notes\n"
        "- follow up with block office re: ration delay complaints\n"
        "- check if March visit sheets got uploaded for all villages\n"
        "- old office landline stopped working, use mobile instead\n"
        "- remember to back up this folder before the audit visit\n"
    )
    notes_old = (
        "old notes (probably outdated, keeping just in case)\n"
        "- call Deshmukh sir about the vehicle insurance renewal\n"
        "- pending: last year's donor thank-you letters\n"
        "- laptop in the field office needs a new charger\n"
    )
    todo = (
        "TODO\n"
        "[ ] scan paper attendance registers older than 2024\n"
        "[ ] merge duplicate donor entries\n"
        "[ ] confirm training calendar with block office\n"
        "[x] send March report to donor\n"
    )
    files = {
        "readme.txt": readme, "notes.txt": notes,
        "notes_old.txt": notes_old, "TODO.txt": todo,
    }
    for fname, content in files.items():
        with open(os.path.join(OUT, fname), "w", encoding="utf-8") as f:
            f.write(content)
    return list(files.keys())


# ===========================================================================
# Questions: golds recomputed by re-reading the actual written files
# ===========================================================================


def compute_questions(visits: dict, attendance: dict, misc: dict) -> list[dict]:
    questions = []

    # ---- single (4): answerable from exactly one file --------------------

    s1_village = VILLAGES[2]
    s1_fname = f"whatsapp_{s1_village.lower()}_group.txt"
    with open(os.path.join(OUT, s1_fname), encoding="utf-8") as f:
        s1_lines = f.readlines()
    s1_gold = sum(1 for ln in s1_lines if "meeting" in ln.lower())
    questions.append({
        "q": f"In whatsapp_{s1_village.lower()}_group.txt, how many messages "
             f"mention a meeting (i.e. contain the word 'meeting')?",
        "kind": "single",
        "gold": s1_gold,
        "gold_note": "Count of lines in the file containing the substring "
                      "'meeting' (case-insensitive), including the header "
                      "system lines.",
        "files_needed": [s1_fname],
    })

    s2_village, s2_month = VILLAGES[0], VISIT_MONTHS[0]
    s2_fname = f"visits_{s2_village.lower()}_{s2_month}.csv"
    s2_df = pd.read_csv(os.path.join(OUT, s2_fname))
    s2_counts = s2_df["staff_name"].value_counts()
    s2_staff = s2_counts.idxmax()
    s2_gold = int(s2_counts.max())
    questions.append({
        "q": f"How many visits did {s2_staff} make in {s2_fname}?",
        "kind": "single",
        "gold": s2_gold,
        "gold_note": f"Count of rows in {s2_fname} where staff_name == "
                      f"'{s2_staff}' ({s2_staff} is that file's most "
                      f"frequent staff_name).",
        "files_needed": [s2_fname],
    })

    s3_village, s3_month = VILLAGES[0], ATTEND_MONTHS[0]
    s3_fname = f"attendance_{center_slug(s3_village)}_{s3_month}.csv"
    s3_df = pd.read_csv(os.path.join(OUT, s3_fname))
    s3_gold = round(100 * (s3_df["present"] == "Yes").sum() / len(s3_df), 1)
    questions.append({
        "q": f"In {s3_fname}, what percentage of attendance records are "
             f"marked present? (round to 1 decimal place)",
        "kind": "single",
        "gold": s3_gold,
        "gold_note": "100 * count(present == 'Yes') / total rows in that "
                      "one file, rounded to 1 decimal place.",
        "files_needed": [s3_fname],
    })

    s4_village, s4_month = VILLAGES[1], VISIT_MONTHS[1]
    s4_fname = f"visits_{s4_village.lower()}_{s4_month}.csv"
    s4_df = pd.read_csv(os.path.join(OUT, s4_fname))
    s4_gold = int(s4_df["households_met"].sum())
    questions.append({
        "q": f"What is the total households_met in {s4_fname}?",
        "kind": "single",
        "gold": s4_gold,
        "gold_note": f"Sum of the households_met column in {s4_fname}.",
        "files_needed": [s4_fname],
    })

    # ---- aggregate (4): need many files ------------------------------

    a1_month = VISIT_MONTHS[2]
    a1_files = [f"visits_{v.lower()}_{a1_month}.csv" for v in VILLAGES]
    a1_gold = 0
    for fn in a1_files:
        a1_gold += int(pd.read_csv(os.path.join(OUT, fn))["households_met"].sum())
    questions.append({
        "q": f"What is the total households_met across all villages in "
             f"{a1_month}?",
        "kind": "aggregate",
        "gold": a1_gold,
        "gold_note": f"Sum of households_met across the 12 "
                      f"visits_<village>_{a1_month}.csv files.",
        "files_needed": a1_files,
    })

    a2_files = visits["all"]
    a2_village_totals = {}
    for village in VILLAGES:
        total = 0
        for month in VISIT_MONTHS:
            fn = f"visits_{village.lower()}_{month}.csv"
            total += len(pd.read_csv(os.path.join(OUT, fn)))
        a2_village_totals[village] = total
    a2_top = max(a2_village_totals, key=a2_village_totals.get)
    questions.append({
        "q": "Across all visit records (all villages, all months), which "
             "village had the most total visits, and how many?",
        "kind": "aggregate",
        "gold": {"village": a2_top, "visits": a2_village_totals[a2_top]},
        "gold_note": "Row counts of visits_<village>_<month>.csv summed "
                      "per village across all 5 months, then argmax.",
        "files_needed": a2_files,
    })

    a3_month = ATTEND_MONTHS[1] if len(ATTEND_MONTHS) > 1 else ATTEND_MONTHS[0]
    a3_files = [f"attendance_{center_slug(v)}_{a3_month}.csv" for v in VILLAGES]
    rates = []
    for fn in a3_files:
        df = pd.read_csv(os.path.join(OUT, fn))
        rates.append(100 * (df["present"] == "Yes").sum() / len(df))
    a3_gold = round(sum(rates) / len(rates), 1)
    questions.append({
        "q": f"What is the average attendance present-rate (%) across all "
             f"centres in {a3_month}? (round to 1 decimal place)",
        "kind": "aggregate",
        "gold": a3_gold,
        "gold_note": "For each of the 12 attendance_<center>_"
                      f"{a3_month}.csv files, compute 100*count(present=="
                      "'Yes')/rows, then take the unweighted mean of those "
                      "12 per-centre rates and round to 1 decimal place.",
        "files_needed": a3_files,
    })

    a4_files = visits["all"]
    a4_staff_totals = {}
    for fn in a4_files:
        df = pd.read_csv(os.path.join(OUT, fn))
        for name, cnt in df["staff_name"].value_counts().items():
            a4_staff_totals[name] = a4_staff_totals.get(name, 0) + int(cnt)
    a4_top = max(a4_staff_totals, key=a4_staff_totals.get)
    questions.append({
        "q": "Across all 60 visit CSVs, which staff member made the most "
             "total visits, and how many?",
        "kind": "aggregate",
        "gold": {"staff": a4_top, "visits": a4_staff_totals[a4_top]},
        "gold_note": "Value counts of staff_name summed across all 60 "
                      "visits_<village>_<month>.csv files, then argmax.",
        "files_needed": a4_files,
    })

    # ---- needle (4): one buried fact, exactly once in the folder ------

    hp_village = NEEDLES["handpump_village"]
    hp_files = [f"visits_{hp_village.lower()}_{m}.csv" for m in VISIT_MONTHS]
    hp_hits = []
    for fn in hp_files:
        df = pd.read_csv(os.path.join(OUT, fn))
        matches = df[df["issue"].str.contains("Handpump", case=False, na=False)]
        for _, row in matches.iterrows():
            hp_hits.append((row["staff_name"], row["date"]))
    assert len(hp_hits) == 1, f"expected exactly one handpump needle hit, got {hp_hits}"
    questions.append({
        "q": f"Who reported a broken handpump in {hp_village}, and on what "
             f"date? (search that village's monthly visit CSVs)",
        "kind": "needle",
        "gold": {"staff": hp_hits[0][0], "date": hp_hits[0][1]},
        "gold_note": f"The single row across the 5 visits_{hp_village.lower()}"
                      f"_<month>.csv files whose issue field contains "
                      f"'Handpump'.",
        "files_needed": hp_files,
    })

    grant_idx = 1
    grant = misc["grants"][grant_idx]
    with open(os.path.join(OUT, grant["filename"]), encoding="utf-8") as f:
        grant_text = f.read()
    m = re.search(r"INR\s*([\d,]+)", grant_text)
    parsed_amount = int(m.group(1).replace(",", ""))
    assert parsed_amount == grant["amount_inr"], (parsed_amount, grant)
    questions.append({
        "q": f"What amount did the {grant['donor']} grant letter commit, "
             f"and for what purpose?",
        "kind": "needle",
        "gold": {"amount_inr": grant["amount_inr"], "purpose": grant["purpose"]},
        "gold_note": f"Amount and purpose stated in {grant['filename']}, "
                      f"extracted with regex r'INR\\s*([\\d,]+)' against the "
                      f"written file text.",
        "files_needed": [grant["filename"]],
    })

    wa_files = [f"whatsapp_{v.lower()}_group.txt" for v in VILLAGES]
    wa_hits = []
    needle_phrase = "abhi tak nahi mila"
    for fn in wa_files:
        with open(os.path.join(OUT, fn), encoding="utf-8") as f:
            for line in f:
                if needle_phrase in line and str(NEEDLES["wa_amount"]) in line:
                    sender = line.split(" - ", 1)[1].split(":", 1)[0]
                    wa_hits.append((fn, sender))
    assert len(wa_hits) == 1, f"expected exactly one whatsapp payment-delay needle hit, got {wa_hits}"
    hit_fname, hit_sender = wa_hits[0]
    hit_village_slug = hit_fname.replace("whatsapp_", "").replace("_group.txt", "")
    hit_village = next(v for v in VILLAGES if v.lower() == hit_village_slug)
    questions.append({
        "q": f"In which village's WhatsApp group export was a payment of "
             f"Rs {NEEDLES['wa_amount']} reported as delayed, and who sent "
             f"that message?",
        "kind": "needle",
        "gold": {"village": hit_village, "sender": hit_sender},
        "gold_note": f"The single line across the 12 whatsapp_<village>_"
                      f"group.txt files containing both "
                      f"'{NEEDLES['wa_amount']}' and the phrase "
                      f"'{needle_phrase}'; sender parsed from the "
                      f"'dd/mm/yy, hh:mm - Name: msg' line format.",
        "files_needed": wa_files,
    })

    ls_files = visits["all"]
    ls_hits = []
    for fn in ls_files:
        df = pd.read_csv(os.path.join(OUT, fn))
        matches = df[df["issue"].str.contains("Landslide", case=False, na=False)]
        for _, row in matches.iterrows():
            village_from_fname = fn.split("_")[1]
            ls_hits.append((row["staff_name"], village_from_fname, row["date"]))
    assert len(ls_hits) == 1, f"expected exactly one landslide needle hit, got {ls_hits}"
    ls_staff, ls_village_slug, ls_date = ls_hits[0]
    ls_village = next(v for v in VILLAGES if v.lower() == ls_village_slug)
    questions.append({
        "q": "Across all village visit CSVs, which staff member reported a "
             "landslide blocking a road, in which village, and on what "
             "date?",
        "kind": "needle",
        "gold": {"staff": ls_staff, "village": ls_village, "date": ls_date},
        "gold_note": "The single row across all 60 visits_<village>_"
                      "<month>.csv files whose issue field contains "
                      "'Landslide'.",
        "files_needed": ls_files,
    })

    assert len(questions) == 12
    kinds = [q["kind"] for q in questions]
    assert kinds.count("single") == 4
    assert kinds.count("aggregate") == 4
    assert kinds.count("needle") == 4
    return questions


# ===========================================================================
# Main
# ===========================================================================


def main():
    if os.path.isdir(OUT):
        shutil.rmtree(OUT)
    ensure_dir(OUT)

    visits = gen_visits()
    attendance = gen_attendance()
    whatsapp = gen_whatsapp()
    reports = gen_reports()
    misc = gen_misc()
    junk = gen_junk()

    questions = compute_questions(visits, attendance, misc)
    with open(QUESTIONS_PATH, "w", encoding="utf-8") as f:
        json.dump({"seed": SEED, "questions": questions}, f, indent=2,
                   ensure_ascii=False, default=str)

    all_files = (visits["all"] + attendance["all"] + whatsapp["all"]
                 + reports + misc["all"] + junk)
    total_bytes = 0
    for fn in all_files:
        total_bytes += os.path.getsize(os.path.join(OUT, fn))

    print(f"Wrote {len(all_files)} files to {OUT}")
    print(f"  visits: {len(visits['all'])}, attendance: {len(attendance['all'])}, "
          f"whatsapp: {len(whatsapp['all'])}, reports: {len(reports)}, "
          f"misc: {len(misc['all'])}, junk: {len(junk)}")
    print(f"Total size: {total_bytes / (1024 * 1024):.2f} MB")
    print(f"VISIT_MONTHS={VISIT_MONTHS} ATTEND_MONTHS={ATTEND_MONTHS}")
    print(f"Wrote {len(questions)} questions to {QUESTIONS_PATH}")


if __name__ == "__main__":
    main()
