#!/usr/bin/env python3
"""
Maternal Health Coverage Dashboard - Python Server
Serves the website and provides JSON API for Excel workbook data.
"""

import http.server
import json
import os
import sys
import urllib.parse
import openpyxl

EXCEL_PATH = os.environ.get('EXCEL_PATH', '/workspace/maternal_health.xlsx')
PORT = int(os.environ.get('PORT', 8000))
DIRECTORY = os.path.dirname(os.path.abspath(__file__))

def read_excel_data(file_path):
    """Parses maternal_health.xlsx and returns structured JSON."""
    if not os.path.exists(file_path):
        return {"error": f"File not found: {file_path}", "sheets": {}}

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets_data = {}

        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                sheets_data[sheetname] = {"headers": [], "rows": []}
                continue

            raw_headers = rows[0]
            headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]
            
            data_rows = []
            for row in rows[1:]:
                if not any(row):
                    continue
                row_dict = {}
                for i, h in enumerate(headers):
                    val = row[i] if i < len(row) else None
                    row_dict[h] = val
                data_rows.append(row_dict)

            sheets_data[sheetname] = {
                "headers": headers,
                "rows": data_rows,
                "total_rows": len(data_rows)
            }

        return {
            "success": True,
            "filename": os.path.basename(file_path),
            "filepath": file_path,
            "sheet_names": wb.sheetnames,
            "sheets": sheets_data
        }
    except Exception as e:
        return {"success": False, "error": str(e)}

class DashboardRequestHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=DIRECTORY, **kwargs)

    def end_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.end_headers()

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        
        if parsed.path == '/api/data':
            data = read_excel_data(EXCEL_PATH)
            response_bytes = json.dumps(data).encode('utf-8')
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', str(len(response_bytes)))
            self.end_headers()
            self.wfile.write(response_bytes)
            return

        if parsed.path == '/api/health':
            res = json.dumps({"status": "ok", "file_exists": os.path.exists(EXCEL_PATH)}).encode('utf-8')
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', str(len(res)))
            self.end_headers()
            self.wfile.write(res)
            return

        return super().do_GET()

def run_server():
    server_address = ('', PORT)
    httpd = http.server.ThreadingHTTPServer(server_address, DashboardRequestHandler)
    print(f"==================================================")
    print(f" Maternal Health Analytics Dashboard Server")
    print(f" Serving at: http://localhost:{PORT}")
    print(f" Reading Excel: {EXCEL_PATH}")
    print(f"==================================================")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping server...")
        httpd.server_close()

if __name__ == '__main__':
    run_server()
