"""Generate the downloadable workbook for the current comparison view:
postnatal check within 48 hours coverage, Gaya vs Nalanda, 2022 and 2023,
with the 2022->2023 change in percentage points.

All values are read/recomputed from maternal_health.xlsx — nothing is typed
by hand. A second sheet records the source file, sheets, formula, period and
caveats. The script re-opens the saved file and asserts every value.
"""
import openpyxl
from openpyxl.styles import Font

SRC = 'maternal_health.xlsx'
OUT = 'postnatal_check_Gaya_Nalanda_2022_2023.xlsx'
INDICATOR = 'Postnatal check within 48 hours coverage'
NUMCOL = 'postnatal_check_48h'
DENS = 'pregnancies_registered'
DISTRICTS = ['Gaya', 'Nalanda']
Y0, Y1 = 2022, 2023

src = openpyxl.load_workbook(SRC, data_only=True)
dd = list(src['District Data'].iter_rows(values_only=True))
header = list(dd[0])
rows = [dict(zip(header, r)) for r in dd[1:]]
notes = {n['indicator']: n for n in
         [dict(zip(['indicator', 'definition', 'formula', 'unit', 'source'], r))
          for r in list(src['Indicator Notes'].iter_rows(values_only=True))[1:]]}
ind = notes[INDICATOR]
assert ind['formula'] == f'{NUMCOL} / {DENS} * 100', ind['formula']

def row_of(d, y):
    return next(r for r in rows if r['district'] == d and r['year'] == y)

def cov(r):
    return r[NUMCOL] / r[DENS] * 100

def f1(x):
    return f'{x:.1f}'

r0 = {d: row_of(d, Y0) for d in DISTRICTS}
r1 = {d: row_of(d, Y1) for d in DISTRICTS}
top1 = max(cov(r1[d]) for d in DISTRICTS)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Comparison'
headers = [
    'district',
    f'{Y0} {NUMCOL} (count)', f'{Y0} {DENS} (count)', f'{Y0} coverage (%)',
    f'{Y1} {NUMCOL} (count)', f'{Y1} {DENS} (count)', f'{Y1} coverage (%)',
    f'{Y1} vs top selected district',
    f'change {Y0} to {Y1} (percentage points)', 'change direction',
]
ws.append(headers)
for d in DISTRICTS:
    c0, c1 = cov(r0[d]), cov(r1[d])
    delta = c1 - c0
    gap = top1 - c1
    ws.append([
        d,
        r0[d][NUMCOL], r0[d][DENS], round(c0, 1),
        r1[d][NUMCOL], r1[d][DENS], round(c1, 1),
        'top' if gap < 0.0005 else f'-{f1(gap)} pts',
        f'+{f1(delta)}' if delta > 0.0005 else (f'-{f1(abs(delta))}' if delta < -0.0005 else '0.0'),
        'up' if delta > 0.0005 else ('down' if delta < -0.0005 else 'no change'),
    ])
for c in ws[1]:
    c.font = Font(bold=True)
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
    row[0].number_format = '0.0'
for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
    row[0].number_format = '0.0'
for col, w in zip('ABCDEFGHIJ', (12, 26, 28, 15, 26, 28, 15, 26, 26, 15)):
    ws.column_dimensions[col].width = w

ws2 = wb.create_sheet('Source')
kv = [
    ['indicator', INDICATOR],
    ['definition', ind['definition']],
    ['formula (unmodified, from "Indicator Notes")', ind['formula']],
    ['unit', ind['unit']],
    ['periods', f'{Y0}, {Y1}'],
    ['districts shown', ', '.join(DISTRICTS)],
    ['raw count column', NUMCOL],
    ['denominator column', DENS],
    ['source file', SRC],
    ['source sheet (raw counts)', 'District Data'],
    ['source sheet (definition, formula, unit)', 'Indicator Notes'],
    ['note on hidden rows',
     f'The "District Data" sheet also contains '
     f'{sum(1 for r in rows if r["year"] not in (Y0, Y1))} row(s) for 2021, '
     'which are hidden per your request.'],
    ['data caveat', f'Important (from "Indicator Notes"): {notes["Important"]["definition"]}'],
]
ws2.append(['item', 'value'])
for k, v in kv:
    ws2.append([k, v])
for c in ws2[1]:
    c.font = Font(bold=True)
ws2.column_dimensions['A'].width = 38
ws2.column_dimensions['B'].width = 90

wb.save(OUT)

# ---- re-open the artifact and assert every value ---------------------------
chk = openpyxl.load_workbook(OUT, data_only=True)
c1s = chk['Comparison']
got = [[c for c in row] for row in c1s.iter_rows(values_only=True)]
assert got[0] == headers, got[0]
for i, d in enumerate(DISTRICTS):
    c0, c1 = cov(r0[d]), cov(r1[d])
    gap = top1 - c1
    assert got[1 + i] == [d, r0[d][NUMCOL], r0[d][DENS], round(c0, 1),
                          r1[d][NUMCOL], r1[d][DENS], round(c1, 1),
                          'top' if gap < 0.0005 else f'-{f1(gap)} pts',
                          f'+{f1(c1 - c0)}', 'up'], got[1 + i]
s1 = chk['Source']
srows = {row[0]: row[1] for row in s1.iter_rows(values_only=True) if row[0] != 'item'}
assert srows['formula (unmodified, from "Indicator Notes")'] == ind['formula']
assert srows['definition'] == ind['definition']
assert srows['source file'] == SRC
print(f'OK: wrote {OUT}')
for row in got:
    print(' | '.join(str(v) for v in row))
