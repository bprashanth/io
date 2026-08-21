"""Build index.html from maternal_health.xlsx.

Reads the two sheets, validates structure, and embeds the real rows (JSON)
into template.html. The page recomputes coverage from these raw counts, so
every number shown on the site traces back to this workbook.
"""
import json
import openpyxl

XLSX = "maternal_health.xlsx"
REQUIRED_COLS = ["district", "year", "pregnancies_registered",
                 "institutional_deliveries", "postnatal_check_48h"]

wb = openpyxl.load_workbook(XLSX, data_only=True)
assert wb.sheetnames == ["District Data", "Indicator Notes"], wb.sheetnames


def read_rows(ws):
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() for c in next(it)]
    out = []
    for r in it:
        if all(v is None for v in r):
            continue
        out.append({h: v for h, v in zip(header, r)})
    return header, out


SHOWN_YEARS = [2022, 2023]  # user request: show only 2022 and 2023

dd_header, all_data = read_rows(wb["District Data"])
notes_header, notes = read_rows(wb["Indicator Notes"])
data = [r for r in all_data if int(r["year"]) in SHOWN_YEARS]
hidden_rows = [r for r in all_data if int(r["year"]) not in SHOWN_YEARS]
hidden_note = (f'The sheet also contains {len(hidden_rows)} row(s) for '
               + ' and '.join(sorted({str(int(r["year"])) for r in hidden_rows}))
               + ', which are hidden per your request.') if hidden_rows else ''

missing = [c for c in REQUIRED_COLS if c not in dd_header]
assert not missing, f"District Data missing columns: {missing}"

for r in data:
    assert isinstance(r["district"], str) and r["district"].strip()
    assert isinstance(r["year"], int), r
    for c in REQUIRED_COLS[2:]:
        v = r[c]
        assert isinstance(v, (int, float)) and v >= 0, (r, c)
    assert r["pregnancies_registered"] > 0, r

years = sorted({int(r["year"]) for r in data})
districts = sorted({r["district"] for r in data})
# every district present for every year
for d in districts:
    for y in years:
        assert any(r["district"] == d and r["year"] == y for r in data), (d, y)


def find_note(label):
    for n in notes:
        if n["indicator"] == label:
            return n
    raise SystemExit(f"Indicator Notes missing row: {label}")


inst = find_note("Institutional delivery coverage")
pnat = find_note("Postnatal check within 48 hours coverage")

ind = {
    "institutional": {
        "label": "Institutional delivery coverage",
        "short": "institutional delivery coverage",
        "num": "institutional_deliveries",
        "den": "pregnancies_registered",
        "definition": inst["definition"],
        "formula": inst["formula"],
        "unit": inst["unit"],
    },
    "postnatal48": {
        "label": "Postnatal check within 48 hours coverage",
        "short": "postnatal check within 48 hours coverage",
        "num": "postnatal_check_48h",
        "den": "pregnancies_registered",
        "definition": pnat["definition"],
        "formula": pnat["formula"],
        "unit": pnat["unit"],
    },
}

html = open("template.html", encoding="utf-8").read()
html = (html
        .replace("__DATA__", json.dumps(data, ensure_ascii=False))
        .replace("__NOTES__", json.dumps(notes, ensure_ascii=False))
        .replace("__IND__", json.dumps(ind, ensure_ascii=False))
        .replace("__YEARS__", json.dumps(years))
        .replace("__RAWWNOTE__", hidden_note))
assert "__" not in html.replace("__DSH", ""), "unreplaced placeholder"

with open("index.html", "w", encoding="utf-8") as f:
    f.write(html)
print(f"OK: {len(data)} rows shown (of {len(all_data)} in file), "
      f"districts={districts}, years={years}")
print("formulas:", ind["institutional"]["formula"], "|", ind["postnatal48"]["formula"])
