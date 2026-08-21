"""Cross-check index.html: (1) embedded JSON == xlsx, (2) rendered HTML from
the page's own JS == independently recomputed values, (3) default view is the
requested state: postnatal check, Gaya vs Nalanda, 2022->2023 change."""
import json
import re
import subprocess
import sys

import openpyxl

SHOWN_YEARS = {2022, 2023}
Y0, Y1 = 2022, 2023

html = open('index.html', encoding='utf-8').read()
wb = openpyxl.load_workbook('maternal_health.xlsx', data_only=True)

def embedded(name):
    m = re.search(rf'const {name} = (\[.*?\]);\n', html, re.S)
    return json.loads(m.group(1))

data = embedded('DATA')
notes = embedded('NOTES')

ws = wb['District Data']
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
xlsx_all = [dict(zip(header, r)) for r in rows[1:]]
xlsx_rows = [r for r in xlsx_all if int(r['year']) in SHOWN_YEARS]
assert data == xlsx_rows, 'embedded DATA differs from xlsx (filtered)'
assert len(data) == 6 and len(xlsx_all) == 9
ws = wb['Indicator Notes']
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
xlsx_notes = [dict(zip(header, r)) for r in rows[1:]]
assert notes == xlsx_notes, 'embedded NOTES differs from xlsx'
print('1) embedded DATA + NOTES identical to xlsx: PASS')

def run_verify(search=''):
    out = subprocess.run(['node', 'verify_page.js', search],
                         capture_output=True, text=True, check=True)
    return json.loads(out.stdout)

def cov(r, col):
    return r[col] / r['pregnancies_registered'] * 100

def f1(x):
    return f'{x:.1f}'

NUMCOL = {'institutional': 'institutional_deliveries', 'postnatal48': 'postnatal_check_48h'}
LABEL = {'institutional': 'Institutional delivery coverage',
         'postnatal48': 'Postnatal check within 48 hours coverage'}
fail = []
def check(cond, msg):
    if not cond:
        fail.append(msg)

# ---------- default load: the requested state ----------
rep = run_verify('')
st = rep['state']
check(st == {'indicator': 'postnatal48', 'year': 2023, 'districts': ['Gaya', 'Nalanda']},
      f'default state wrong: {st}')
check(rep['controls']['indicator'] == 'postnatal48', 'default indicator select wrong')
check(rep['controls']['year'] == '2023', 'default year select wrong')
chips = {c['v']: c['checked'] for c in rep['controls']['chips']}
check(chips == {'Gaya': True, 'Nalanda': True, 'Purnia': False}, f'chips wrong: {chips}')
check(rep['init']['indTitle'] == LABEL['postnatal48'], f'indTitle: {rep["init"]["indTitle"]!r}')
check('Change from 2022 to 2023' in rep['init']['chgTitle'], f'chgTitle: {rep["init"]["chgTitle"]!r}')

# main table, default view (postnatal, 2023, Gaya+Nalanda)
cmp_body = rep['init']['cmpBody']
check('Purnia' not in cmp_body, 'Purnia shown although not selected')
vals = {}
for d in ('Gaya', 'Nalanda'):
    r = next(x for x in xlsx_rows if x['district'] == d and x['year'] == 2023)
    c = cov(r, 'postnatal_check_48h')
    vals[d] = c
    check(f'<td>{d}</td><td class="num">{r["postnatal_check_48h"]}</td><td class="num">{r["pregnancies_registered"]}</td>' in cmp_body,
          f'{d}: raw counts wrong in default table')
    check(f'>{f1(c)}%<' in cmp_body, f'{d}: coverage {f1(c)}% not rendered')
top_d = max(vals, key=vals.get)
gap_d = min(vals, key=vals.get)
check(f'class="top"' in cmp_body and f'>\u2212{f1(vals[top_d] - vals[gap_d])} pts<' in cmp_body,
      f'top/gap marking wrong (top={top_d})')
ranked = sorted(vals.items(), key=lambda x: -x[1])
expected_summary = (f'In 2023, {ranked[0][0]} had the highest postnatal check within 48 hours coverage '
                    f'at {f1(ranked[0][1])}%, ahead of '
                    + ' and '.join(f'{d} ({f1(c)}%)' for d, c in ranked[1:]) + '.')
check(rep['init']['summary'] == expected_summary,
      f'summary mismatch\n got: {rep["init"]["summary"]}\n want: {expected_summary}')

# change table, default view
chg = rep['init']['chgBody']
for d in ('Gaya', 'Nalanda'):
    r0 = next(x for x in xlsx_rows if x['district'] == d and x['year'] == Y0)
    r1 = next(x for x in xlsx_rows if x['district'] == d and x['year'] == Y1)
    c0, c1 = cov(r0, 'postnatal_check_48h'), cov(r1, 'postnatal_check_48h')
    delta = c1 - c0
    dir_ = 'up' if delta > 0 else ('down' if delta < 0 else 'no change')
    sign = '+' if delta > 0 else ('\u2212' if delta < 0 else '')
    row = (f'<td>{d}</td><td class="num">{f1(c0)}%</td><td class="num">{f1(c1)}%</td>'
           f'<td class="num">{dir_} {sign}{f1(abs(delta))} pts</td>')
    check(row in chg, f'{d}: change row wrong\n want: {row}\n got: {chg}')
check('Purnia' not in chg, 'Purnia shown in change table although not selected')

# ---------- current comparison CSV (what the Download button exports) ----------
def cell(s):
    s = str(s)
    return '"' + s.replace('"', '""') + '"' if any(ch in s for ch in ',"\n') else s

ind_num = 'postnatal_check_48h'
head = ['district', f'{Y0} {ind_num} (count)', f'{Y0} pregnancies_registered (count)', f'{Y0} coverage (%)',
        f'{Y1} {ind_num} (count)', f'{Y1} pregnancies_registered (count)', f'{Y1} coverage (%)',
        f'{Y1} vs top selected district', f'change {Y0} to {Y1} (percentage points)', 'change direction']
exp = [','.join(cell(h) for h in head)]
vals1 = {d: cov(next(x for x in xlsx_rows if x['district'] == d and x['year'] == Y1), ind_num)
         for d in ('Gaya', 'Nalanda')}
topv = max(vals1.values())
for d in ('Gaya', 'Nalanda'):
    r0 = next(x for x in xlsx_rows if x['district'] == d and x['year'] == Y0)
    r1 = next(x for x in xlsx_rows if x['district'] == d and x['year'] == Y1)
    c0, c1 = cov(r0, ind_num), cov(r1, ind_num)
    gap = topv - c1
    exp.append(','.join(cell(v) for v in [
        d, r0[ind_num], r0['pregnancies_registered'], f1(c0),
        r1[ind_num], r1['pregnancies_registered'], f1(c1),
        'top' if gap < 0.0005 else f'-{f1(gap)} pts', f'+{f1(c1 - c0)}', 'up']))
exp.append(','.join(cell(v) for v in ['formula (from "Indicator Notes" sheet)',
                                      'postnatal_check_48h / pregnancies_registered * 100'] + [''] * 8))
exp.append(','.join(cell(v) for v in ['source',
                                      'maternal_health.xlsx - "District Data" sheet (raw counts), '
                                      '"Indicator Notes" sheet (definition and formula)'] + [''] * 8))
expected_csv = '\n'.join(exp)
check(rep['csv'] == expected_csv, f'default CSV mismatch:\n got: {rep["csv"]!r}\n want: {expected_csv!r}')
check(rep['csvName'] == 'comparison_postnatal48_2022_2023_Gaya_Nalanda.csv', f'csvName: {rep["csvName"]}')

# ---------- parameterized load: all 3 districts, institutional, 2022 ----------
rep2 = run_verify('?indicator=institutional&year=2022&districts=Gaya,Nalanda,Purnia')
check(rep2['state'] == {'indicator': 'institutional', 'year': 2022,
                        'districts': ['Gaya', 'Nalanda', 'Purnia']},
      f'parameterized state wrong: {rep2["state"]}')
chips2 = {c['v']: c['checked'] for c in rep2['controls']['chips']}
check(all(chips2.values()) and set(chips2) == {'Gaya', 'Nalanda', 'Purnia'}, f'chips2: {chips2}')
t = rep2['defaults']['institutional_2023_all']['table']
vals2 = {}
for r in xlsx_rows:
    if r['year'] != 2023:
        continue
    c = cov(r, 'institutional_deliveries')
    vals2[r['district']] = c
    check(f'<td>{r["district"]}</td><td class="num">{r["institutional_deliveries"]}</td>'
          f'<td class="num">{r["pregnancies_registered"]}</td>' in t,
          f'{r["district"]} 2023 institutional: raw counts wrong')
    check(f'>{f1(c)}%<' in t, f'{r["district"]} 2023 institutional: coverage wrong')
ch2 = rep2['defaults']['institutional_2023_all']['change']
for d, c0v, c1v in [(d,
                     cov(next(x for x in xlsx_rows if x['district'] == d and x['year'] == Y0), 'institutional_deliveries'),
                     cov(next(x for x in xlsx_rows if x['district'] == d and x['year'] == Y1), 'institutional_deliveries'))
                    for d in ('Gaya', 'Nalanda', 'Purnia')]:
    delta = c1v - c0v
    dir_ = 'up' if delta > 0 else 'down'
    row = f'<td>{d}</td><td class="num">{f1(c0v)}%</td><td class="num">{f1(c1v)}%</td><td class="num">{dir_} +{f1(abs(delta))} pts</td>'
    check(row in ch2, f'institutional change row for {d} wrong\n want: {row}')
check('Select at least one district' in rep['defaults']['postnatal48_2023_none']['table'],
      'empty-district message missing')
check('Select at least one district' in rep['defaults']['postnatal48_2023_none']['change'],
      'empty-district message missing in change table')

# ---------- static content ----------
check('3 row(s) for 2021, which are hidden per your request' in html, 'raw-note missing')
check('simple difference between the two coverage rates' in html, 'percentage-point explanation missing')
for n in xlsx_notes:
    check(f'<td>{n["indicator"]}</td><td>{n["definition"]}</td>' in rep['init']['notesBody'],
          f'notes row missing: {n["indicator"]}')
for r in xlsx_rows:
    frag = (f'<td>{r["district"]}</td><td class="num">{r["year"]}</td>'
            f'<td class="num">{r["pregnancies_registered"]}</td>'
            f'<td class="num">{r["institutional_deliveries"]}</td>'
            f'<td class="num">{r["postnatal_check_48h"]}</td>')
    check(frag in rep['init']['rawBody'], f'raw row missing: {r["district"]} {r["year"]}')
check('2021' not in rep['init']['rawBody'], '2021 rows still shown in raw table')
check('Illustrative benchmark data' in rep['init']['caveat'], 'caveat missing')

if fail:
    print('FAILURES:')
    for f in fail:
        print(' -', f)
    sys.exit(1)
print('2) default view is the requested state (postnatal, Gaya vs Nalanda, 2022->2023 change): PASS')
print('3) all rendered values match independent recomputation (incl. 3-district and empty selections): PASS')
print('4) source sheets, formulas, raw rows and caveat rendered: PASS')
