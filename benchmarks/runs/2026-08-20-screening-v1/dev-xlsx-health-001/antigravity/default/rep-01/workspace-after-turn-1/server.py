#!/usr/bin/env python3
"""
Maternal Health Dashboard Server
Reads Excel sheets from maternal_health.xlsx and serves indicator calculations,
district comparisons, raw sheet views, and formula breakdowns.
"""

import http.server
import json
import os
import sys
import urllib.parse
from http import HTTPStatus
import openpyxl

EXCEL_PATH = os.environ.get('EXCEL_PATH', '/workspace/maternal_health.xlsx')
if not os.path.exists(EXCEL_PATH):
    EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'maternal_health.xlsx')

PORT = int(os.environ.get('PORT', 8080))
STATIC_DIR = os.path.dirname(os.path.abspath(__file__))


def load_workbook_data(filepath):
    if not os.path.exists(filepath):
        return None
    wb = openpyxl.load_workbook(filepath, data_only=False)
    result = {
        'sheets': {},
        'indicators': [],
        'district_data': [],
        'years': [],
        'districts': []
    }
    
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            if any(v is not None for v in row):
                rows.append(list(row))
        result['sheets'][sheetname] = rows

    # Parse Indicator Notes sheet
    if 'Indicator Notes' in result['sheets']:
        sheet_rows = result['sheets']['Indicator Notes']
        if len(sheet_rows) > 1:
            headers = [str(h).strip().lower() if h is not None else '' for h in sheet_rows[0]]
            for row in sheet_rows[1:]:
                item = {}
                for idx, val in enumerate(row):
                    if idx < len(headers):
                        item[headers[idx]] = val
                if item.get('indicator') and item.get('indicator') != 'Important':
                    result['indicators'].append(item)
                elif item.get('indicator') == 'Important':
                    result['disclaimer'] = item.get('definition')

    # Parse District Data sheet
    if 'District Data' in result['sheets']:
        sheet_rows = result['sheets']['District Data']
        if len(sheet_rows) > 1:
            headers = [str(h).strip().lower() if h is not None else '' for h in sheet_rows[0]]
            for row in sheet_rows[1:]:
                item = {}
                for idx, val in enumerate(row):
                    if idx < len(headers):
                        item[headers[idx]] = val
                if item.get('district') and item.get('year'):
                    try:
                        item['year'] = int(item['year'])
                    except (ValueError, TypeError):
                        pass
                    result['district_data'].append(item)

    years = sorted(list(set(d['year'] for d in result['district_data'] if 'year' in d and isinstance(d['year'], int))))
    districts = sorted(list(set(d['district'] for d in result['district_data'] if 'district' in d and d['district'])))
    result['years'] = years
    result['districts'] = districts
    return result


def compute_comparisons(data, indicator_key, year=None, selected_districts=None):
    """
    Computes coverage rates for the specified indicator and year.
    indicator_key: 'institutional_delivery' or 'postnatal_check'
    """
    numerator_field = 'institutional_deliveries'
    formula_str = 'institutional_deliveries / pregnancies_registered * 100'
    indicator_name = 'Institutional delivery coverage'
    definition = 'Share of registered pregnancies with an institutional delivery'

    if 'postnatal' in indicator_key.lower():
        numerator_field = 'postnatal_check_48h'
        formula_str = 'postnatal_check_48h / pregnancies_registered * 100'
        indicator_name = 'Postnatal check within 48 hours coverage'
        definition = 'Share of registered pregnancies with a recorded postnatal check within 48 hours'
    else:
        for ind in data.get('indicators', []):
            if 'institutional' in ind.get('indicator', '').lower():
                formula_str = ind.get('formula', formula_str)
                indicator_name = ind.get('indicator', indicator_name)
                definition = ind.get('definition', definition)

    denominator_field = 'pregnancies_registered'

    rows = data.get('district_data', [])
    if year is not None and str(year).lower() != 'all':
        try:
            year_int = int(year)
            rows = [r for r in rows if r.get('year') == year_int]
        except (ValueError, TypeError):
            pass

    if selected_districts:
        rows = [r for r in rows if r.get('district') in selected_districts]

    results = []
    for r in rows:
        district = r.get('district')
        y = r.get('year')
        num = r.get(numerator_field)
        den = r.get(denominator_field)

        coverage = None
        calc_str = 'N/A'
        if num is not None and den is not None and den > 0:
            coverage = round((float(num) / float(den)) * 100.0, 2)
            calc_str = f"{num:,} / {den:,} * 100 = {coverage:.2f}%"

        results.append({
            'district': district,
            'year': y,
            'numerator_field': numerator_field,
            'numerator_value': num,
            'denominator_field': denominator_field,
            'denominator_value': den,
            'coverage_percent': coverage,
            'calculation_string': calc_str
        })

    if year is not None and str(year).lower() != 'all':
        results.sort(key=lambda x: (x['coverage_percent'] is None, -(x['coverage_percent'] or 0)))

    total_num = sum(r.get(numerator_field, 0) or 0 for r in rows)
    total_den = sum(r.get(denominator_field, 0) or 0 for r in rows)
    overall_coverage = round((total_num / total_den * 100.0), 2) if total_den > 0 else 0.0

    return {
        'indicator_name': indicator_name,
        'definition': definition,
        'formula': formula_str,
        'numerator_field': numerator_field,
        'denominator_field': denominator_field,
        'unit': 'percent (%)',
        'source_sheets': {
            'data_sheet': 'District Data',
            'metadata_sheet': 'Indicator Notes',
            'workbook': os.path.basename(EXCEL_PATH)
        },
        'selected_year': year,
        'summary': {
            'total_numerator': total_num,
            'total_denominator': total_den,
            'overall_coverage': overall_coverage,
            'district_count': len(set(r.get('district') for r in rows))
        },
        'districts_comparison': results
    }


class DashboardHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=STATIC_DIR, **kwargs)

    def end_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(HTTPStatus.OK)
        self.end_headers()

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        params = urllib.parse.parse_qs(parsed.query)

        if path == '/api/data':
            data = load_workbook_data(EXCEL_PATH)
            if data is None:
                self.send_json_response({'error': f'Excel file not found at {EXCEL_PATH}'}, HTTPStatus.NOT_FOUND)
                return
            self.send_json_response(data)
            return

        elif path == '/api/compare':
            data = load_workbook_data(EXCEL_PATH)
            if data is None:
                self.send_json_response({'error': 'Excel file not found'}, HTTPStatus.NOT_FOUND)
                return
            ind = params.get('indicator', ['institutional'])[0]
            year = params.get('year', ['2023'])[0]
            districts = params.get('districts', None)
            if districts:
                districts = districts[0].split(',')
            result = compute_comparisons(data, ind, year, districts)
            self.send_json_response(result)
            return

        elif path == '/api/sheets':
            data = load_workbook_data(EXCEL_PATH)
            if data is None:
                self.send_json_response({'error': 'Excel file not found'}, HTTPStatus.NOT_FOUND)
                return
            self.send_json_response({'sheets': data.get('sheets', {})})
            return

        return super().do_GET()

    def send_json_response(self, obj, status=HTTPStatus.OK):
        content = json.dumps(obj, indent=2).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(content)))
        self.end_headers()
        self.wfile.write(content)


def run_server():
    server_address = ('', PORT)
    httpd = http.server.HTTPServer(server_address, DashboardHandler)
    print(f"Serving Maternal Health Dashboard at http://localhost:{PORT}")
    print(f"Loading Excel file from: {EXCEL_PATH}")
    httpd.serve_forever()


if __name__ == '__main__':
    run_server()
