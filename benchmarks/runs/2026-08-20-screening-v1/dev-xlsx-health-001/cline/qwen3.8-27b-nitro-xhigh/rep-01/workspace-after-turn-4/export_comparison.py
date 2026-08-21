#!/usr/bin/env python3
"""Export the current comparison table to postnatal_check_gaya_nalanda_2022_2023.xlsx.

Current comparison: 'Postnatal check within 48 hours coverage',
Gaya vs Nalanda, change from 2022 to 2023 in percentage points.

Usage:
    python3 export_comparison.py
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).resolve().parent
WORKBOOK = BASE / "maternal_health.xlsx"
OUT = BASE / "postnatal_check_gaya_nalanda_2022_2023.xlsx"

INDICATOR = "Postnatal check within 48 hours coverage"
DISTRICTS = ["Gaya", "Nalanda"]
YEARS = [2022, 2023]


def read_source():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["District Data"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header = [str(h).strip() for h in rows[0]]
    i_d, i_y = header.index("district"), header.index("year")
    data = {}
    for raw in rows[1:]:
        if raw[i_d] is None:
            continue
        data[(str(raw[i_d]).strip(), int(raw[i_y]))] = {
            col: raw[i] for i, col in enumerate(header)
        }

    ws2 = wb["Indicator Notes"]
    n_rows = [list(r) for r in ws2.iter_rows(values_only=True)]
    n_header = [str(h).strip() for h in n_rows[0]]
    meta = {}
    for raw in n_rows[1:]:
        if raw[0] is not None:
            meta[str(raw[0]).strip()] = {
                col: ("" if raw[i] is None else raw[i])
                for i, col in enumerate(n_header)
            }
    return data, meta


def main():
    data, meta = read_source()
    ind = meta[INDICATOR]
    formula = str(ind["formula"]).strip()
    num, den = (p.strip() for p in formula.split("*")[0].split("/"))
    note = meta.get("Important", {}).get("definition", "")

    out = openpyxl.Workbook()
    sh = out.active
    sh.title = "Comparison"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF7")

    sh["A1"] = f"{INDICATOR} - {DISTRICTS[0]} vs {DISTRICTS[1]} ({YEARS[0]}-{YEARS[1]})"
    sh["A1"].font = Font(bold=True, size=14)
    sh["A2"] = f"Formula: {formula}   |   Source: {WORKBOOK.name}, sheet 'District Data'"
    sh["A2"].font = Font(color="6B7280")

    # main comparison table
    head = 4
    cols = ["district", f"{YEARS[0]}", f"{YEARS[1]}",
            f"change {YEARS[0]} to {YEARS[1]} (percentage points)"]
    for j, c in enumerate(cols, start=1):
        cell = sh.cell(row=head, column=j, value=c)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, d in enumerate(DISTRICTS):
        r0 = data[(d, YEARS[0])][den]
        r1 = data[(d, YEARS[1])][den]
        n0, n1 = data[(d, YEARS[0])][num], data[(d, YEARS[1])][num]
        p0, p1 = n0 / r0, n1 / r1
        row = head + 1 + i
        sh.cell(row=row, column=1, value=d).font = bold
        for col, p in enumerate((p0, p1), start=2):
            cell = sh.cell(row=row, column=col, value=p)
            cell.number_format = "0.0%"
        delta = sh.cell(row=row, column=4, value=round((p1 - p0) * 100, 1))
        delta.number_format = "+0.0;-0.0;0.0"

    # raw counts table
    raw_head = head + 1 + len(DISTRICTS) + 2
    raw_cols = ["district", "year", den, num, "coverage"]
    for j, c in enumerate(raw_cols, start=1):
        cell = sh.cell(row=raw_head, column=j, value=c)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    r = raw_head + 1
    for d in DISTRICTS:
        for y in YEARS:
            sh.cell(row=r, column=1, value=d)
            sh.cell(row=r, column=2, value=y)
            sh.cell(row=r, column=3, value=int(data[(d, y)][den]))
            sh.cell(row=r, column=4, value=int(data[(d, y)][num]))
            cell = sh.cell(row=r, column=5,
                           value=int(data[(d, y)][num]) / int(data[(d, y)][den]))
            cell.number_format = "0.0%"
            r += 1

    note_row = r + 1
    sh.cell(row=note_row, column=1,
            value=f"Important: {note}").font = Font(italic=True, color="92400E")

    for col, w in zip("ABCDE", (12, 12, 22, 20, 42)):
        sh.column_dimensions[col].width = w

    out.save(OUT)
    print(f"Wrote {OUT.relative_to(BASE)}")


if __name__ == "__main__":
    main()
