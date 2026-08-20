#!/usr/bin/env python3
"""Build site/index.html from maternal_health.xlsx.

Reads the 'District Data' sheet (counts per district/year) and the
'Indicator Notes' sheet (definitions and formulas), then injects the data
into template.html and writes a self-contained static page.

Usage:
    python3 build.py
"""
import json
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
WORKBOOK = BASE / "maternal_health.xlsx"
TEMPLATE = BASE / "template.html"
OUT = BASE / "site" / "index.html"


def read_district_data(wb):
    ws = wb["District Data"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header = [str(h).strip() for h in rows[0]]
    records = []
    for raw in rows[1:]:
        if raw[0] is None:
            continue
        rec = {col: raw[i] for i, col in enumerate(header)}
        out = {"district": str(rec["district"]).strip(), "year": int(rec["year"])}
        for col in header:
            if col not in ("district", "year"):
                out[col] = int(rec[col])
        records.append(out)
    return header, records


def read_indicators(wb):
    ws = wb["Indicator Notes"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header = [str(h).strip() for h in rows[0]]
    indicators, notes = [], []
    for raw in rows[1:]:
        if raw[0] is None:
            continue
        row = {col: ("" if raw[i] is None else raw[i]) for i, col in enumerate(header)}
        formula = str(row.get("formula", "")).strip()
        if formula:
            num, den = (p.strip() for p in formula.split("*")[0].split("/"))
            indicators.append(
                {
                    "name": str(row["indicator"]).strip(),
                    "definition": str(row.get("definition", "")).strip(),
                    "formula": formula,
                    "unit": str(row.get("unit", "")).strip(),
                    "source": str(row.get("source", "")).strip(),
                    "numerator": num,
                    "denominator": den,
                }
            )
        else:
            notes.append(str(row.get("definition", "")).strip())
    return indicators, notes


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    data_header, records = read_district_data(wb)
    indicators, notes = read_indicators(wb)
    if not indicators:
        raise SystemExit("No indicators found in 'Indicator Notes' sheet")

    years = sorted({r["year"] for r in records})
    districts = []
    for r in records:
        if r["district"] not in districts:
            districts.append(r["district"])

    data = {
        "workbook": WORKBOOK.name,
        "sourceSheet": "District Data",
        "dataColumns": data_header,
        "records": records,
        "indicators": indicators,
        "years": years,
        "districts": districts,
        "notes": notes,
    }

    html = TEMPLATE.read_text(encoding="utf-8").replace("__DATA__", json.dumps(data, indent=2))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(html, encoding="utf-8")
    print(f"Wrote {OUT.relative_to(BASE)} ({OUT.stat().st_size:,} bytes)")
    print(f"  districts : {', '.join(districts)}")
    print(f"  years     : {years}")
    print(f"  indicators: {', '.join(i['name'] for i in indicators)}")


if __name__ == "__main__":
    main()
