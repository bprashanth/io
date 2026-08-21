import openpyxl
import json
import os

def parse_excel_workbook(file_path):
    """
    Parses the school attendance nested workbook into structured JSON.
    Returns:
        dict containing:
        - 'sheets': list of sheet names
        - 'primary_attendance': structured primary school attendance list
        - 'secondary_attendance': structured secondary school attendance list
        - 'readme': dict of readme metadata
        - 'enrolment_raw': list of raw enrolment records
        - 'enrolment_summary': aggregation by block
        - 'raw_sheets': full grid data for every sheet with merge info
        - 'years': list of available years
        - 'blocks': list of blocks
        - 'indicators': list of available indicators
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    result = {
        'filename': os.path.basename(file_path),
        'sheets': wb.sheetnames,
        'raw_sheets': {},
        'primary_attendance': [],
        'secondary_attendance': [],
        'readme': {},
        'enrolment_raw': [],
        'enrolment_summary': {},
        'years': ['2022', '2023'],
        'blocks': [],
        'indicators': [
            {'id': 'all', 'label': 'All Indicators (Boys & Girls)'},
            {'id': 'boys', 'label': 'Boys Attendance (%)'},
            {'id': 'girls', 'label': 'Girls Attendance (%)'},
            {'id': 'both_compare', 'label': 'Boys vs Girls Comparison'},
            {'id': 'avg', 'label': 'Average Attendance ((Boys + Girls) / 2)'},
            {'id': 'gap', 'label': 'Gender Gap (Boys - Girls %)'}
        ]
    }

    # 1. Capture Raw Sheet Grids with merged ranges
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        grid = []
        for r in range(1, sheet.max_row + 1):
            row = []
            for c in range(1, sheet.max_column + 1):
                val = sheet.cell(r, c).value
                row.append(val)
            grid.append(row)

        merged_cells = []
        for rng in sheet.merged_cells.ranges:
            merged_cells.append({
                'range': str(rng),
                'min_row': rng.min_row,
                'max_row': rng.max_row,
                'min_col': rng.min_col,
                'max_col': rng.max_col
            })

        result['raw_sheets'][sheet_name] = {
            'name': sheet_name,
            'max_row': sheet.max_row,
            'max_col': sheet.max_column,
            'grid': grid,
            'merged_cells': merged_cells
        }

    # 2. Parse Attendance Report Sheet
    if 'Attendance Report' in wb.sheetnames:
        att_sheet = wb['Attendance Report']
        
        # Primary school table is located in Rows 3-7
        # Row 3: Headers ['Block', 'Boys attendance (%)', None, 'Girls attendance (%)', None]
        # Row 4: Years [None, '2022', '2023', '2022', '2023']
        # Rows 5-7: Block rows
        primary_blocks = []
        for r in range(5, 8):
            block_name = att_sheet.cell(r, 1).value
            if block_name:
                boys_2022 = att_sheet.cell(r, 2).value or 0
                boys_2023 = att_sheet.cell(r, 3).value or 0
                girls_2022 = att_sheet.cell(r, 4).value or 0
                girls_2023 = att_sheet.cell(r, 5).value or 0
                
                primary_blocks.append({
                    'block': block_name,
                    'attendance': {
                        '2022': {
                            'boys': boys_2022,
                            'girls': girls_2022,
                            'avg': round((boys_2022 + girls_2022) / 2, 1),
                            'gap': round(boys_2022 - girls_2022, 1)
                        },
                        '2023': {
                            'boys': boys_2023,
                            'girls': girls_2023,
                            'avg': round((boys_2023 + girls_2023) / 2, 1),
                            'gap': round(boys_2023 - girls_2023, 1)
                        }
                    },
                    'yoy': {
                        'boys': round(boys_2023 - boys_2022, 1),
                        'girls': round(girls_2023 - girls_2022, 1),
                        'avg': round(((boys_2023 + girls_2023) / 2) - ((boys_2022 + girls_2022) / 2), 1)
                    }
                })
        result['primary_attendance'] = primary_blocks
        result['blocks'] = [item['block'] for item in primary_blocks]

        # Secondary school table is located in Rows 13-17
        secondary_blocks = []
        for r in range(15, 18):
            block_name = att_sheet.cell(r, 1).value
            if block_name:
                boys_2022 = att_sheet.cell(r, 2).value or 0
                boys_2023 = att_sheet.cell(r, 3).value or 0
                girls_2022 = att_sheet.cell(r, 4).value or 0
                girls_2023 = att_sheet.cell(r, 5).value or 0
                
                secondary_blocks.append({
                    'block': block_name,
                    'attendance': {
                        '2022': {
                            'boys': boys_2022,
                            'girls': girls_2022,
                            'avg': round((boys_2022 + girls_2022) / 2, 1),
                            'gap': round(boys_2022 - girls_2022, 1)
                        },
                        '2023': {
                            'boys': boys_2023,
                            'girls': girls_2023,
                            'avg': round((boys_2023 + girls_2023) / 2, 1),
                            'gap': round(boys_2023 - girls_2023, 1)
                        }
                    },
                    'yoy': {
                        'boys': round(boys_2023 - boys_2022, 1),
                        'girls': round(girls_2023 - girls_2022, 1),
                        'avg': round(((boys_2023 + girls_2023) / 2) - ((boys_2022 + girls_2022) / 2), 1)
                    }
                })
        result['secondary_attendance'] = secondary_blocks

    # 3. Parse Read Me Sheet
    if 'Read Me' in wb.sheetnames:
        readme_sheet = wb['Read Me']
        readme_dict = {}
        for r in range(1, readme_sheet.max_row + 1):
            key = readme_sheet.cell(r, 1).value
            val = readme_sheet.cell(r, 2).value
            if key:
                readme_dict[str(key).strip()] = str(val).strip() if val is not None else ''
        result['readme'] = readme_dict

    # 4. Parse Enrolment Raw Sheet
    if 'Enrolment Raw' in wb.sheetnames:
        enr_sheet = wb['Enrolment Raw']
        headers = [enr_sheet.cell(1, c).value for c in range(1, enr_sheet.max_column + 1)]
        records = []
        block_summary = {}
        
        for r in range(2, enr_sheet.max_row + 1):
            school_code = enr_sheet.cell(r, 1).value
            block = enr_sheet.cell(r, 2).value
            year = enr_sheet.cell(r, 3).value
            enrolled = enr_sheet.cell(r, 4).value
            if school_code is not None:
                rec = {
                    'school_code': school_code,
                    'block': block,
                    'year': year,
                    'enrolled': enrolled
                }
                records.append(rec)
                
                if block:
                    if block not in block_summary:
                        block_summary[block] = {
                            'school_count': 0,
                            'total_enrolled': 0,
                            'min_enrolled': float('inf'),
                            'max_enrolled': 0
                        }
                    block_summary[block]['school_count'] += 1
                    if enrolled is not None:
                        block_summary[block]['total_enrolled'] += enrolled
                        block_summary[block]['min_enrolled'] = min(block_summary[block]['min_enrolled'], enrolled)
                        block_summary[block]['max_enrolled'] = max(block_summary[block]['max_enrolled'], enrolled)

        # Average enrolled per block
        for b, stats in block_summary.items():
            if stats['school_count'] > 0:
                stats['avg_enrolled'] = round(stats['total_enrolled'] / stats['school_count'], 1)
            if stats['min_enrolled'] == float('inf'):
                stats['min_enrolled'] = 0

        result['enrolment_raw'] = records
        result['enrolment_summary'] = block_summary

    return result

if __name__ == '__main__':
    data = parse_excel_workbook('/workspace/school_attendance_nested.xlsx')
    print("Parsed data keys:", list(data.keys()))
    print("Primary Attendance count:", len(data['primary_attendance']))
    print("Blocks:", data['blocks'])
